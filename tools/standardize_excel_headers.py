from __future__ import annotations

import argparse
from calendar import monthrange
import json
import re
from dataclasses import dataclass
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from tools.csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message
except ModuleNotFoundError:
    from csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message


DEFAULT_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "header-standardizer.json"
COMMENT_DATE_AND_PRODUCT_HEADER = "评论日期与产品"
COMMENT_DATE_HEADER = "评论日期"
PRODUCT_NAME_HEADER = "产品名"
TIMESTAMP_HEADER = "timestamp"
BEIJING_TZ = timezone(timedelta(hours=8))
BEIJING_TIMESTAMP_FORMAT = "%Y-%m-%d"
RELATIVE_COUNT_PATTERN = r"\d+|[一二两三四五六七八九十]+"
PLATFORM_DATETIME_HEADERS = (
    "timestamp",
    "createTime",
    "create_time",
    "createdAt",
    "created_at",
    "createDate",
    "create_date",
    "publishedAt",
    "published_at",
    "publishedTime",
    "published_time",
    "published",
    "date",
    "Date",
    "time",
    "Time",
    "commentTime",
    "comment_time",
    "Comment Published",
    "Published At",
)
COMMENT_DATE_AND_PRODUCT_PATTERN = re.compile(
    r"^\s*(?P<date>(?:\d{4}年\d{1,2}月\d{1,2}日|\d{4}[/-]\d{1,2}[/-]\d{1,2}))"
    r"\s*(?:(?:已购|已购买)[:：])?\s*(?P<product>.*)$"
)


class HeaderNotFoundError(ValueError):
    pass


class DuplicateHeaderError(ValueError):
    pass


class OutputPathConflictError(ValueError):
    pass


@dataclass(frozen=True)
class StandardColumn:
    header: str
    aliases: tuple[str, ...]
    required: bool = True


@dataclass(frozen=True)
class HeaderStandardizerConfig:
    header_row: int
    output_columns: tuple[StandardColumn, ...]
    drop_headers: tuple[str, ...]


@dataclass(frozen=True)
class SelectedColumn:
    output_header: str
    source_header: str | None
    source_column: int | None


@dataclass(frozen=True)
class SheetStandardizeSummary:
    sheet_name: str
    data_rows_written: int
    selected_columns: tuple[SelectedColumn, ...]
    omitted_headers: tuple[str, ...]
    configured_drop_headers_found: tuple[str, ...]


@dataclass(frozen=True)
class StandardizeResult:
    input_path: Path
    output_xlsx: Path
    summary_json: Path
    sheets_processed: int
    data_rows_written: int


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split())


def value_for_json(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def split_comment_date_and_product(value: Any) -> tuple[Any, Any]:
    if value is None:
        return None, None

    text = str(value).strip()
    if not text:
        return None, None

    match = COMMENT_DATE_AND_PRODUCT_PATTERN.match(text)
    if not match:
        return value, None

    comment_date = match.group("date").strip()
    product_name = match.group("product").strip()
    return comment_date or None, product_name or None


def convert_unix_timestamp_to_beijing(value: Any) -> Any:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None
    if not re.fullmatch(r"[+-]?\d+(?:\.\d+)?", text):
        return value

    try:
        timestamp = float(text)
        if abs(timestamp) >= 10_000_000_000:
            timestamp = timestamp / 1000
        return datetime.fromtimestamp(timestamp, tz=timezone.utc).astimezone(BEIJING_TZ).strftime(BEIJING_TIMESTAMP_FORMAT)
    except (OSError, OverflowError, ValueError):
        return value


def current_beijing_date() -> date:
    return datetime.now(BEIJING_TZ).date()


def subtract_months(base: date, months: int) -> date:
    total_months = base.year * 12 + (base.month - 1) - months
    year = total_months // 12
    month = total_months % 12 + 1
    day = min(base.day, monthrange(year, month)[1])
    return date(year, month, day)


def parse_relative_count(value: str) -> int | None:
    if re.fullmatch(r"\d+", value):
        return int(value)

    digits = {
        "一": 1,
        "二": 2,
        "两": 2,
        "三": 3,
        "四": 4,
        "五": 5,
        "六": 6,
        "七": 7,
        "八": 8,
        "九": 9,
    }
    if value in digits:
        return digits[value]
    if value == "十":
        return 10
    if "十" in value:
        tens_text, ones_text = value.split("十", 1)
        tens = 1 if not tens_text else digits.get(tens_text)
        ones = 0 if not ones_text else digits.get(ones_text)
        if tens is None or ones is None:
            return None
        return tens * 10 + ones
    return None


def convert_relative_platform_time_to_beijing(value: Any, today: date | None = None) -> Any:
    if value is None:
        return None

    text = str(value).strip()
    if not text:
        return None

    base = today if today else current_beijing_date()
    normalized = re.sub(r"\s+", " ", text).strip()
    lower_text = normalized.lower()
    if normalized in {"今天", "今日"} or lower_text == "today":
        return base.strftime(BEIJING_TIMESTAMP_FORMAT)
    if normalized == "昨天" or lower_text == "yesterday":
        return (base - timedelta(days=1)).strftime(BEIJING_TIMESTAMP_FORMAT)
    if normalized == "前天":
        return (base - timedelta(days=2)).strftime(BEIJING_TIMESTAMP_FORMAT)

    year_match = re.fullmatch(
        rf"(?P<count>{RELATIVE_COUNT_PATTERN})\s*(?:年\s*前|years?\s+ago|yrs?\s+ago)",
        normalized,
        flags=re.IGNORECASE,
    )
    if year_match:
        count = parse_relative_count(year_match.group("count"))
        if count is not None:
            return f"{base.year - count:04d}"

    month_match = re.fullmatch(
        rf"(?P<count>{RELATIVE_COUNT_PATTERN})\s*(?:(?:个|個)?月\s*前|months?\s+ago|mos?\s+ago)",
        normalized,
        flags=re.IGNORECASE,
    )
    if month_match:
        count = parse_relative_count(month_match.group("count"))
        if count is not None:
            converted = subtract_months(base, count)
            return f"{converted.year:04d}-{converted.month:02d}"

    week_match = re.fullmatch(
        rf"(?P<count>{RELATIVE_COUNT_PATTERN})\s*(?:周\s*前|星期\s*前|weeks?\s+ago|wks?\s+ago)",
        normalized,
        flags=re.IGNORECASE,
    )
    if week_match:
        count = parse_relative_count(week_match.group("count"))
        if count is not None:
            return (base - timedelta(days=count * 7)).strftime(BEIJING_TIMESTAMP_FORMAT)

    day_match = re.fullmatch(
        rf"(?P<count>{RELATIVE_COUNT_PATTERN})\s*(?:天\s*前|日\s*前|days?\s+ago)",
        normalized,
        flags=re.IGNORECASE,
    )
    if day_match:
        count = parse_relative_count(day_match.group("count"))
        if count is not None:
            return (base - timedelta(days=count)).strftime(BEIJING_TIMESTAMP_FORMAT)

    return value


def convert_platform_datetime_to_beijing_date(value: Any, today: date | None = None) -> Any:
    if value is None:
        return None

    if isinstance(value, datetime):
        if value.tzinfo:
            return value.astimezone(BEIJING_TZ).strftime(BEIJING_TIMESTAMP_FORMAT)
        return value.strftime(BEIJING_TIMESTAMP_FORMAT)
    if isinstance(value, date):
        return value.strftime(BEIJING_TIMESTAMP_FORMAT)

    converted_timestamp = convert_unix_timestamp_to_beijing(value)
    if converted_timestamp != value:
        return converted_timestamp

    text = str(value).strip()
    if not text:
        return None

    converted_relative_time = convert_relative_platform_time_to_beijing(text, today=today)
    if converted_relative_time != text:
        return converted_relative_time

    iso_text = text.replace("Z", "+00:00")
    try:
        parsed = datetime.fromisoformat(iso_text)
        if parsed.tzinfo:
            parsed = parsed.astimezone(BEIJING_TZ)
        return parsed.strftime(BEIJING_TIMESTAMP_FORMAT)
    except ValueError:
        pass

    for pattern in (
        "%Y-%m-%d %H:%M:%S",
        "%Y/%m/%d %H:%M:%S",
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y年%m月%d日",
    ):
        try:
            return datetime.strptime(text, pattern).strftime(BEIJING_TIMESTAMP_FORMAT)
        except ValueError:
            continue

    return value


def value_for_selected_column(row: tuple[Any, ...], column: SelectedColumn, today: date | None = None) -> Any:
    if column.source_column is None:
        return None
    if column.source_column - 1 >= len(row):
        return None

    raw_value = row[column.source_column - 1]
    source_key = normalize_header(column.source_header)
    output_key = normalize_header(column.output_header)
    if source_key == normalize_header(COMMENT_DATE_AND_PRODUCT_HEADER):
        comment_date, product_name = split_comment_date_and_product(raw_value)
        if output_key == normalize_header(COMMENT_DATE_HEADER):
            return comment_date
        if output_key == normalize_header(PRODUCT_NAME_HEADER):
            return product_name
    platform_datetime_keys = {normalize_header(header) for header in PLATFORM_DATETIME_HEADERS}
    if source_key in platform_datetime_keys and output_key == normalize_header(COMMENT_DATE_HEADER):
        return convert_platform_datetime_to_beijing_date(raw_value, today=today)

    return raw_value


def load_config(path: Path | None = None) -> HeaderStandardizerConfig:
    config_path = path if path else DEFAULT_CONFIG_PATH
    data = json.loads(config_path.read_text(encoding="utf-8"))
    output_columns = tuple(
        StandardColumn(
            header=str(item["header"]),
            aliases=tuple(str(alias) for alias in item.get("aliases", [])),
            required=bool(item.get("required", True)),
        )
        for item in data["output_columns"]
    )
    return HeaderStandardizerConfig(
        header_row=int(data.get("header_row", 1)),
        output_columns=output_columns,
        drop_headers=tuple(str(header) for header in data.get("drop_headers", [])),
    )


def make_output_paths(input_path: Path, output_dir: Path | None) -> tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d")
    parent = output_dir if output_dir else input_path.parent
    stem = f"{timestamp}_{input_path.stem}"
    return (
        parent / f"{stem}.standardized.xlsx",
        parent / f"{stem}.standardized.summary.json",
    )


def build_header_lookup(headers: list[Any]) -> dict[str, list[int]]:
    lookup: dict[str, list[int]] = {}
    for column_index, header in enumerate(headers, start=1):
        key = normalize_header(header)
        if not key:
            continue
        lookup.setdefault(key, []).append(column_index)
    return lookup


def select_columns(headers: list[Any], config: HeaderStandardizerConfig) -> tuple[SelectedColumn, ...]:
    lookup = build_header_lookup(headers)
    available_headers = [str(header) for header in headers if normalize_header(header)]
    selected: list[SelectedColumn] = []

    for output_column in config.output_columns:
        alias_keys = [normalize_header(output_column.header)]
        alias_keys.extend(normalize_header(alias) for alias in output_column.aliases)
        matched_columns: list[int] = []
        for key in dict.fromkeys(alias_keys):
            matched_columns.extend(lookup.get(key, []))

        matched_columns = sorted(set(matched_columns))
        if not matched_columns:
            if not output_column.required:
                selected.append(
                    SelectedColumn(
                        output_header=output_column.header,
                        source_header=None,
                        source_column=None,
                    )
                )
                continue
            raise HeaderNotFoundError(
                f"Required header not found: {output_column.header}. "
                f"Accepted aliases: {list(output_column.aliases)}. "
                f"Available headers: {available_headers}"
            )
        if len(matched_columns) > 1:
            raise DuplicateHeaderError(f"Multiple columns match required header: {output_column.header}")

        source_column = matched_columns[0]
        source_header = headers[source_column - 1]
        selected.append(
            SelectedColumn(
                output_header=output_column.header,
                source_header="" if source_header is None else str(source_header),
                source_column=source_column,
            )
        )

    return tuple(selected)


def standardize_sheet(
    source_sheet: Worksheet,
    output_sheet: Worksheet,
    config: HeaderStandardizerConfig,
    today: date | None = None,
) -> SheetStandardizeSummary:
    header_values = next(
        source_sheet.iter_rows(
            min_row=config.header_row,
            max_row=config.header_row,
            max_col=source_sheet.max_column,
            values_only=True,
        ),
        None,
    )
    if header_values is None:
        raise HeaderNotFoundError(f"Worksheet has no header row: {source_sheet.title}")

    headers = list(header_values)
    selected_columns = select_columns(headers, config)
    selected_column_indexes = {column.source_column for column in selected_columns if column.source_column is not None}

    output_sheet.append([column.output_header for column in selected_columns])
    data_rows_written = 0

    for row in source_sheet.iter_rows(
        min_row=config.header_row + 1,
        max_row=source_sheet.max_row,
        max_col=source_sheet.max_column,
        values_only=True,
    ):
        output_sheet.append([value_for_selected_column(row, column, today=today) for column in selected_columns])
        data_rows_written += 1

    omitted_headers = tuple(
        "" if header is None else str(header)
        for column_index, header in enumerate(headers, start=1)
        if column_index not in selected_column_indexes and normalize_header(header)
    )
    drop_header_keys = {normalize_header(header) for header in config.drop_headers}
    configured_drop_headers_found = tuple(
        header for header in omitted_headers if normalize_header(header) in drop_header_keys
    )

    return SheetStandardizeSummary(
        sheet_name=source_sheet.title,
        data_rows_written=data_rows_written,
        selected_columns=selected_columns,
        omitted_headers=omitted_headers,
        configured_drop_headers_found=configured_drop_headers_found,
    )


def standardize_workbook(
    input_path: Path,
    config: HeaderStandardizerConfig,
    output_dir: Path | None = None,
    output_path: Path | None = None,
    today: date | None = None,
) -> StandardizeResult:
    input_path = input_path.resolve()
    if not is_supported_input_path(input_path):
        raise ValueError(unsupported_input_message())
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    if output_path is None:
        output_xlsx, summary_json = make_output_paths(input_path, output_dir)
    else:
        output_xlsx = output_path.resolve()
        summary_json = output_xlsx.with_suffix(".standardized.summary.json")

    if output_xlsx.resolve() == input_path:
        raise OutputPathConflictError("Output path must be a new workbook path, not the input file.")

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    input_workbook = load_workbook_for_processing(input_path, read_only=True, data_only=True)
    output_workbook = Workbook()

    sheet_summaries: list[SheetStandardizeSummary] = []
    for sheet_index, source_sheet in enumerate(input_workbook.worksheets):
        output_sheet = output_workbook.active if sheet_index == 0 else output_workbook.create_sheet()
        output_sheet.title = source_sheet.title
        sheet_summaries.append(standardize_sheet(source_sheet, output_sheet, config, today=today))

    output_workbook.save(output_xlsx)

    summary = {
        "input_path": str(input_path),
        "output_xlsx": str(output_xlsx),
        "sheets_processed": len(sheet_summaries),
        "data_rows_written": sum(sheet.data_rows_written for sheet in sheet_summaries),
        "header_row": config.header_row,
        "output_headers": [column.header for column in config.output_columns],
        "drop_headers": list(config.drop_headers),
        "sheets": [
            {
                "sheet_name": sheet.sheet_name,
                "data_rows_written": sheet.data_rows_written,
                "selected_columns": [
                    {
                        "output_header": column.output_header,
                        "source_header": value_for_json(column.source_header),
                        "source_column": column.source_column,
                    }
                    for column in sheet.selected_columns
                ],
                "omitted_headers": [value_for_json(header) for header in sheet.omitted_headers],
                "configured_drop_headers_found": [
                    value_for_json(header) for header in sheet.configured_drop_headers_found
                ],
            }
            for sheet in sheet_summaries
        ],
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return StandardizeResult(
        input_path=input_path,
        output_xlsx=output_xlsx,
        summary_json=summary_json,
        sheets_processed=len(sheet_summaries),
        data_rows_written=sum(sheet.data_rows_written for sheet in sheet_summaries),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Standardize Bazhuayu Excel/CSV headers into a fixed safe schema.")
    parser.add_argument("input_path", type=Path, help="需要整理表头的 .xlsx/.xlsm/.csv 文件")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH, help="表头标准化配置文件")
    parser.add_argument("--output-dir", type=Path, default=None, help="输出目录")
    parser.add_argument("--output", type=Path, default=None, help="输出 .xlsx 文件路径")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = standardize_workbook(
        input_path=args.input_path,
        config=load_config(args.config),
        output_dir=args.output_dir,
        output_path=args.output,
    )
    print(f"Standardized xlsx: {result.output_xlsx}")
    print(f"Summary: {result.summary_json}")
    print(f"Sheets processed: {result.sheets_processed}")
    print(f"Data rows written: {result.data_rows_written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
