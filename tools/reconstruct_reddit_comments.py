"""Reconstruct Reddit comment rows using exact Comment ID matches."""

from __future__ import annotations

import argparse
import csv
import json
import os
import shutil
import sys
from contextlib import contextmanager
from pathlib import Path
from typing import Iterator
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.utils.exceptions import IllegalCharacterError

try:
    from tools.output_path_safety import (
        OutputPathConflictError,
        atomic_output_path,
        ensure_output_paths_safe,
    )
    from tools.reddit_free_csv import FreeRedditExport, parse_free_reddit_csv
    from tools.reddit_json_export import RedditJsonError, parse_reddit_json
    from tools.reddit_json_text_merge import (
        JSON_TEXT_OUTPUT_HEADERS,
        reconstruct_json_text_rows,
    )
    from tools.reddit_page_text import RedditPageTextError, parse_reddit_page_text
    from tools.reddit_saved_html import SavedRedditHtml, parse_saved_reddit_html
except ModuleNotFoundError:
    sys.path.insert(0, str(Path(__file__).resolve().parent.parent))
    from tools.output_path_safety import (
        OutputPathConflictError,
        atomic_output_path,
        ensure_output_paths_safe,
    )
    from tools.reddit_free_csv import FreeRedditExport, parse_free_reddit_csv
    from tools.reddit_json_export import RedditJsonError, parse_reddit_json
    from tools.reddit_json_text_merge import (
        JSON_TEXT_OUTPUT_HEADERS,
        reconstruct_json_text_rows,
    )
    from tools.reddit_page_text import RedditPageTextError, parse_reddit_page_text
    from tools.reddit_saved_html import SavedRedditHtml, parse_saved_reddit_html


OUTPUT_HEADERS = (
    "Title",
    "Post Body",
    "Post URL",
    "Post Author",
    "Post Score",
    "Post Comment Count",
    "Author",
    "Time",
    "Score",
    "Thread Level",
    "Is Reply",
    "Comment",
    "Comment URL",
    "Comment ID",
    "Parent ID",
)
_XLSX_STRING_LIMIT = 32_767


def _validate_input_aliases(
    input_paths: tuple[Path, Path],
    output_xlsx: Path,
    output_csv: Path,
) -> None:
    resolved_inputs = {path.resolve() for path in input_paths}
    for output in (output_xlsx.resolve(), output_csv.resolve()):
        if output in resolved_inputs:
            raise OutputPathConflictError(
                f"Output path must be a new path, not an input file: {output}"
            )


def _validate_output_roles(output_xlsx: Path, output_csv: Path) -> None:
    if output_xlsx.resolve() == output_csv.resolve():
        raise OutputPathConflictError(
            f"Duplicate output path is not allowed: {output_xlsx.resolve()}"
        )
    if output_xlsx.suffix.lower() != ".xlsx":
        raise ValueError("XLSX output path must use the .xlsx suffix")
    if output_csv.suffix.lower() != ".csv":
        raise ValueError("CSV output path must use the .csv suffix")


def _validate_xlsx_string_lengths(
    rows: list[dict[str, str | int]],
    headers: tuple[str, ...],
) -> None:
    for column_number, header in enumerate(headers, start=1):
        if len(header) > _XLSX_STRING_LIMIT:
            raise ValueError(
                "XLSX string exceeds 32,767 characters at "
                f"header row, column {column_number}"
            )
    for row_number, row in enumerate(rows, start=1):
        for header in headers:
            value = row[header]
            if isinstance(value, str) and len(value) > _XLSX_STRING_LIMIT:
                raise ValueError(
                    "XLSX string exceeds 32,767 characters at "
                    f"data row {row_number}, column {header}"
                )


def _write_xlsx(
    rows: list[dict[str, str | int]],
    output_path: Path,
    headers: tuple[str, ...],
) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Reddit Comments"
    for column_number, header in enumerate(headers, start=1):
        cell = sheet.cell(1, column_number, header)
        cell.data_type = "s"
    for row_number, row in enumerate(rows, start=2):
        for column_number, header in enumerate(headers, start=1):
            value = row[header]
            cell = sheet.cell(row_number, column_number, value)
            if isinstance(value, str):
                cell.data_type = "s"
    workbook.save(output_path)


def _write_csv(
    rows: list[dict[str, str | int]],
    output_path: Path,
    headers: tuple[str, ...],
) -> None:
    with output_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=headers)
        writer.writeheader()
        writer.writerows(rows)


def _transaction_path(target: Path, role: str) -> Path:
    return target.with_name(
        f".{target.stem}.{uuid4().hex}.reddit-{role}{target.suffix}"
    )


def _rollback_committed_outputs(
    committed_targets: list[Path],
    backups: dict[Path, Path],
) -> tuple[list[str], set[Path]]:
    failures: list[str] = []
    retained_backups: set[Path] = set()
    for target in reversed(committed_targets):
        backup = backups.get(target)
        if backup is not None:
            try:
                os.replace(backup, target)
            except BaseException as error:
                retained_backups.add(backup)
                failures.append(
                    f"failed to restore {target} from retained backup "
                    f"{backup}: {type(error).__name__}: {error}"
                )
        else:
            try:
                target.unlink(missing_ok=True)
            except BaseException as error:
                failures.append(
                    f"failed to remove newly created output; surviving target "
                    f"{target}: {type(error).__name__}: {error}"
                )
    return failures, retained_backups


def _replace_output_pair(
    staged_and_targets: tuple[tuple[Path, Path], tuple[Path, Path]],
) -> None:
    backups: dict[Path, Path] = {}
    try:
        for _, target in staged_and_targets:
            if target.exists():
                backup = _transaction_path(target, "backup")
                backups[target] = backup
                shutil.copyfile(target, backup)
    except BaseException:
        for backup in backups.values():
            backup.unlink(missing_ok=True)
        raise
    retained_backups: set[Path] = set()
    committed_targets: list[Path] = []
    try:
        for staged, target in staged_and_targets:
            os.replace(staged, target)
            committed_targets.append(target)
    except BaseException as commit_error:
        rollback_failures, retained_backups = _rollback_committed_outputs(
            committed_targets,
            backups,
        )
        if rollback_failures:
            failure_details = "; ".join(rollback_failures)
            raise RuntimeError(
                f"Output pair commit failed: {type(commit_error).__name__}: "
                f"{commit_error}; rollback failure(s): {failure_details}"
            ) from commit_error
        raise
    finally:
        for staged, _ in staged_and_targets:
            staged.unlink(missing_ok=True)
        for backup in backups.values():
            if backup not in retained_backups:
                backup.unlink(missing_ok=True)


def _reservation_path(target: Path) -> Path:
    return target.with_name(f".{target.name}.reddit-output.lock")


@contextmanager
def _reserve_output_paths(output_paths: tuple[Path, Path]) -> Iterator[None]:
    acquired: list[Path] = []
    try:
        for target in sorted(output_paths, key=lambda path: str(path).casefold()):
            target.parent.mkdir(parents=True, exist_ok=True)
            reservation = _reservation_path(target)
            try:
                descriptor = os.open(
                    reservation,
                    os.O_CREAT | os.O_EXCL | os.O_WRONLY,
                    0o600,
                )
            except FileExistsError as error:
                raise OutputPathConflictError(
                    f"Output path is reserved by another writer: {target}"
                ) from error
            acquired.append(reservation)
            os.close(descriptor)
        yield
    finally:
        for reservation in reversed(acquired):
            reservation.unlink(missing_ok=True)


def write_outputs(
    rows: list[dict[str, str | int]],
    *,
    headers: tuple[str, ...] = OUTPUT_HEADERS,
    input_paths: tuple[Path, Path],
    output_xlsx: Path,
    output_csv: Path,
    overwrite: bool,
) -> None:
    _validate_input_aliases(input_paths, output_xlsx, output_csv)
    _validate_output_roles(output_xlsx, output_csv)
    _validate_xlsx_string_lengths(rows, headers)
    requested_outputs = (output_xlsx.resolve(), output_csv.resolve())
    with _reserve_output_paths(requested_outputs):
        resolved_xlsx, resolved_csv = ensure_output_paths_safe(
            input_paths,
            requested_outputs,
            overwrite=overwrite,
        )
        staged_xlsx = _transaction_path(resolved_xlsx, "stage")
        staged_csv = _transaction_path(resolved_csv, "stage")
        try:
            with atomic_output_path(staged_xlsx) as temporary_xlsx:
                _write_xlsx(rows, temporary_xlsx, headers)
            with atomic_output_path(staged_csv) as temporary_csv:
                _write_csv(rows, temporary_csv, headers)
            _replace_output_pair(
                ((staged_xlsx, resolved_xlsx), (staged_csv, resolved_csv))
            )
        finally:
            staged_xlsx.unlink(missing_ok=True)
            staged_csv.unlink(missing_ok=True)


def _required_post_value(
    html_value: str,
    fallback: str | None,
    field_name: str,
    cli_name: str,
) -> str:
    if html_value.strip():
        return html_value
    if fallback is not None and fallback.strip():
        return fallback
    raise ValueError(
        f"Missing required post field {field_name}; explicit CLI value "
        f"{cli_name} is needed"
    )


def reconstruct_rows(
    free: FreeRedditExport,
    html: SavedRedditHtml,
    *,
    post_author: str | None = None,
    post_score: str | None = None,
    post_comment_count: str | None = None,
) -> list[dict[str, str | int]]:
    if free.post_id != html.post_id:
        raise ValueError(
            f"Reddit post ID mismatch: free CSV {free.post_id!r}, "
            f"saved HTML {html.post_id!r}"
        )

    resolved_post_author = _required_post_value(
        html.post_author, post_author, "Post Author", "post_author"
    )
    resolved_post_score = _required_post_value(
        html.post_score, post_score, "Post Score", "post_score"
    )
    resolved_post_comment_count = _required_post_value(
        html.post_comment_count,
        post_comment_count,
        "Post Comment Count",
        "post_comment_count",
    )

    missing_ids: list[str] = []
    invalid_hierarchy_ids: list[str] = []
    for comment in free.comments:
        html_comment = html.comments.get(comment.comment_id)
        if html_comment is None:
            missing_ids.append(comment.comment_id)
        elif html_comment.parent_id == "" or html_comment.thread_level is None:
            invalid_hierarchy_ids.append(comment.comment_id)

    if missing_ids or invalid_hierarchy_ids:
        categories: list[str] = []
        if missing_ids:
            categories.append("Missing HTML comments: " + ", ".join(missing_ids))
        if invalid_hierarchy_ids:
            categories.append(
                "Invalid hierarchy: " + ", ".join(invalid_hierarchy_ids)
            )
        raise ValueError("; ".join(categories))

    rows: list[dict[str, str | int]] = []
    for comment in free.comments:
        html_comment = html.comments[comment.comment_id]
        thread_level = html_comment.thread_level
        assert thread_level is not None
        rows.append(
            {
                "Title": free.title,
                "Post Body": free.body,
                "Post URL": free.url,
                "Post Author": resolved_post_author,
                "Post Score": resolved_post_score,
                "Post Comment Count": resolved_post_comment_count,
                "Author": comment.author,
                "Time": comment.time,
                "Score": html_comment.score,
                "Thread Level": thread_level,
                "Is Reply": "No" if thread_level == 0 else "Yes",
                "Comment": comment.comment,
                "Comment URL": comment.comment_url,
                "Comment ID": comment.comment_id,
                "Parent ID": html_comment.parent_id,
            }
        )
    return rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description=(
            "Deterministically combine a Reddit JSON export with copied "
            "Reddit page text."
        )
    )
    parser.add_argument("--json", required=True, type=Path)
    parser.add_argument("--page-text", required=True, type=Path)
    parser.add_argument("--output-xlsx", required=True, type=Path)
    parser.add_argument("--output-csv", required=True, type=Path)
    parser.add_argument("--overwrite", action="store_true")
    return parser


def _safe_cli_error(error: BaseException) -> str:
    if isinstance(error, IllegalCharacterError):
        return "XLSX output contains an unsupported control character"
    if isinstance(error, json.JSONDecodeError):
        return "JSON is invalid"
    return str(error)


def main(argv: list[str] | None = None) -> int:
    arguments = build_parser().parse_args(argv)
    try:
        json_path = arguments.json.resolve()
        page_text_path = arguments.page_text.resolve()
        output_xlsx = arguments.output_xlsx.resolve()
        output_csv = arguments.output_csv.resolve()
        export = parse_reddit_json(json_path)
        page = parse_reddit_page_text(page_text_path, export)
        rows = reconstruct_json_text_rows(export, page)
        write_outputs(
            rows,
            headers=JSON_TEXT_OUTPUT_HEADERS,
            input_paths=(json_path, page_text_path),
            output_xlsx=output_xlsx,
            output_csv=output_csv,
            overwrite=arguments.overwrite,
        )
    except (
        RedditJsonError,
        RedditPageTextError,
        IllegalCharacterError,
        OSError,
        csv.Error,
        RuntimeError,
        ValueError,
    ) as error:
        print(f"Error: {_safe_cli_error(error)}", file=sys.stderr)
        return 1

    print(f"Reddit JSON input: {json_path}")
    print(f"Reddit page text input: {page_text_path}")
    print(f"XLSX output: {output_xlsx}")
    print(f"CSV output: {output_csv}")
    print(f"JSON comment count: {len(export.comments)}")
    print(f"Page comment match count: {len(page.comments)}")
    print(
        "Missing comment score count: "
        f"{sum(item.score is None for item in page.comments)}"
    )
    print(f"Unavailable reported comment gap: {export.meta.discrepancy}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
