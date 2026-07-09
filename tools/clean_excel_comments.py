from __future__ import annotations

import argparse
import csv
import json
import re
from dataclasses import dataclass, replace
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

try:
    from tools.csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message
except ModuleNotFoundError:
    from csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message


DEFAULT_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "comment-cleaner.json"
CHINESE_CHAR_PATTERN = re.compile(r"[\u4e00-\u9fff]")
ALNUM_TOKEN_PATTERN = re.compile(r"[A-Za-z0-9]+")
DEFAULT_DELETE_CONTAINS_TEXTS = (
    "链接",
    "凑字数",
    "水经验",
    "赚积分",
    "为了金币",
    "赚硬币",
    "赚京豆",
    "淘气值",
    "为了评论而评论",
    "混个脸熟",
    "完成任务",
    "代下",
    "代买",
    "内部券",
    "加微",
    "加v",
    "私聊我",
    "主页看",
    "点击链接",
    "http://",
    "https://",
    "第一",
    "打卡",
    "路过",
    "来了",
    "冒泡",
    "占座",
    "测试",
    "test",
    "无",
    "无内容",
    "略",
    "暂无评价",
    "蹲",
    "蹲一个",
    "求链接",
    "求分享",
    "多少钱",
    "怎么卖",
    "啥牌子",
    "什么牌子",
    "求品牌",
    "求私",
    "加群",
    "裙内",
    "互赞",
    "互粉",
    "互关",
    "回关",
    "秒回",
    "交朋友",
    "リンク",
    "プロフィール見て",
    "プロフ見て",
    "DMして",
    "フォロー返し",
    "相互フォロー",
    "テスト",
    "内容なし",
    "評価なし",
    "コメント稼ぎ",
    "링크",
    "맞팔",
    "테스트",
    "내용 없음",
)
DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS = (
    "加v",
    "link in bio",
    "click link",
    "click the link",
    "check my profile",
    "see my profile",
    "visit my profile",
    "dm me",
    "message me",
    "follow me",
    "follow back",
    "follow for follow",
    "sub4sub",
    "sub for sub",
    "subscribe to my channel",
    "earn coins",
    "free coins",
    "for coins",
    "comment for points",
    "promo code",
    "coupon code",
    "discount code",
    "whatsapp",
    "telegram",
    "first",
    "test",
    "n/a",
    "no content",
    "no comment",
    "nothing to say",
)


@dataclass(frozen=True)
class CleanerConfig:
    target_column: int = 3
    target_header: str | None = None
    first_data_row: int = 2
    min_trimmed_length: int = 8
    delete_exact_texts: tuple[str, ...] = ("该用户未填写评价内容", "此用户未填写评价内容")
    delete_contains_texts: tuple[str, ...] = DEFAULT_DELETE_CONTAINS_TEXTS
    delete_contains_case_insensitive_texts: tuple[str, ...] = DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS
    delete_random_alnum_without_chinese: bool = True
    random_digit_min_length: int = 9
    random_letter_min_length: int = 10
    random_mixed_min_length: int = 10
    random_letter_max_vowel_ratio: float = 0.2
    random_letter_min_consonant_run: int = 5
    subcomment_deduplicate_headers: tuple[str, ...] = ("一级评论", "二级评论", "三级评论")
    subcomment_min_trimmed_length: int = 6
    duplicate_keep: str = "last"
    export_first_sheet_csv: bool = True
    csv_encoding: str = "utf-8-sig"


@dataclass(frozen=True)
class DeletedRow:
    sheet: str
    row_number: int
    reason: str
    value: str


@dataclass(frozen=True)
class ClearedCell:
    sheet: str
    row_number: int
    column_header: str
    reason: str
    value: str


@dataclass(frozen=True)
class CleanResult:
    input_path: Path
    output_xlsx: Path
    output_csv: Path | None
    deletion_log_csv: Path
    summary_json: Path
    sheets_processed: int
    rows_deleted: int
    cells_cleared: int


def load_config(path: Path) -> CleanerConfig:
    data = json.loads(path.read_text(encoding="utf-8"))
    return CleanerConfig(
        target_column=int(data.get("target_column", 3)),
        target_header=data.get("target_header"),
        first_data_row=int(data.get("first_data_row", 2)),
        min_trimmed_length=int(data.get("min_trimmed_length", 8)),
        delete_exact_texts=tuple(data.get("delete_exact_texts", [])),
        delete_contains_texts=tuple(data.get("delete_contains_texts", DEFAULT_DELETE_CONTAINS_TEXTS)),
        delete_contains_case_insensitive_texts=tuple(
            data.get("delete_contains_case_insensitive_texts", DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS)
        ),
        delete_random_alnum_without_chinese=bool(data.get("delete_random_alnum_without_chinese", True)),
        random_digit_min_length=int(data.get("random_digit_min_length", 9)),
        random_letter_min_length=int(data.get("random_letter_min_length", 10)),
        random_mixed_min_length=int(data.get("random_mixed_min_length", 10)),
        random_letter_max_vowel_ratio=float(data.get("random_letter_max_vowel_ratio", 0.2)),
        random_letter_min_consonant_run=int(data.get("random_letter_min_consonant_run", 5)),
        subcomment_deduplicate_headers=tuple(
            data.get("subcomment_deduplicate_headers", ["一级评论", "二级评论", "三级评论"])
        ),
        subcomment_min_trimmed_length=int(data.get("subcomment_min_trimmed_length", 6)),
        duplicate_keep=str(data.get("duplicate_keep", "last")),
        export_first_sheet_csv=bool(data.get("export_first_sheet_csv", True)),
        csv_encoding=str(data.get("csv_encoding", "utf-8-sig")),
    )


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split())


def max_consonant_run(value: str) -> int:
    max_run = 0
    current_run = 0
    for char in value.lower():
        if char.isalpha() and char not in "aeiou":
            current_run += 1
            max_run = max(max_run, current_run)
        else:
            current_run = 0
    return max_run


def is_random_alnum_without_chinese(comment: str, config: CleanerConfig) -> bool:
    if not comment or CHINESE_CHAR_PATTERN.search(comment):
        return False

    for token in ALNUM_TOKEN_PATTERN.findall(comment):
        has_alpha = any(char.isalpha() for char in token)
        has_digit = any(char.isdigit() for char in token)
        if has_digit and not has_alpha and len(token) >= config.random_digit_min_length:
            return True
        if has_digit and has_alpha and len(token) >= config.random_mixed_min_length:
            return True
        if has_alpha and not has_digit and len(token) >= config.random_letter_min_length:
            vowel_count = sum(1 for char in token.lower() if char in "aeiou")
            vowel_ratio = vowel_count / len(token)
            if vowel_ratio <= config.random_letter_max_vowel_ratio:
                return True
            if max_consonant_run(token) >= config.random_letter_min_consonant_run:
                return True
    return False


def resolve_target_column(sheet: Worksheet, config: CleanerConfig) -> int:
    if not config.target_header:
        return config.target_column

    header_row = max(1, config.first_data_row - 1)
    target_key = normalize_header(config.target_header)
    headers = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, max_col=sheet.max_column, values_only=True),
        (),
    )
    matches = [
        column_index
        for column_index, header in enumerate(headers, start=1)
        if normalize_header(header) == target_key
    ]
    if not matches:
        raise ValueError(f"未找到评论列表头: {config.target_header}")
    if len(matches) > 1:
        raise ValueError(f"评论列表头重复: {config.target_header}")
    return matches[0]


def resolve_optional_header_columns(sheet: Worksheet, config: CleanerConfig, headers: tuple[str, ...]) -> list[tuple[int, str]]:
    header_row = max(1, config.first_data_row - 1)
    wanted = {normalize_header(header): header for header in headers}
    if not wanted:
        return []

    sheet_headers = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, max_col=sheet.max_column, values_only=True),
        (),
    )
    columns: list[tuple[int, str]] = []
    for column_index, header in enumerate(sheet_headers, start=1):
        key = normalize_header(header)
        if key in wanted:
            columns.append((column_index, wanted[key]))
    return columns


def should_delete_comment(
    comment: str,
    seen_comments: set[str],
    config: CleanerConfig,
    clean_words: tuple[str, ...],
) -> str | None:
    if len(comment) < config.min_trimmed_length:
        return f"评论长度小于 {config.min_trimmed_length}"

    if comment in config.delete_exact_texts:
        return "评论等于占位文案"

    for clean_word in clean_words:
        if clean_word and clean_word in comment:
            return f"评论包含清理词: {clean_word}"

    for text in config.delete_contains_texts:
        if text and text in comment:
            return f"评论包含固定删除词: {text}"

    casefolded_comment = comment.casefold()
    for text in config.delete_contains_case_insensitive_texts:
        if text and text.casefold() in casefolded_comment:
            return f"评论包含固定删除词: {text}"

    if config.delete_random_alnum_without_chinese and is_random_alnum_without_chinese(comment, config):
        return "评论为无中文随机英文/数字堆砌"

    if comment in seen_comments:
        return "同一工作表内重复评论"

    return None


def iter_row_numbers(sheet: Worksheet, config: CleanerConfig) -> range:
    if config.duplicate_keep == "last":
        return range(sheet.max_row, config.first_data_row - 1, -1)
    if config.duplicate_keep == "first":
        return range(config.first_data_row, sheet.max_row + 1)
    raise ValueError("duplicate_keep 只能是 first 或 last")


def iter_cell_positions(sheet: Worksheet, config: CleanerConfig, columns: list[tuple[int, str]]) -> list[tuple[int, int, str]]:
    if config.duplicate_keep == "last":
        row_numbers = range(sheet.max_row, config.first_data_row - 1, -1)
        ordered_columns = list(reversed(columns))
    elif config.duplicate_keep == "first":
        row_numbers = range(config.first_data_row, sheet.max_row + 1)
        ordered_columns = columns
    else:
        raise ValueError("duplicate_keep 只能是 first 或 last")

    return [(row_number, column_index, header) for row_number in row_numbers for column_index, header in ordered_columns]


def clear_duplicate_subcomments(sheet: Worksheet, config: CleanerConfig) -> list[ClearedCell]:
    columns = resolve_optional_header_columns(sheet, config, config.subcomment_deduplicate_headers)
    if not columns:
        return []

    seen_values: set[str] = set()
    cleared_cells: list[ClearedCell] = []

    for row_number, column_index, header in iter_cell_positions(sheet, config, columns):
        cell = sheet.cell(row=row_number, column=column_index)
        value = normalize_cell(cell.value)
        if not value:
            continue
        if len(value) < config.subcomment_min_trimmed_length:
            cell.value = None
            cleared_cells.append(
                ClearedCell(
                    sheet=sheet.title,
                    row_number=row_number,
                    column_header=header,
                    reason=f"子评论长度小于 {config.subcomment_min_trimmed_length}: {header}",
                    value=value,
                )
            )
            continue
        if value in seen_values:
            cell.value = None
            cleared_cells.append(
                ClearedCell(
                    sheet=sheet.title,
                    row_number=row_number,
                    column_header=header,
                    reason=f"同一工作表内重复子评论: {header}",
                    value=value,
                )
            )
        else:
            seen_values.add(value)

    return sorted(cleared_cells, key=lambda cell: (cell.row_number, cell.column_header))


def clean_sheet(sheet: Worksheet, config: CleanerConfig, clean_words: tuple[str, ...]) -> tuple[list[DeletedRow], list[ClearedCell]]:
    deleted: list[DeletedRow] = []
    seen_comments: set[str] = set()
    pending_deletions: list[DeletedRow] = []
    target_column = resolve_target_column(sheet, config)

    for row_number in iter_row_numbers(sheet, config):
        comment = normalize_cell(sheet.cell(row=row_number, column=target_column).value)
        reason = should_delete_comment(comment, seen_comments, config, clean_words)

        if reason:
            pending_deletions.append(
                DeletedRow(
                    sheet=sheet.title,
                    row_number=row_number,
                    reason=reason,
                    value=comment,
                )
            )
        else:
            seen_comments.add(comment)

    for deleted_row in sorted(pending_deletions, key=lambda row: row.row_number, reverse=True):
        sheet.delete_rows(deleted_row.row_number, 1)
        deleted.append(deleted_row)

    cleared_cells = clear_duplicate_subcomments(sheet, config)

    return deleted, cleared_cells


def make_output_paths(input_path: Path, output_dir: Path | None, output_path: Path | None = None) -> tuple[Path, Path, Path, Path]:
    if output_path:
        output_xlsx = output_path.resolve()
        return (
            output_xlsx,
            output_xlsx.with_suffix(".csv"),
            output_xlsx.with_suffix(".deletions.csv"),
            output_xlsx.with_suffix(".summary.json"),
        )

    timestamp = datetime.now().strftime("%Y%m%d")
    parent = output_dir if output_dir else input_path.parent
    stem = f"{timestamp}_{input_path.stem}"
    return (
        parent / f"{stem}.cleaned.xlsx",
        parent / f"{stem}.cleaned.csv",
        parent / f"{stem}.deletions.csv",
        parent / f"{stem}.summary.json",
    )


def write_first_sheet_csv(sheet: Worksheet, path: Path, encoding: str) -> None:
    with path.open("w", newline="", encoding=encoding) as csv_file:
        writer = csv.writer(csv_file)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def write_deletion_log(path: Path, deleted_rows: list[DeletedRow], cleared_cells: list[ClearedCell], encoding: str) -> None:
    with path.open("w", newline="", encoding=encoding) as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["action", "sheet", "row_number", "column_header", "reason", "value"])
        writer.writeheader()
        for row in deleted_rows:
            writer.writerow(
                {
                    "action": "delete_row",
                    "sheet": row.sheet,
                    "row_number": row.row_number,
                    "column_header": "",
                    "reason": row.reason,
                    "value": row.value,
                }
            )
        for cell in cleared_cells:
            writer.writerow(
                {
                    "action": "clear_cell",
                    "sheet": cell.sheet,
                    "row_number": cell.row_number,
                    "column_header": cell.column_header,
                    "reason": cell.reason,
                    "value": cell.value,
                }
            )


def clean_workbook(
    input_path: Path,
    config: CleanerConfig,
    clean_words: tuple[str, ...],
    output_dir: Path | None = None,
    output_path: Path | None = None,
) -> CleanResult:
    if not is_supported_input_path(input_path):
        raise ValueError(unsupported_input_message())

    output_xlsx, output_csv, deletion_log_csv, summary_json = make_output_paths(input_path, output_dir, output_path)
    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook_for_processing(input_path, read_only=False, data_only=False)
    deleted_rows: list[DeletedRow] = []
    cleared_cells: list[ClearedCell] = []

    for sheet in workbook.worksheets:
        sheet_deleted_rows, sheet_cleared_cells = clean_sheet(sheet, config, clean_words)
        deleted_rows.extend(sheet_deleted_rows)
        cleared_cells.extend(sheet_cleared_cells)

    workbook.save(output_xlsx)

    actual_output_csv: Path | None = None
    if config.export_first_sheet_csv and workbook.worksheets:
        write_first_sheet_csv(workbook.worksheets[0], output_csv, config.csv_encoding)
        actual_output_csv = output_csv

    write_deletion_log(deletion_log_csv, deleted_rows, cleared_cells, config.csv_encoding)

    summary = {
        "input_path": str(input_path),
        "output_xlsx": str(output_xlsx),
        "output_csv": str(actual_output_csv) if actual_output_csv else None,
        "deletion_log_csv": str(deletion_log_csv),
        "sheets_processed": len(workbook.worksheets),
        "rows_deleted": len(deleted_rows),
        "cells_cleared": len(cleared_cells),
        "target_column": config.target_column,
        "target_header": config.target_header,
        "first_data_row": config.first_data_row,
        "clean_words": list(clean_words),
        "delete_contains_texts": list(config.delete_contains_texts),
        "delete_contains_case_insensitive_texts": list(config.delete_contains_case_insensitive_texts),
        "delete_random_alnum_without_chinese": config.delete_random_alnum_without_chinese,
        "random_digit_min_length": config.random_digit_min_length,
        "random_letter_min_length": config.random_letter_min_length,
        "random_mixed_min_length": config.random_mixed_min_length,
        "random_letter_max_vowel_ratio": config.random_letter_max_vowel_ratio,
        "random_letter_min_consonant_run": config.random_letter_min_consonant_run,
        "duplicate_keep": config.duplicate_keep,
        "subcomment_deduplicate_headers": list(config.subcomment_deduplicate_headers),
        "subcomment_min_trimmed_length": config.subcomment_min_trimmed_length,
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return CleanResult(
        input_path=input_path,
        output_xlsx=output_xlsx,
        output_csv=actual_output_csv,
        deletion_log_csv=deletion_log_csv,
        summary_json=summary_json,
        sheets_processed=len(workbook.worksheets),
        rows_deleted=len(deleted_rows),
        cells_cleared=len(cleared_cells),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="清洗八爪鱼导出的 Excel/CSV 评论表")
    parser.add_argument("input_path", type=Path, help="需要清洗的 .xlsx/.xlsm/.csv 文件")
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help=f"清洗规则配置文件，默认 {DEFAULT_CONFIG_PATH}",
    )
    parser.add_argument(
        "--clean-word",
        action="append",
        default=[],
        help="额外清理词，可重复传入。例如 --clean-word KOL清理词1 --clean-word KOL清理词2",
    )
    parser.add_argument("--target-header", default=None, help="按表头定位评论列，例如 评论内容")
    parser.add_argument("--output-dir", type=Path, default=None, help="输出目录，默认写到输入文件同目录")
    parser.add_argument("--output", type=Path, default=None, help="输出 .xlsx 文件路径")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    config = load_config(args.config)
    if args.target_header:
        config = replace(config, target_header=args.target_header)
    clean_words = tuple(word.strip() for word in args.clean_word if word and word.strip())
    result = clean_workbook(args.input_path.resolve(), config, clean_words, args.output_dir, args.output)

    print(f"处理完成: {result.input_path}")
    print(f"工作表数量: {result.sheets_processed}")
    print(f"删除行数: {result.rows_deleted}")
    print(f"清空单元格数: {result.cells_cleared}")
    print(f"清洗后 xlsx: {result.output_xlsx}")
    if result.output_csv:
        print(f"首个工作表 csv: {result.output_csv}")
    print(f"删除日志: {result.deletion_log_csv}")
    print(f"摘要: {result.summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
