from __future__ import annotations

import csv
import json
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest

from openpyxl import load_workbook

from reddit_reconstruction.merge import JSON_TEXT_OUTPUT_HEADERS


class RedditPackageCliTests(unittest.TestCase):
    def test_json_primary_cli_writes_post_before_all_json_comments(self) -> None:
        repo_root = Path(__file__).resolve().parents[2]
        with tempfile.TemporaryDirectory() as temporary_directory:
            directory = Path(temporary_directory)
            json_path = directory / "export.json"
            page_text_path = directory / "page.txt"
            output_xlsx = directory / "result.xlsx"
            output_csv = directory / "result.csv"
            json_path.write_text(
                json.dumps(
                    {
                        "meta": {
                            "completeness": "complete",
                            "collectedCommentCount": 2,
                            "reportedByApi": 2,
                            "discrepancy": 0,
                            "failedMore": 0,
                            "failedNodes": [],
                            "failedReasons": [],
                            "failedDetails": [],
                        },
                        "post": {
                            "id": "postone",
                            "subreddit": "python",
                            "title": "CLI title",
                            "content": "",
                            "author": "poster",
                            "num_comments": 2,
                        },
                        "comments": [
                            {
                                "id": "rootone",
                                "parent_id": "postone",
                                "content": "matched comment",
                                "depth": 0,
                                "username": "commenter",
                                "date": "2026-07-28",
                                "created_utc": 1,
                            },
                            {
                                "id": "childone",
                                "parent_id": "rootone",
                                "content": "unmatched comment",
                                "depth": 1,
                                "username": "replier",
                                "date": "2026-07-28",
                                "created_utc": 2,
                            },
                        ],
                    }
                ),
                encoding="utf-8",
            )
            page_text_path.write_text(
                "\n".join(
                    (
                        "Reddit",
                        "r/python",
                        "u/poster",
                        "poster 头像",
                        "8小时前",
                        "CLI title",
                        "正文",
                        "赞同",
                        "11",
                        "反对",
                        "2",
                        "转到评论",
                        "评论区域",
                        "commenter",
                        "•8小时前",
                        "matched comment",
                        "",
                        "赞同",
                        "7",
                        "反对",
                        "回复",
                        "奖励",
                        "分享",
                    )
                ),
                encoding="utf-8",
            )

            completed = subprocess.run(
                [
                    sys.executable,
                    "-m",
                    "reddit_reconstruction",
                    "--json",
                    str(json_path),
                    "--page-text",
                    str(page_text_path),
                    "--output-xlsx",
                    str(output_xlsx),
                    "--output-csv",
                    str(output_csv),
                    "--json-primary-page-metrics",
                ],
                cwd=repo_root,
                capture_output=True,
                text=True,
                check=False,
            )

            self.assertEqual(0, completed.returncode, completed.stderr)
            self.assertTrue(output_xlsx.is_file())
            self.assertTrue(output_csv.is_file())
            workbook = load_workbook(output_xlsx, data_only=False)
            sheet = workbook.active
            xlsx_rows = list(sheet.iter_rows(values_only=True))
            workbook.close()
            with output_csv.open("r", encoding="utf-8-sig", newline="") as handle:
                csv_rows = list(csv.DictReader(handle))

        comment_id_header = JSON_TEXT_OUTPUT_HEADERS[9]
        score_header = JSON_TEXT_OUTPUT_HEADERS[5]
        self.assertEqual(list(JSON_TEXT_OUTPUT_HEADERS), list(xlsx_rows[0]))
        self.assertEqual(
            ["postone", "rootone", "childone"],
            [row[9] for row in xlsx_rows[1:]],
        )
        self.assertEqual("", csv_rows[2][score_header])
        self.assertEqual(
            ["postone", "rootone", "childone"],
            [row[comment_id_header] for row in csv_rows],
        )
