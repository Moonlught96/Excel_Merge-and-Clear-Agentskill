from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook

try:
    from tools.csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message
except ModuleNotFoundError:
    from csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message


class HeaderMismatchError(ValueError):
    pass


class OutputPathConflictError(ValueError):
    pass


@dataclass(frozen=True)
class MergeResult:
    output_path: Path
    summary_path: Path
    files_processed: int
    sheets_processed: int
    data_rows_written: int


def normalize_for_header(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def value_for_json(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def is_blank_row(row: tuple[Any, ...]) -> bool:
    return all(value is None or str(value).strip() == "" for value in row)


def validate_input_paths(input_paths: list[Path]) -> list[Path]:
    if not input_paths:
        raise ValueError("At least one Excel file is required.")

    validated: list[Path] = []
    for path in input_paths:
        resolved = path.resolve()
        if resolved.is_dir():
            raise ValueError("Merge input must be explicit Excel files, not a folder.")
        if not is_supported_input_path(resolved):
            raise ValueError(f"{unsupported_input_message()}: {resolved}")
        if resolved.name.startswith("~$"):
            continue
        if not resolved.exists():
            raise FileNotFoundError(resolved)
        validated.append(resolved)

    if not validated:
        raise ValueError("No usable Excel files were provided.")
    return validated


def default_output_path() -> Path:
    timestamp = datetime.now().strftime("%Y%m%d")
    return Path.cwd() / "outputs" / "merged" / f"{timestamp}_merged.xlsx"


def merge_workbooks(
    input_paths: list[Path],
    output_path: Path,
    *,
    add_source_columns: bool = False,
) -> MergeResult:
    paths = validate_input_paths(input_paths)
    output_path = output_path.resolve()
    if output_path in paths:
        raise OutputPathConflictError("Output path must be a new workbook path, not one of the input files.")

    output_path.parent.mkdir(parents=True, exist_ok=True)

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "总表"

    canonical_header: list[Any] | None = None
    canonical_header_key: list[str] | None = None
    data_rows_written = 0
    sheets_processed = 0
    file_summaries: list[dict[str, Any]] = []

    for path in paths:
        workbook = load_workbook_for_processing(path, read_only=True, data_only=False)
        file_rows = 0
        sheet_summaries: list[dict[str, Any]] = []

        for sheet in workbook.worksheets:
            rows = sheet.iter_rows(values_only=True)
            try:
                header_tuple = next(rows)
            except StopIteration:
                continue

            header = list(header_tuple)
            header_key = [normalize_for_header(value) for value in header]

            if canonical_header_key is None:
                canonical_header = header
                canonical_header_key = header_key
                output_header = list(header)
                if add_source_columns:
                    output_header.extend(["source_file", "source_sheet"])
                output_sheet.append(output_header)
            elif header_key != canonical_header_key:
                raise HeaderMismatchError(
                    f"Header mismatch in {path} / {sheet.title}. "
                    f"Expected {canonical_header_key}, got {header_key}."
                )

            sheet_rows = 0
            for row_tuple in rows:
                if is_blank_row(row_tuple):
                    continue
                row = list(row_tuple)
                if add_source_columns:
                    row.extend([path.name, sheet.title])
                output_sheet.append(row)
                sheet_rows += 1

            if sheet_rows:
                sheets_processed += 1
                file_rows += sheet_rows
                sheet_summaries.append(
                    {
                        "sheet_name": sheet.title,
                        "data_rows": sheet_rows,
                    }
                )

        file_summaries.append(
            {
                "file": str(path),
                "data_rows": file_rows,
                "sheets": sheet_summaries,
            }
        )
        data_rows_written += file_rows
        workbook.close()

    if canonical_header is None:
        raise ValueError("No non-empty worksheets were found in the provided files.")

    output_workbook.save(output_path)
    output_workbook.close()
    summary_path = output_path.with_suffix(".summary.json")
    summary = {
        "output_path": str(output_path),
        "files_processed": len(paths),
        "sheets_processed": sheets_processed,
        "data_rows_written": data_rows_written,
        "add_source_columns": add_source_columns,
        "header": [value_for_json(value) for value in canonical_header],
        "files": file_summaries,
    }
    summary_path.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return MergeResult(
        output_path=output_path,
        summary_path=summary_path,
        files_processed=len(paths),
        sheets_processed=sheets_processed,
        data_rows_written=data_rows_written,
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Merge explicitly provided user-comment Excel/CSV files into one workbook.")
    parser.add_argument("input_paths", type=Path, nargs="+", help="Explicit .xlsx/.xlsm/.csv files to merge.")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path.")
    parser.add_argument(
        "--add-source-columns",
        action="store_true",
        help="Append source_file and source_sheet columns for audit.",
    )
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    output_path = args.output if args.output else default_output_path()
    result = merge_workbooks(args.input_paths, output_path, add_source_columns=args.add_source_columns)
    print(f"Merged xlsx: {result.output_path}")
    print(f"Summary: {result.summary_path}")
    print(f"Files processed: {result.files_processed}")
    print(f"Sheets processed: {result.sheets_processed}")
    print(f"Data rows written: {result.data_rows_written}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
