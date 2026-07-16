from __future__ import annotations

import csv
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


EXCEL_SUFFIXES = {".xlsx", ".xlsm"}
CSV_SUFFIXES = {".csv"}
SUPPORTED_INPUT_SUFFIXES = EXCEL_SUFFIXES | CSV_SUFFIXES
CSV_INPUT_ENCODINGS = ("utf-8-sig", "gb18030")
INVALID_SHEET_TITLE_CHARS = re.compile(r"[\[\]:*?/\\]")


class CsvDecodeError(ValueError):
    pass


@dataclass(frozen=True)
class CsvRows:
    rows: list[list[str]]
    encoding: str


def unsupported_input_message() -> str:
    return "当前工具支持 .xlsx、.xlsm 和 .csv；老版 .xls 请先另存为 .xlsx"


def is_supported_input_path(path: Path) -> bool:
    return path.suffix.lower() in SUPPORTED_INPUT_SUFFIXES


def safe_sheet_title(path: Path) -> str:
    title = INVALID_SHEET_TITLE_CHARS.sub("_", path.stem).strip()
    if not title:
        title = "Sheet1"
    return title[:31]


def read_csv_rows(path: Path) -> CsvRows:
    for encoding in CSV_INPUT_ENCODINGS:
        try:
            with path.open("r", newline="", encoding=encoding) as csv_file:
                return CsvRows(rows=list(csv.reader(csv_file)), encoding=encoding)
        except UnicodeDecodeError:
            continue
    raise CsvDecodeError(f"CSV 解码失败，已尝试编码: {', '.join(CSV_INPUT_ENCODINGS)}")


def workbook_from_csv(path: Path) -> Workbook:
    csv_rows = read_csv_rows(path)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = safe_sheet_title(path)
    for row in csv_rows.rows:
        sheet.append(row)
    return workbook


def load_workbook_for_processing(path: Path, *, read_only: bool = False, data_only: bool = True) -> Any:
    if path.suffix.lower() == ".csv":
        return workbook_from_csv(path)
    return load_workbook(path, read_only=read_only, data_only=data_only)
