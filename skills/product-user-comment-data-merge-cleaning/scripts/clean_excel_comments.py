from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from contextlib import ExitStack
from dataclasses import dataclass, field, replace
from decimal import Decimal, InvalidOperation
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

try:
    from tools.csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message
    from tools.output_path_safety import (
        add_confirmed_overwrite_arguments,
        atomic_output_path,
        beijing_date_text,
        ensure_output_paths_safe,
    )
except ModuleNotFoundError:
    from csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message
    from output_path_safety import (
        add_confirmed_overwrite_arguments,
        atomic_output_path,
        beijing_date_text,
        ensure_output_paths_safe,
    )


DEFAULT_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "comment-cleaner.json"
STANDARDIZED_COMMENT_HEADER = "评论内容"


class ExternalCleanerConfigError(ValueError):
    pass


class FinalizedCleaningAuditArtifactsError(ValueError):
    pass


class CleanerConfigError(ValueError):
    pass


def require_canonical_cleaner_config_path(config_path: Path) -> Path:
    """Reject per-run or copied cleaner configurations at the executable boundary."""
    configured_path = config_path.resolve()
    canonical_path = DEFAULT_CONFIG_PATH.resolve()
    if configured_path != canonical_path:
        raise ExternalCleanerConfigError(
            "External or temporary cleaner configs are forbidden. "
            f"Use the bundled canonical config only: {canonical_path}"
        )
    return canonical_path


def refuse_finalized_audit_artifact_restoration(
    output_xlsx: Path,
    deletion_log_csv: Path,
    summary_json: Path,
) -> None:
    """Keep default-retention cleanup irreversible for a finalized output path."""
    if output_xlsx.exists() and (
        not deletion_log_csv.exists() or not summary_json.exists()
    ):
        raise FinalizedCleaningAuditArtifactsError(
            "Refusing to regenerate a finalized cleaning audit artifact that was removed by the retention policy. "
            "Choose a new confirmed final output path for a fresh run; do not restore finalized audit artifacts."
        )


CHINESE_CHAR_PATTERN = re.compile(r"[\u4e00-\u9fff]")
JAPANESE_KANA_PATTERN = re.compile(r"[\u3040-\u30ff]")
KOREAN_HANGUL_PATTERN = re.compile(r"[\uac00-\ud7af\u1100-\u11ff]")
THAI_PATTERN = re.compile(r"[\u0e00-\u0e7f]")
DEVANAGARI_PATTERN = re.compile(r"[\u0900-\u097f]")
ALNUM_TOKEN_PATTERN = re.compile(r"[A-Za-z0-9]+")
LIKE_COUNT_PATTERN = re.compile(r"[+-]?\d+(?:\.\d+)?")
LATIN_LETTER_PATTERN = re.compile(r"[A-Za-zÀ-ÖØ-öø-ÿĀ-ž]")
HTTPS_URL_PATTERN = re.compile(r"https://\S+")
FIXED_TERM_SCRIPT_GROUPS = frozenset(
    {"chinese", "japanese", "korean", "thai", "hindi", "latin", "neutral"}
)
DEFAULT_DELETE_CONTAINS_TEXTS = (
    "链接",
    "凑字数",
    "水经验",
    "赚积分",
    "为了金币",
    "赚硬币",
    "赚京豆",
    "淘气值",
    "为了评论而评论",
    "混个脸熟",
    "完成任务",
    "代下",
    "代买",
    "内部券",
    "加微",
    "加v",
    "私聊我",
    "主页看",
    "点击链接",
    "http://",
    "https://",
    "打卡",
    "冒泡",
    "占座",
    "无内容",
    "暂无评价",
    "蹲",
    "蹲一个",
    "求链接",
    "求分享",
    "多少钱",
    "怎么卖",
    "啥牌子",
    "什么牌子",
    "求品牌",
    "求私",
    "加群",
    "裙内",
    "互赞",
    "互粉",
    "互关",
    "回关",
    "秒回",
    "交朋友",
    "优惠",
    "好物",
    "红包",
    "特价",
    "国补",
    "リンク",
    "プロフィール見て",
    "プロフ見て",
    "DMして",
    "フォロー返し",
    "相互フォロー",
    "テスト",
    "内容なし",
    "評価なし",
    "コメント稼ぎ",
    "割引",
    "良いもの",
    "お年玉",
    "特価",
    "国の補助金",
    "링크",
    "맞팔",
    "테스트",
    "내용 없음",
    "할인",
    "좋은 물건",
    "홍바오",
    "특가",
    "국가 보조금",
)
DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS = (
    "加v",
    "link in bio",
    "click link",
    "click the link",
    "check my profile",
    "see my profile",
    "visit my profile",
    "dm me",
    "message me",
    "follow me",
    "follow back",
    "follow for follow",
    "sub4sub",
    "sub for sub",
    "subscribe to my channel",
    "earn coins",
    "free coins",
    "for coins",
    "comment for points",
    "promo code",
    "coupon code",
    "discount code",
    "discount",
    "good stuff",
    "red envelope",
    "special price",
    "national subsidy",
    "whatsapp",
    "telegram",
    "n/a",
    "no content",
    "no comment",
    "nothing to say",
    "what brand",
    "brand?",
    "share link",
    "need link",
    "passing by",
    "check in",
    "enlace",
    "link en bio",
    "haz clic en el enlace",
    "mira mi perfil",
    "revisa mi perfil",
    "mándame dm",
    "mandame dm",
    "escríbeme",
    "escribeme",
    "sígueme",
    "sigueme",
    "te sigo",
    "cupón",
    "cupon",
    "código promocional",
    "codigo promocional",
    "descuento",
    "cosas buenas",
    "sobre rojo",
    "precio especial",
    "subsidio nacional",
    "primero",
    "prueba",
    "sin contenido",
    "sin comentario",
    "nada que decir",
    "cuánto cuesta",
    "cuanto cuesta",
    "precio",
    "qué marca",
    "que marca",
    "marca?",
    "pásame el link",
    "pasame el link",
    "necesito el link",
    "ลิงก์",
    "ขอลิงก์",
    "ส่งลิงก์",
    "ดูโปรไฟล์",
    "ทัก dm",
    "dm มา",
    "ติดตามกลับ",
    "ฟอลกลับ",
    "ทดสอบ",
    "ไม่มีเนื้อหา",
    "ไม่มีความคิดเห็น",
    "ปั๊มคอมเมนต์",
    "เก็บแต้ม",
    "ส่วนลด",
    "ของดี",
    "อั่งเปา",
    "ราคาพิเศษ",
    "เงินอุดหนุนจากรัฐ",
    "ราคาเท่าไหร่",
    "กี่บาท",
    "ยี่ห้ออะไร",
    "แบรนด์อะไร",
    "ผ่านมา",
    "เช็คชื่อ",
    "ทำภารกิจ",
    "लिंक",
    "लिंक दो",
    "लिंक भेजो",
    "प्रोफाइल देखें",
    "डीएम करें",
    "dm करें",
    "फॉलो बैक",
    "मुझे फॉलो करें",
    "टेस्ट",
    "कोई सामग्री नहीं",
    "कोई टिप्पणी नहीं",
    "पॉइंट कमाने",
    "सिक्के कमाने",
    "छूट",
    "अच्छी चीज़",
    "लाल लिफाफा",
    "विशेष कीमत",
    "राष्ट्रीय सब्सिडी",
    "कितने का है",
    "कीमत",
    "कौन सा ब्रांड",
    "ब्रांड क्या है",
    "बस गुजर रहा",
    "चेक इन",
    "काम पूरा",
)


def parse_fixed_term_script_group_overrides(
    raw_overrides: Any,
    fixed_terms: tuple[str, ...],
) -> tuple[tuple[str, str], ...]:
    if raw_overrides is None:
        return ()
    if not isinstance(raw_overrides, dict):
        raise CleanerConfigError("fixed_term_script_group_overrides must be an object")

    configured_terms = set(fixed_terms)
    overrides: list[tuple[str, str]] = []
    for term, script_group in raw_overrides.items():
        if not isinstance(term, str) or not term:
            raise CleanerConfigError(
                "fixed_term_script_group_overrides keys must be non-empty strings"
            )
        if term not in configured_terms:
            raise CleanerConfigError(
                "fixed_term_script_group_overrides may only classify an active fixed delete term: "
                f"{term}"
            )
        if script_group not in FIXED_TERM_SCRIPT_GROUPS:
            raise CleanerConfigError(
                "fixed_term_script_group_overrides has an unsupported script group: "
                f"{script_group}"
            )
        overrides.append((term, script_group))
    return tuple(overrides)


def default_fixed_term_script_group_overrides() -> tuple[tuple[str, str], ...]:
    data = json.loads(DEFAULT_CONFIG_PATH.read_text(encoding="utf-8"))
    fixed_terms = tuple(
        data.get("delete_contains_texts", DEFAULT_DELETE_CONTAINS_TEXTS)
    ) + tuple(
        data.get(
            "delete_contains_case_insensitive_texts",
            DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS,
        )
    )
    return parse_fixed_term_script_group_overrides(
        data.get("fixed_term_script_group_overrides"),
        fixed_terms,
    )


@dataclass(frozen=True)
class CleanerConfig:
    # Kept only as an ignored compatibility field for callers that still pass it.
    # Cleaning always resolves the locked standardized comment header instead.
    target_column: int | None = None
    target_header: str | None = STANDARDIZED_COMMENT_HEADER
    platform: str = ""
    twitter_comments_strip_https_urls_from_comment_content: bool = False
    first_data_row: int = 2
    min_trimmed_length: int = 8
    non_chinese_max_short_words: int = 2
    non_chinese_max_short_unspaced_chars: int = 4
    delete_exact_texts: tuple[str, ...] = ("该用户未填写评价内容", "此用户未填写评价内容")
    delete_contains_texts: tuple[str, ...] = DEFAULT_DELETE_CONTAINS_TEXTS
    delete_contains_case_insensitive_texts: tuple[str, ...] = DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS
    fixed_term_script_group_overrides: tuple[tuple[str, str], ...] = field(
        default_factory=default_fixed_term_script_group_overrides
    )
    delete_random_alnum_without_chinese: bool = True
    random_digit_min_length: int = 9
    random_letter_min_length: int = 10
    random_mixed_min_length: int = 10
    random_letter_max_vowel_ratio: float = 0.2
    random_letter_min_consonant_run: int = 5
    subcomment_deduplicate_headers: tuple[str, ...] = ("一级评论", "二级评论", "三级评论")
    subcomment_min_trimmed_length: int = 6
    main_comment_duplicate_keep: str = "max_likes_last_tiebreak"
    main_comment_duplicate_like_header: str = "点赞数"
    duplicate_keep: str = "last"
    export_first_sheet_csv: bool = True
    csv_encoding: str = "utf-8-sig"


@dataclass(frozen=True)
class DeletedRow:
    sheet: str
    row_number: int
    reason: str
    value: str


@dataclass(frozen=True)
class ClearedCell:
    sheet: str
    row_number: int
    column_header: str
    reason: str
    value: str


@dataclass(frozen=True)
class CleanResult:
    input_path: Path
    output_xlsx: Path
    output_csv: Path | None
    deletion_log_csv: Path
    summary_json: Path
    sheets_processed: int
    rows_deleted: int
    cells_cleared: int
    https_urls_stripped: int


def load_config(path: Path, platform: str | None = None) -> CleanerConfig:
    data = json.loads(path.read_text(encoding="utf-8"))
    if "platform_profiles" in data:
        raise CleanerConfigError(
            "platform_profiles are forbidden. All platforms must use the canonical fixed cleaning rules."
        )
    if "target_column" in data:
        raise CleanerConfigError(
            "target_column is not supported. Cleaning must resolve the standardized 评论内容 header."
        )
    configured_target_header = data.get("target_header", STANDARDIZED_COMMENT_HEADER)
    if configured_target_header != STANDARDIZED_COMMENT_HEADER:
        raise CleanerConfigError(
            "target_header must be exactly 评论内容 for the standardized cleaning workflow."
        )
    delete_exact_texts = tuple(data.get("delete_exact_texts", []))
    delete_contains_texts = tuple(
        data.get("delete_contains_texts", DEFAULT_DELETE_CONTAINS_TEXTS)
    )
    delete_contains_case_insensitive_texts = tuple(
        data.get(
            "delete_contains_case_insensitive_texts",
            DEFAULT_DELETE_CONTAINS_CASE_INSENSITIVE_TEXTS,
        )
    )
    normalized_platform = (platform or "").strip().casefold()
    config = CleanerConfig(
        target_header=STANDARDIZED_COMMENT_HEADER,
        platform=normalized_platform,
        twitter_comments_strip_https_urls_from_comment_content=bool(
            data.get("twitter_comments_strip_https_urls_from_comment_content", False)
        ),
        first_data_row=int(data.get("first_data_row", 2)),
        min_trimmed_length=int(data.get("min_trimmed_length", 8)),
        non_chinese_max_short_words=int(data.get("non_chinese_max_short_words", 2)),
        non_chinese_max_short_unspaced_chars=int(data.get("non_chinese_max_short_unspaced_chars", 4)),
        delete_exact_texts=delete_exact_texts,
        delete_contains_texts=delete_contains_texts,
        delete_contains_case_insensitive_texts=delete_contains_case_insensitive_texts,
        fixed_term_script_group_overrides=parse_fixed_term_script_group_overrides(
            data.get("fixed_term_script_group_overrides"),
            delete_contains_texts + delete_contains_case_insensitive_texts,
        ),
        delete_random_alnum_without_chinese=bool(data.get("delete_random_alnum_without_chinese", True)),
        random_digit_min_length=int(data.get("random_digit_min_length", 9)),
        random_letter_min_length=int(data.get("random_letter_min_length", 10)),
        random_mixed_min_length=int(data.get("random_mixed_min_length", 10)),
        random_letter_max_vowel_ratio=float(data.get("random_letter_max_vowel_ratio", 0.2)),
        random_letter_min_consonant_run=int(data.get("random_letter_min_consonant_run", 5)),
        subcomment_deduplicate_headers=tuple(
            data.get("subcomment_deduplicate_headers", ["一级评论", "二级评论", "三级评论"])
        ),
        subcomment_min_trimmed_length=int(data.get("subcomment_min_trimmed_length", 6)),
        main_comment_duplicate_keep=str(data.get("main_comment_duplicate_keep", "max_likes_last_tiebreak")),
        main_comment_duplicate_like_header=str(data.get("main_comment_duplicate_like_header", "点赞数")),
        duplicate_keep=str(data.get("duplicate_keep", "last")),
        export_first_sheet_csv=bool(data.get("export_first_sheet_csv", True)),
        csv_encoding=str(data.get("csv_encoding", "utf-8-sig")),
    )
    return config


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split())


def count_non_chinese_words(comment: str) -> int | None:
    if not comment or CHINESE_CHAR_PATTERN.search(comment):
        return None

    has_latin_letter = LATIN_LETTER_PATTERN.search(comment) is not None
    has_whitespace = any(char.isspace() for char in comment)
    if not has_latin_letter and not has_whitespace:
        return None

    tokens = 0
    in_token = False
    for char in comment:
        category = unicodedata.category(char)
        is_token_char = category[0] in {"L", "N", "M"} and not CHINESE_CHAR_PATTERN.fullmatch(char)
        if is_token_char:
            if not in_token:
                tokens += 1
            in_token = True
        else:
            in_token = False
    return tokens


def should_delete_for_length(comment: str, config: CleanerConfig) -> str | None:
    if fixed_term_script_group(comment) == "chinese":
        if len(comment) < config.min_trimmed_length:
            return f"评论长度小于 {config.min_trimmed_length}"
        return None

    if re.fullmatch(r"\d+", comment):
        if len(comment) < config.min_trimmed_length:
            return f"评论长度小于 {config.min_trimmed_length}"
        return None

    word_count = count_non_chinese_words(comment)
    if word_count is not None:
        if word_count <= config.non_chinese_max_short_words:
            return f"非中文评论词数小于等于 {config.non_chinese_max_short_words}"
        return None

    if len(comment) <= config.non_chinese_max_short_unspaced_chars:
        return f"非中文无空格短文本长度小于等于 {config.non_chinese_max_short_unspaced_chars}"
    return None


def max_consonant_run(value: str) -> int:
    max_run = 0
    current_run = 0
    for char in value.lower():
        if char.isalpha() and char not in "aeiou":
            current_run += 1
            max_run = max(max_run, current_run)
        else:
            current_run = 0
    return max_run


def is_random_alnum_without_chinese(comment: str, config: CleanerConfig) -> bool:
    if not comment or CHINESE_CHAR_PATTERN.search(comment):
        return False

    for token in ALNUM_TOKEN_PATTERN.findall(comment):
        has_alpha = any(char.isalpha() for char in token)
        has_digit = any(char.isdigit() for char in token)
        if has_digit and not has_alpha and len(token) >= config.random_digit_min_length:
            return True
        if has_digit and has_alpha and len(token) >= config.random_mixed_min_length:
            return True
        if has_alpha and not has_digit and len(token) >= config.random_letter_min_length:
            vowel_count = sum(1 for char in token.lower() if char in "aeiou")
            vowel_ratio = vowel_count / len(token)
            if vowel_ratio <= config.random_letter_max_vowel_ratio:
                return True
            if max_consonant_run(token) >= config.random_letter_min_consonant_run:
                return True
    return False


def resolve_target_column(sheet: Worksheet, config: CleanerConfig) -> int:
    if config.target_header != STANDARDIZED_COMMENT_HEADER:
        raise ValueError(
            "Cleaning requires the standardized 评论内容 header; numeric target_column fallback is disabled."
        )

    header_row = max(1, config.first_data_row - 1)
    target_key = normalize_header(STANDARDIZED_COMMENT_HEADER)
    headers = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, max_col=sheet.max_column, values_only=True),
        (),
    )
    matches = [
        column_index
        for column_index, header in enumerate(headers, start=1)
        if normalize_header(header) == target_key
    ]
    if not matches:
        raise ValueError(f"未找到评论列表头: {STANDARDIZED_COMMENT_HEADER}")
    if len(matches) > 1:
        raise ValueError(f"评论列表头重复: {STANDARDIZED_COMMENT_HEADER}")
    return matches[0]


def resolve_optional_header_columns(sheet: Worksheet, config: CleanerConfig, headers: tuple[str, ...]) -> list[tuple[int, str]]:
    header_row = max(1, config.first_data_row - 1)
    wanted = {normalize_header(header): header for header in headers}
    if not wanted:
        return []

    sheet_headers = next(
        sheet.iter_rows(min_row=header_row, max_row=header_row, max_col=sheet.max_column, values_only=True),
        (),
    )
    columns: list[tuple[int, str]] = []
    for column_index, header in enumerate(sheet_headers, start=1):
        key = normalize_header(header)
        if key in wanted:
            columns.append((column_index, wanted[key]))
    return columns


def resolve_optional_single_header_column(sheet: Worksheet, config: CleanerConfig, header: str) -> int | None:
    columns = resolve_optional_header_columns(sheet, config, (header,))
    if len(columns) > 1:
        raise ValueError(f"重复列标题: {header}")
    return columns[0][0] if columns else None


def parse_like_count(value: Any) -> Decimal:
    text = normalize_cell(value)
    if not text or not LIKE_COUNT_PATTERN.fullmatch(text):
        return Decimal(0)
    try:
        return Decimal(text)
    except InvalidOperation:
        return Decimal(0)


def fixed_term_script_group(
    text: str,
    *,
    overrides: tuple[tuple[str, str], ...] = (),
) -> str:
    for term, script_group in overrides:
        if term == text:
            return script_group
    normalized = text.casefold()
    if normalized in {"http://", "https://"}:
        return "neutral"
    if JAPANESE_KANA_PATTERN.search(text):
        return "japanese"
    if KOREAN_HANGUL_PATTERN.search(text):
        return "korean"
    if THAI_PATTERN.search(text):
        return "thai"
    if DEVANAGARI_PATTERN.search(text):
        return "hindi"
    if CHINESE_CHAR_PATTERN.search(text):
        return "chinese"

    letters = [character for character in text if unicodedata.category(character).startswith("L")]
    if letters and all("LATIN" in unicodedata.name(character, "") for character in letters):
        return "latin"
    return "neutral"


def contains_configured_fixed_term(
    comment: str,
    term: str,
    *,
    case_sensitive: bool,
    term_script_group_overrides: tuple[tuple[str, str], ...] = (),
) -> bool:
    term_group = fixed_term_script_group(term, overrides=term_script_group_overrides)
    comment_group = fixed_term_script_group(comment)
    if term_group != "neutral" and term_group != comment_group:
        return False

    if term_group == "latin":
        prefix = r"(?<!\w)" if term[0].isalnum() else ""
        suffix = r"(?!\w)" if term[-1].isalnum() else ""
        flags = 0 if case_sensitive else re.IGNORECASE
        return re.search(f"{prefix}{re.escape(term)}{suffix}", comment, flags) is not None

    if case_sensitive:
        return term in comment
    return term.casefold() in comment.casefold()


def should_delete_comment(
    comment: str,
    seen_comments: set[str],
    config: CleanerConfig,
    clean_words: tuple[str, ...],
) -> str | None:
    length_reason = should_delete_for_length(comment, config)
    if length_reason:
        return length_reason

    if comment in config.delete_exact_texts:
        return "评论等于占位文案"

    for clean_word in clean_words:
        if clean_word and clean_word in comment:
            return f"评论包含清理词: {clean_word}"

    for text in config.delete_contains_texts:
        if text and contains_configured_fixed_term(
            comment,
            text,
            case_sensitive=True,
            term_script_group_overrides=config.fixed_term_script_group_overrides,
        ):
            return f"评论包含固定删除词: {text}"

    for text in config.delete_contains_case_insensitive_texts:
        if text and contains_configured_fixed_term(
            comment,
            text,
            case_sensitive=False,
            term_script_group_overrides=config.fixed_term_script_group_overrides,
        ):
            return f"评论包含固定删除词: {text}"

    if config.delete_random_alnum_without_chinese and is_random_alnum_without_chinese(comment, config):
        return "评论为无中文随机英文/数字堆砌"

    if comment in seen_comments:
        return "同一工作表内重复评论"

    return None


def iter_row_numbers(sheet: Worksheet, config: CleanerConfig) -> range:
    if config.duplicate_keep == "last":
        return range(sheet.max_row, config.first_data_row - 1, -1)
    if config.duplicate_keep == "first":
        return range(config.first_data_row, sheet.max_row + 1)
    raise ValueError("duplicate_keep 只能是 first 或 last")


def collect_main_comment_deletions(
    sheet: Worksheet,
    config: CleanerConfig,
    clean_words: tuple[str, ...],
) -> list[DeletedRow]:
    if config.main_comment_duplicate_keep != "max_likes_last_tiebreak":
        raise ValueError("main_comment_duplicate_keep 只能是 max_likes_last_tiebreak")

    target_column = resolve_target_column(sheet, config)
    like_column = resolve_optional_single_header_column(sheet, config, config.main_comment_duplicate_like_header)
    candidates: dict[str, list[tuple[int, Decimal]]] = {}
    pending_deletions: list[DeletedRow] = []

    for row_number in range(config.first_data_row, sheet.max_row + 1):
        comment = normalize_cell(sheet.cell(row=row_number, column=target_column).value)
        reason = should_delete_comment(comment, set(), config, clean_words)
        if reason:
            pending_deletions.append(
                DeletedRow(
                    sheet=sheet.title,
                    row_number=row_number,
                    reason=reason,
                    value=comment,
                )
            )
            continue

        like_count = parse_like_count(sheet.cell(row=row_number, column=like_column).value) if like_column else Decimal(0)
        candidates.setdefault(comment, []).append((row_number, like_count))

    for comment, duplicate_rows in candidates.items():
        if len(duplicate_rows) < 2:
            continue
        retained_row, _ = max(duplicate_rows, key=lambda row: (row[1], row[0]))
        for row_number, _ in duplicate_rows:
            if row_number == retained_row:
                continue
            pending_deletions.append(
                DeletedRow(
                    sheet=sheet.title,
                    row_number=row_number,
                    reason="同一工作表内重复评论（保留点赞数最高，点赞相同保留最后一条）",
                    value=comment,
                )
            )

    return pending_deletions


def iter_cell_positions(sheet: Worksheet, config: CleanerConfig, columns: list[tuple[int, str]]) -> list[tuple[int, int, str]]:
    if config.duplicate_keep == "last":
        row_numbers = range(sheet.max_row, config.first_data_row - 1, -1)
        ordered_columns = list(reversed(columns))
    elif config.duplicate_keep == "first":
        row_numbers = range(config.first_data_row, sheet.max_row + 1)
        ordered_columns = columns
    else:
        raise ValueError("duplicate_keep 只能是 first 或 last")

    return [(row_number, column_index, header) for row_number in row_numbers for column_index, header in ordered_columns]


def clear_duplicate_subcomments(sheet: Worksheet, config: CleanerConfig) -> list[ClearedCell]:
    columns = resolve_optional_header_columns(sheet, config, config.subcomment_deduplicate_headers)
    if not columns:
        return []

    seen_values: set[str] = set()
    cleared_cells: list[ClearedCell] = []

    for row_number, column_index, header in iter_cell_positions(sheet, config, columns):
        cell = sheet.cell(row=row_number, column=column_index)
        value = normalize_cell(cell.value)
        if not value:
            continue
        if len(value) < config.subcomment_min_trimmed_length:
            cell.value = None
            cleared_cells.append(
                ClearedCell(
                    sheet=sheet.title,
                    row_number=row_number,
                    column_header=header,
                    reason=f"子评论长度小于 {config.subcomment_min_trimmed_length}: {header}",
                    value=value,
                )
            )
            continue
        if value in seen_values:
            cell.value = None
            cleared_cells.append(
                ClearedCell(
                    sheet=sheet.title,
                    row_number=row_number,
                    column_header=header,
                    reason=f"同一工作表内重复子评论: {header}",
                    value=value,
                )
            )
        else:
            seen_values.add(value)

    return sorted(cleared_cells, key=lambda cell: (cell.row_number, cell.column_header))


def strip_https_urls_from_twitter_comment_content(
    sheet: Worksheet,
    config: CleanerConfig,
) -> int:
    """Remove inline HTTPS URLs from X comment text before normal fixed-rule cleaning."""
    if (
        config.platform != "twitter-comments"
        or not config.twitter_comments_strip_https_urls_from_comment_content
    ):
        return 0

    target_column = resolve_target_column(sheet, config)
    stripped_url_count = 0
    for row_number in range(config.first_data_row, sheet.max_row + 1):
        cell = sheet.cell(row=row_number, column=target_column)
        comment = normalize_cell(cell.value)
        stripped_comment, replacements = HTTPS_URL_PATTERN.subn("", comment)
        if replacements:
            cell.value = stripped_comment.strip()
            stripped_url_count += replacements
    return stripped_url_count


def clean_sheet(sheet: Worksheet, config: CleanerConfig, clean_words: tuple[str, ...]) -> tuple[list[DeletedRow], list[ClearedCell]]:
    deleted: list[DeletedRow] = []
    pending_deletions = collect_main_comment_deletions(sheet, config, clean_words)

    for deleted_row in sorted(pending_deletions, key=lambda row: row.row_number, reverse=True):
        sheet.delete_rows(deleted_row.row_number, 1)
        deleted.append(deleted_row)

    cleared_cells = clear_duplicate_subcomments(sheet, config)

    return deleted, cleared_cells


def make_output_paths(input_path: Path, output_dir: Path | None, output_path: Path | None = None) -> tuple[Path, Path, Path, Path]:
    if output_path:
        output_xlsx = output_path.resolve()
        return (
            output_xlsx,
            output_xlsx.with_suffix(".csv"),
            output_xlsx.with_suffix(".deletions.csv"),
            output_xlsx.with_suffix(".summary.json"),
        )

    timestamp = beijing_date_text()
    parent = output_dir if output_dir else input_path.parent
    stem = f"{timestamp}_{input_path.stem}"
    return (
        parent / f"{stem}.cleaned.xlsx",
        parent / f"{stem}.cleaned.csv",
        parent / f"{stem}.deletions.csv",
        parent / f"{stem}.summary.json",
    )


def write_first_sheet_csv(sheet: Worksheet, path: Path, encoding: str) -> None:
    with path.open("w", newline="", encoding=encoding) as csv_file:
        writer = csv.writer(csv_file)
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=sheet.max_column, values_only=True):
            writer.writerow(["" if value is None else value for value in row])


def write_deletion_log(path: Path, deleted_rows: list[DeletedRow], cleared_cells: list[ClearedCell], encoding: str) -> None:
    with path.open("w", newline="", encoding=encoding) as csv_file:
        writer = csv.DictWriter(csv_file, fieldnames=["action", "sheet", "row_number", "column_header", "reason", "value"])
        writer.writeheader()
        for row in deleted_rows:
            writer.writerow(
                {
                    "action": "delete_row",
                    "sheet": row.sheet,
                    "row_number": row.row_number,
                    "column_header": "",
                    "reason": row.reason,
                    "value": row.value,
                }
            )
        for cell in cleared_cells:
            writer.writerow(
                {
                    "action": "clear_cell",
                    "sheet": cell.sheet,
                    "row_number": cell.row_number,
                    "column_header": cell.column_header,
                    "reason": cell.reason,
                    "value": cell.value,
                }
            )


def clean_workbook(
    input_path: Path,
    config: CleanerConfig,
    clean_words: tuple[str, ...],
    output_dir: Path | None = None,
    output_path: Path | None = None,
    *,
    overwrite: bool = True,
    overwrite_confirmations: tuple[Path, ...] | list[Path] | None = None,
) -> CleanResult:
    input_path = input_path.resolve()
    if not is_supported_input_path(input_path):
        raise ValueError(unsupported_input_message())
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    output_xlsx, output_csv, deletion_log_csv, summary_json = make_output_paths(input_path, output_dir, output_path)
    derived_outputs = (
        output_xlsx.resolve(),
        output_csv.resolve(),
        deletion_log_csv.resolve(),
        summary_json.resolve(),
    )
    if input_path in derived_outputs:
        if output_xlsx.resolve() == input_path:
            raise ValueError("Output path must be a new workbook path, not the input file.")
        raise ValueError("A derived output path would overwrite the input file.")
    outputs_to_create = [output_xlsx, deletion_log_csv, summary_json]
    if config.export_first_sheet_csv:
        outputs_to_create.append(output_csv)
    refuse_finalized_audit_artifact_restoration(
        output_xlsx,
        deletion_log_csv,
        summary_json,
    )
    ensure_output_paths_safe(
        [input_path],
        outputs_to_create,
        overwrite=overwrite,
        overwrite_confirmations=overwrite_confirmations,
    )

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook_for_processing(input_path, read_only=False, data_only=False)
    deleted_rows: list[DeletedRow] = []
    cleared_cells: list[ClearedCell] = []
    https_urls_stripped = 0

    try:
        for sheet in workbook.worksheets:
            https_urls_stripped += strip_https_urls_from_twitter_comment_content(sheet, config)
            sheet_deleted_rows, sheet_cleared_cells = clean_sheet(sheet, config, clean_words)
            deleted_rows.extend(sheet_deleted_rows)
            cleared_cells.extend(sheet_cleared_cells)
    except BaseException:
        workbook.close()
        raise

    actual_output_csv: Path | None = None
    if config.export_first_sheet_csv and workbook.worksheets:
        actual_output_csv = output_csv

    summary = {
        "input_path": str(input_path),
        "output_xlsx": str(output_xlsx),
        "output_csv": str(actual_output_csv) if actual_output_csv else None,
        "deletion_log_csv": str(deletion_log_csv),
        "sheets_processed": len(workbook.worksheets),
        "rows_deleted": len(deleted_rows),
        "cells_cleared": len(cleared_cells),
        "platform": config.platform,
        "https_urls_stripped": https_urls_stripped,
        "target_header": config.target_header,
        "first_data_row": config.first_data_row,
        "min_trimmed_length": config.min_trimmed_length,
        "non_chinese_max_short_words": config.non_chinese_max_short_words,
        "non_chinese_max_short_unspaced_chars": config.non_chinese_max_short_unspaced_chars,
        "clean_words": list(clean_words),
        "delete_contains_texts": list(config.delete_contains_texts),
        "delete_contains_case_insensitive_texts": list(config.delete_contains_case_insensitive_texts),
        "delete_random_alnum_without_chinese": config.delete_random_alnum_without_chinese,
        "random_digit_min_length": config.random_digit_min_length,
        "random_letter_min_length": config.random_letter_min_length,
        "random_mixed_min_length": config.random_mixed_min_length,
        "random_letter_max_vowel_ratio": config.random_letter_max_vowel_ratio,
        "random_letter_min_consonant_run": config.random_letter_min_consonant_run,
        "main_comment_duplicate_keep": config.main_comment_duplicate_keep,
        "main_comment_duplicate_like_header": config.main_comment_duplicate_like_header,
        "duplicate_keep": config.duplicate_keep,
        "subcomment_deduplicate_headers": list(config.subcomment_deduplicate_headers),
        "subcomment_min_trimmed_length": config.subcomment_min_trimmed_length,
    }
    try:
        with ExitStack() as stack:
            staged_xlsx = stack.enter_context(atomic_output_path(output_xlsx))
            staged_deletion_log = stack.enter_context(atomic_output_path(deletion_log_csv))
            staged_summary = stack.enter_context(atomic_output_path(summary_json))
            staged_csv = (
                stack.enter_context(atomic_output_path(output_csv))
                if actual_output_csv is not None
                else None
            )
            workbook.save(staged_xlsx)
            if staged_csv is not None:
                write_first_sheet_csv(workbook.worksheets[0], staged_csv, config.csv_encoding)
            write_deletion_log(
                staged_deletion_log,
                deleted_rows,
                cleared_cells,
                config.csv_encoding,
            )
            staged_summary.write_text(
                json.dumps(summary, ensure_ascii=False, indent=2),
                encoding="utf-8",
            )
    finally:
        workbook.close()

    return CleanResult(
        input_path=input_path,
        output_xlsx=output_xlsx,
        output_csv=actual_output_csv,
        deletion_log_csv=deletion_log_csv,
        summary_json=summary_json,
        sheets_processed=len(workbook.worksheets),
        rows_deleted=len(deleted_rows),
        cells_cleared=len(cleared_cells),
        https_urls_stripped=https_urls_stripped,
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="清洗抓取或导出的用户评论 Excel/CSV 表")
    parser.add_argument("input_path", type=Path, help="需要清洗的 .xlsx/.xlsm/.csv 文件")
    parser.add_argument(
        "--config",
        type=Path,
        default=DEFAULT_CONFIG_PATH,
        help=(
            "仅接受 Skill 内置的正式清洗配置；外部或临时配置会被拒绝。"
        ),
    )
    parser.add_argument(
        "--clean-word",
        action="append",
        default=[],
        help="额外清理词，可重复传入。例如 --clean-word KOL清理词1 --clean-word KOL清理词2",
    )
    parser.add_argument(
        "--target-header",
        required=True,
        help="必须显式指定标准化后的评论列表头：评论内容",
    )
    parser.add_argument(
        "--platform",
        required=True,
        help="本轮已确认的平台名称；用于记录已确认的执行上下文，不改变固定清洗规则。",
    )
    parser.add_argument(
        "--output-dir",
        type=Path,
        default=None,
        help="兼容旧程序调用的默认目录；CLI 仍必须传入 --output",
    )
    parser.add_argument(
        "--output",
        type=Path,
        required=True,
        help="命名确认后传入的输出 .xlsx 文件路径",
    )
    add_confirmed_overwrite_arguments(
        parser,
        overwrite_help="仅在用户明确确认每个既有输出路径后才允许覆盖。",
    )
    args = parser.parse_args(argv)
    args.target_header = args.target_header.strip()
    if args.target_header != STANDARDIZED_COMMENT_HEADER:
        parser.error(
            f"--target-header must be exactly {STANDARDIZED_COMMENT_HEADER!r} for the standardized workflow"
        )
    args.platform = args.platform.strip()
    if not args.platform:
        parser.error("--platform must not be blank")
    return args


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    config = load_config(
        require_canonical_cleaner_config_path(args.config),
        platform=args.platform,
    )
    config = replace(config, target_header=args.target_header)
    clean_words = tuple(word.strip() for word in args.clean_word if word and word.strip())
    result = clean_workbook(
        args.input_path.resolve(),
        config,
        clean_words,
        args.output_dir,
        args.output,
        overwrite=args.overwrite,
        overwrite_confirmations=tuple(args.confirm_overwrite),
    )

    print(f"处理完成: {result.input_path}")
    print(f"工作表数量: {result.sheets_processed}")
    print(f"删除行数: {result.rows_deleted}")
    print(f"清空单元格数: {result.cells_cleared}")
    print(f"X 评论评论内容列已移除 https URL 数: {result.https_urls_stripped}")
    print(f"清洗后 xlsx: {result.output_xlsx}")
    if result.output_csv:
        print(f"首个工作表 csv: {result.output_csv}")
    print(f"删除日志: {result.deletion_log_csv}")
    print(f"摘要: {result.summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
