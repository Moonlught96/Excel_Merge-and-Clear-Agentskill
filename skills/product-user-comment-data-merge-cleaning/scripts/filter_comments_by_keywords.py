from __future__ import annotations

import argparse
import json
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl.worksheet.worksheet import Worksheet

try:
    from tools.csv_excel_compat import (
        is_supported_input_path,
        load_workbook_for_processing,
        unsupported_input_message,
    )
    from tools.output_path_safety import atomic_output_path, ensure_output_paths_safe
except ModuleNotFoundError:
    from csv_excel_compat import (
        is_supported_input_path,
        load_workbook_for_processing,
        unsupported_input_message,
    )
    from output_path_safety import atomic_output_path, ensure_output_paths_safe


DEFAULT_TARGET_HEADER = "评论内容"


class KeywordFilterConfigError(ValueError):
    pass


class KeywordTargetHeaderError(ValueError):
    pass


@dataclass(frozen=True)
class KeywordFilterSheetSummary:
    sheet_name: str
    input_rows: int
    kept_rows: int
    deleted_rows: int


@dataclass(frozen=True)
class KeywordFilterResult:
    input_path: Path
    output_xlsx: Path
    summary_json: Path
    target_header: str
    keep_keywords: tuple[str, ...]
    sheets_processed: int
    input_rows: int
    kept_rows: int
    deleted_rows: int


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split())


def normalize_keep_keywords(keywords: list[str] | tuple[str, ...]) -> tuple[str, ...]:
    if not isinstance(keywords, (list, tuple)):
        raise KeywordFilterConfigError("Keep keywords must be a list or tuple of strings")

    normalized: list[str] = []
    seen: set[str] = set()
    for keyword in keywords:
        if not isinstance(keyword, str):
            raise KeywordFilterConfigError("Keep keywords must be strings")
        value = keyword.strip()
        if not value:
            continue
        key = value.casefold()
        if key not in seen:
            seen.add(key)
            normalized.append(value)

    if not normalized:
        raise KeywordFilterConfigError("At least one nonblank keep keyword is required")
    return tuple(normalized)


def resolve_target_column(sheet: Worksheet, target_header: str) -> int:
    normalized_target = normalize_header(target_header)
    if not normalized_target:
        raise KeywordFilterConfigError("Target header must be a nonblank string")

    matches = [
        column_index
        for column_index in range(1, sheet.max_column + 1)
        if normalize_header(sheet.cell(1, column_index).value) == normalized_target
    ]
    if len(matches) != 1:
        raise KeywordTargetHeaderError(
            f"Worksheet {sheet.title!r} must contain exactly one target header: {target_header!r}"
        )
    return matches[0]


def should_keep_comment(value: Any, folded_keywords: tuple[str, ...]) -> bool:
    if value is None:
        return False
    comment = str(value).casefold()
    return any(keyword in comment for keyword in folded_keywords)


def filter_sheet(
    sheet: Worksheet,
    target_header: str,
    folded_keywords: tuple[str, ...],
) -> KeywordFilterSheetSummary:
    target_column = resolve_target_column(sheet, target_header)
    input_rows = max(sheet.max_row - 1, 0)
    deleted_rows: list[int] = []
    for row_index in range(2, sheet.max_row + 1):
        if not should_keep_comment(sheet.cell(row_index, target_column).value, folded_keywords):
            deleted_rows.append(row_index)

    for row_index in reversed(deleted_rows):
        sheet.delete_rows(row_index, 1)

    return KeywordFilterSheetSummary(
        sheet_name=sheet.title,
        input_rows=input_rows,
        kept_rows=input_rows - len(deleted_rows),
        deleted_rows=len(deleted_rows),
    )


def make_output_paths(
    input_path: Path,
    output_dir: Path | None,
    output_path: Path | None,
) -> tuple[Path, Path]:
    if output_path is not None:
        output_xlsx = output_path.resolve()
    else:
        parent = output_dir.resolve() if output_dir else input_path.parent
        output_xlsx = parent / f"{input_path.stem}.keyword-filtered.xlsx"
    return output_xlsx, output_xlsx.with_suffix(".keyword-filter.summary.json")


def filter_workbook(
    input_path: Path,
    keep_keywords: list[str] | tuple[str, ...],
    output_dir: Path | None = None,
    output_path: Path | None = None,
    *,
    target_header: str = DEFAULT_TARGET_HEADER,
    overwrite: bool = False,
) -> KeywordFilterResult:
    input_path = input_path.resolve()
    if not is_supported_input_path(input_path):
        raise ValueError(unsupported_input_message())
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    normalized_keywords = normalize_keep_keywords(keep_keywords)
    folded_keywords = tuple(keyword.casefold() for keyword in normalized_keywords)
    output_xlsx, summary_json = make_output_paths(input_path, output_dir, output_path)
    ensure_output_paths_safe(
        [input_path],
        [output_xlsx, summary_json],
        overwrite=overwrite,
    )

    workbook = load_workbook_for_processing(
        input_path,
        read_only=False,
        data_only=False,
    )
    try:
        sheet_summaries = [
            filter_sheet(sheet, target_header, folded_keywords)
            for sheet in workbook.worksheets
        ]
        with atomic_output_path(output_xlsx) as staged_output:
            workbook.save(staged_output)
    finally:
        workbook.close()

    summary = {
        "input_path": str(input_path),
        "output_xlsx": str(output_xlsx),
        "target_header": target_header,
        "keep_keywords": list(normalized_keywords),
        "sheets_processed": len(sheet_summaries),
        "input_rows": sum(sheet.input_rows for sheet in sheet_summaries),
        "kept_rows": sum(sheet.kept_rows for sheet in sheet_summaries),
        "deleted_rows": sum(sheet.deleted_rows for sheet in sheet_summaries),
        "sheets": [
            {
                "sheet_name": sheet.sheet_name,
                "input_rows": sheet.input_rows,
                "kept_rows": sheet.kept_rows,
                "deleted_rows": sheet.deleted_rows,
            }
            for sheet in sheet_summaries
        ],
    }
    with atomic_output_path(summary_json) as staged_summary:
        staged_summary.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return KeywordFilterResult(
        input_path=input_path,
        output_xlsx=output_xlsx,
        summary_json=summary_json,
        target_header=target_header,
        keep_keywords=normalized_keywords,
        sheets_processed=len(sheet_summaries),
        input_rows=summary["input_rows"],
        kept_rows=summary["kept_rows"],
        deleted_rows=summary["deleted_rows"],
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Keep only standardized comment rows containing confirmed literal keywords."
    )
    parser.add_argument("input_path", type=Path, help="Input standardized .xlsx/.xlsm/.csv file")
    parser.add_argument(
        "--keep-keyword",
        action="append",
        default=[],
        help="Confirmed literal keep keyword; repeat for each keyword.",
    )
    parser.add_argument(
        "--target-header",
        default=DEFAULT_TARGET_HEADER,
        help=f"Exact comment header, default: {DEFAULT_TARGET_HEADER}",
    )
    parser.add_argument("--output-dir", type=Path, default=None, help="Output directory")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path")
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Use only after explicit confirmation of the exact existing output path.",
    )
    return parser.parse_args(argv)


def main() -> int:
    args = parse_args()
    result = filter_workbook(
        args.input_path,
        args.keep_keyword,
        output_dir=args.output_dir,
        output_path=args.output,
        target_header=args.target_header,
        overwrite=args.overwrite,
    )
    print(f"Keyword-filtered rows kept: {result.kept_rows}")
    print(f"Keyword-filtered rows deleted: {result.deleted_rows}")
    print(f"Output xlsx: {result.output_xlsx}")
    print(f"Summary: {result.summary_json}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
