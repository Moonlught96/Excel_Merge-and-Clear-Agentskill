from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from tools.csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message
except ModuleNotFoundError:
    from csv_excel_compat import is_supported_input_path, load_workbook_for_processing, unsupported_input_message


TARGET_HEADER_ALIASES = ("content", "评论内容")
REPLY_PREFIX_PATTERN = re.compile(r"^\s*回复\s*@(?P<target>[^:：]+?)\s*[:：]\s*(?P<body>.*)$", re.DOTALL)


class TargetHeaderNotFoundError(ValueError):
    pass


class DuplicateTargetHeaderError(ValueError):
    pass


class OutputPathConflictError(ValueError):
    pass


@dataclass(frozen=True)
class SheetStripSummary:
    sheet_name: str
    target_headers: tuple[str, ...]
    target_column: int
    input_rows: int
    output_rows: int
    reply_prefixes_stripped: int


@dataclass(frozen=True)
class StripBilibiliReplyPrefixesResult:
    input_path: Path
    output_xlsx: Path
    summary_json: Path
    sheets_processed: int
    input_rows: int
    output_rows: int
    reply_prefixes_stripped: int


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split()).casefold()


def value_for_json(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return value


def strip_reply_prefix(value: Any) -> tuple[Any, bool]:
    if not isinstance(value, str):
        return value, False

    match = REPLY_PREFIX_PATTERN.match(value)
    if not match:
        return value, False
    return match.group("body"), True


def find_target_column(headers: tuple[Any, ...], target_aliases: tuple[str, ...]) -> tuple[int, tuple[str, ...]]:
    alias_keys = {normalize_header(alias) for alias in target_aliases}
    matches: list[tuple[int, str]] = []
    for column_index, header in enumerate(headers, start=1):
        if normalize_header(header) in alias_keys:
            matches.append((column_index, "" if header is None else str(header)))

    if not matches:
        available_headers = [str(header) for header in headers if normalize_header(header)]
        raise TargetHeaderNotFoundError(
            f"Target content header not found. Accepted aliases: {list(target_aliases)}. "
            f"Available headers: {available_headers}"
        )
    if len(matches) > 1:
        raise DuplicateTargetHeaderError(f"Multiple target content headers found: {[header for _, header in matches]}")

    return matches[0][0], (matches[0][1],)


def make_output_paths(input_path: Path, output_dir: Path | None) -> tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d")
    parent = output_dir if output_dir else input_path.parent
    stem = f"{timestamp}_{input_path.stem}.reply-prefix-stripped"
    output_xlsx = parent / f"{stem}.xlsx"
    return output_xlsx, output_xlsx.with_suffix(".reply-prefix-stripped.summary.json")


def strip_sheet_reply_prefixes(
    source_sheet: Worksheet,
    output_sheet: Worksheet,
    target_aliases: tuple[str, ...],
) -> SheetStripSummary:
    rows = source_sheet.iter_rows(values_only=False)
    try:
        header_cells = next(rows)
    except StopIteration:
        raise TargetHeaderNotFoundError(f"Worksheet has no header row: {source_sheet.title}")

    header = tuple(cell.value for cell in header_cells)
    target_column, target_headers = find_target_column(header, target_aliases)
    output_sheet.append(list(header))

    input_rows = 0
    output_rows = 0
    reply_prefixes_stripped = 0
    for row_cells in rows:
        input_rows += 1
        row = [cell.value for cell in row_cells]
        if target_column - 1 < len(row):
            stripped_value, changed = strip_reply_prefix(row[target_column - 1])
            if changed:
                row[target_column - 1] = stripped_value
                reply_prefixes_stripped += 1
        output_sheet.append(row)
        output_row_number = output_sheet.max_row
        for column_index, source_cell in enumerate(row_cells, start=1):
            if source_cell.data_type == "s" and isinstance(row[column_index - 1], str):
                output_sheet.cell(row=output_row_number, column=column_index).data_type = "s"
        output_rows += 1

    return SheetStripSummary(
        sheet_name=source_sheet.title,
        target_headers=target_headers,
        target_column=target_column,
        input_rows=input_rows,
        output_rows=output_rows,
        reply_prefixes_stripped=reply_prefixes_stripped,
    )


def strip_bilibili_reply_prefixes(
    input_path: Path,
    output_path: Path | None = None,
    *,
    output_dir: Path | None = None,
    target_aliases: tuple[str, ...] = TARGET_HEADER_ALIASES,
) -> StripBilibiliReplyPrefixesResult:
    input_path = input_path.resolve()
    if not is_supported_input_path(input_path):
        raise ValueError(unsupported_input_message())
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    if output_path is None:
        output_xlsx, summary_json = make_output_paths(input_path, output_dir)
    else:
        output_xlsx = output_path.resolve()
        summary_json = output_xlsx.with_suffix(".reply-prefix-stripped.summary.json")

    if output_xlsx.resolve() == input_path:
        raise OutputPathConflictError("Output path must be a new workbook path, not the input file.")

    output_xlsx.parent.mkdir(parents=True, exist_ok=True)

    input_workbook = load_workbook_for_processing(input_path, read_only=True, data_only=False)
    output_workbook = Workbook()

    sheet_summaries: list[SheetStripSummary] = []
    for sheet_index, source_sheet in enumerate(input_workbook.worksheets):
        output_sheet = output_workbook.active if sheet_index == 0 else output_workbook.create_sheet()
        output_sheet.title = source_sheet.title
        sheet_summaries.append(strip_sheet_reply_prefixes(source_sheet, output_sheet, target_aliases))

    output_workbook.save(output_xlsx)
    input_workbook.close()
    output_workbook.close()

    summary = {
        "input_path": str(input_path),
        "output_xlsx": str(output_xlsx),
        "sheets_processed": len(sheet_summaries),
        "input_rows": sum(sheet.input_rows for sheet in sheet_summaries),
        "output_rows": sum(sheet.output_rows for sheet in sheet_summaries),
        "reply_prefixes_stripped": sum(sheet.reply_prefixes_stripped for sheet in sheet_summaries),
        "target_aliases": list(target_aliases),
        "reply_prefix_pattern": REPLY_PREFIX_PATTERN.pattern,
        "rows_moved": 0,
        "columns_added": 0,
        "sheets": [
            {
                "sheet_name": sheet.sheet_name,
                "target_headers": [value_for_json(header) for header in sheet.target_headers],
                "target_column": sheet.target_column,
                "input_rows": sheet.input_rows,
                "output_rows": sheet.output_rows,
                "reply_prefixes_stripped": sheet.reply_prefixes_stripped,
            }
            for sheet in sheet_summaries
        ],
    }
    summary_json.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    return StripBilibiliReplyPrefixesResult(
        input_path=input_path,
        output_xlsx=output_xlsx,
        summary_json=summary_json,
        sheets_processed=len(sheet_summaries),
        input_rows=sum(sheet.input_rows for sheet in sheet_summaries),
        output_rows=sum(sheet.output_rows for sheet in sheet_summaries),
        reply_prefixes_stripped=sum(sheet.reply_prefixes_stripped for sheet in sheet_summaries),
    )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Strip deterministic Bilibili reply prefixes from a merged workbook.")
    parser.add_argument("input_path", type=Path, help="Input .xlsx/.xlsm/.csv file.")
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path.")
    parser.add_argument("--output-dir", type=Path, default=None, help="Output directory.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    result = strip_bilibili_reply_prefixes(args.input_path, output_path=args.output, output_dir=args.output_dir)
    print(f"Reply-prefix-stripped xlsx: {result.output_xlsx}")
    print(f"Summary: {result.summary_json}")
    print(f"Sheets processed: {result.sheets_processed}")
    print(f"Input rows: {result.input_rows}")
    print(f"Output rows: {result.output_rows}")
    print(f"Reply prefixes stripped: {result.reply_prefixes_stripped}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
