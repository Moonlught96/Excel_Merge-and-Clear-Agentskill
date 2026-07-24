from __future__ import annotations

import argparse
import json
import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

try:
    from tools.csv_excel_compat import (
        is_supported_input_path,
        load_workbook_for_processing,
        unsupported_input_message,
    )
    from tools.output_path_safety import atomic_output_path, ensure_output_paths_safe
except ModuleNotFoundError:
    from csv_excel_compat import (
        is_supported_input_path,
        load_workbook_for_processing,
        unsupported_input_message,
    )
    from output_path_safety import atomic_output_path, ensure_output_paths_safe


DEFAULT_CONFIG_PATH = Path(__file__).resolve().parents[1] / "config" / "platform-preprocessing.json"
SUPPORTED_OPERATIONS = frozenset(
    {
        "copy",
        "join_trimmed",
        "amazon_review_date",
        "amazon_star_rating",
        "amazon_helpful_count",
        "rakuten_review_date",
        "rakuten_helpful_count",
        "rakuten_user_attribute",
        "rakuten_display_name",
    }
)
AMAZON_REVIEW_DATE_PATTERN = re.compile(
    r"^\s*(?P<year>\d{4})年(?P<month>\d{1,2})月(?P<day>\d{1,2})日(?:在.+发布评论)?\s*$"
)
AMAZON_STAR_RATING_PATTERN = re.compile(
    r"^\s*(?P<rating>\d+(?:\.\d+)?)\s*颗星，最多\s*5\s*颗星\s*$"
)
AMAZON_HELPFUL_COUNT_PATTERN = re.compile(
    r"^\s*(?P<count>\d+)\s*个人发现此评论有用\s*$"
)
RAKUTEN_MONTH_DAY_YEAR_PATTERN = re.compile(
    r"^\s*(?P<month>\d{1,2})/(?P<day>\d{1,2})/(?P<year>\d{4})\s*$"
)
RAKUTEN_YEAR_MONTH_DAY_PATTERN = re.compile(
    r"^\s*(?P<year>\d{4})[/-](?P<month>\d{1,2})[/-](?P<day>\d{1,2})\s*$"
)
RAKUTEN_HELPFUL_COUNT_PATTERN = re.compile(r"^\s*(?P<count>\d+)\s*人\s*$")
RAKUTEN_GENDER_PATTERN = re.compile(r"男性|女性")
RAKUTEN_AGE_PATTERN = re.compile(r"\d{1,3}\s*(?:代(?:以上)?|歳|才)")
RAKUTEN_ANONYMOUS_DISPLAY_NAME = "購入者さん"


class PlatformPreprocessConfigError(ValueError):
    pass


class PlatformNotDetectedError(ValueError):
    pass


class AmbiguousPlatformError(ValueError):
    pass


class PlatformHeaderSignatureError(ValueError):
    pass


@dataclass(frozen=True)
class PreprocessOutputColumn:
    header: str
    operation: str
    source_headers: tuple[str, ...]
    separator: str


@dataclass(frozen=True)
class PlatformPreprocessDefinition:
    namespace: str
    aliases: tuple[str, ...]
    variant_name: str
    header_signature: tuple[str, ...]
    output_columns: tuple[PreprocessOutputColumn, ...]


@dataclass(frozen=True)
class PlatformPreprocessConfig:
    schema_version: int
    header_row: int
    platforms: tuple[PlatformPreprocessDefinition, ...]


@dataclass(frozen=True)
class SheetPreprocessSummary:
    sheet_name: str
    input_rows: int
    output_rows: int
    source_headers: tuple[str, ...]


@dataclass(frozen=True)
class PlatformPreprocessResult:
    input_path: Path
    output_xlsx: Path
    summary_json: Path
    platform: str
    profile_variant: str
    sheets_processed: int
    input_rows: int
    output_rows: int


@dataclass(frozen=True)
class PlatformPreprocessMergeResult:
    output_xlsx: Path
    summary_json: Path
    platform: str
    profile_variants: tuple[str, ...]
    files_processed: int
    sheets_processed: int
    data_rows_written: int


def normalize_header(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).strip().split())


def _require_string_list(value: Any, field_name: str, *, allow_empty: bool = False) -> tuple[str, ...]:
    if not isinstance(value, list) or not all(isinstance(item, str) for item in value):
        raise PlatformPreprocessConfigError(f"{field_name} must be a list of strings")
    values = tuple(item.strip() for item in value)
    if any(not item for item in values):
        raise PlatformPreprocessConfigError(f"{field_name} must not contain blank values")
    if not allow_empty and not values:
        raise PlatformPreprocessConfigError(f"{field_name} must not be empty")
    if len(set(values)) != len(values):
        raise PlatformPreprocessConfigError(f"{field_name} must not contain duplicate values")
    return values


def _load_output_columns(
    raw_output_columns: Any,
    field_prefix: str,
    header_signature: tuple[str, ...],
) -> tuple[PreprocessOutputColumn, ...]:
    if not isinstance(raw_output_columns, list) or not raw_output_columns:
        raise PlatformPreprocessConfigError(
            f"{field_prefix}.output_columns must be a non-empty list"
        )
    output_columns: list[PreprocessOutputColumn] = []
    output_headers: set[str] = set()
    single_source_operations = {
        "copy",
        "amazon_review_date",
        "amazon_star_rating",
        "amazon_helpful_count",
        "rakuten_review_date",
        "rakuten_helpful_count",
        "rakuten_user_attribute",
        "rakuten_display_name",
    }
    for column_index, raw_column in enumerate(raw_output_columns):
        column_prefix = f"{field_prefix}.output_columns[{column_index}]"
        if not isinstance(raw_column, dict):
            raise PlatformPreprocessConfigError(f"{column_prefix} must be an object")
        header = raw_column.get("header")
        if not isinstance(header, str) or not header.strip():
            raise PlatformPreprocessConfigError(
                f"{column_prefix}.header must be a non-empty string"
            )
        header = header.strip()
        normalized_output_header = normalize_header(header)
        if normalized_output_header in output_headers:
            raise PlatformPreprocessConfigError(
                f"{field_prefix}.output_columns must not contain duplicate headers"
            )
        output_headers.add(normalized_output_header)
        operation = raw_column.get("operation")
        if operation not in SUPPORTED_OPERATIONS:
            raise PlatformPreprocessConfigError(
                f"Unsupported preprocessing operation: {operation!r}"
            )
        source_headers = _require_string_list(
            raw_column.get("source_headers"),
            f"{column_prefix}.source_headers",
        )
        if any(source_header not in header_signature for source_header in source_headers):
            raise PlatformPreprocessConfigError(
                f"{column_prefix} references a source header outside header_signature"
            )
        separator = raw_column.get("separator", "")
        if not isinstance(separator, str):
            raise PlatformPreprocessConfigError(f"{column_prefix}.separator must be a string")
        if operation in single_source_operations and len(source_headers) != 1:
            raise PlatformPreprocessConfigError(
                f"{operation} requires exactly one source header"
            )
        if operation == "join_trimmed" and len(source_headers) < 2:
            raise PlatformPreprocessConfigError("join_trimmed requires at least two source headers")
        output_columns.append(
            PreprocessOutputColumn(
                header=header,
                operation=operation,
                source_headers=source_headers,
                separator=separator,
            )
        )
    return tuple(output_columns)


def _load_variant(
    raw_variant: Any,
    field_prefix: str,
) -> tuple[tuple[str, ...], tuple[PreprocessOutputColumn, ...]]:
    if not isinstance(raw_variant, dict):
        raise PlatformPreprocessConfigError(f"{field_prefix} must be an object")
    header_signature = _require_string_list(
        raw_variant.get("header_signature"),
        f"{field_prefix}.header_signature",
    )
    normalized_signature = tuple(normalize_header(header) for header in header_signature)
    if len(set(normalized_signature)) != len(normalized_signature):
        raise PlatformPreprocessConfigError(
            f"{field_prefix}.header_signature must not contain duplicate normalized headers"
        )
    return header_signature, _load_output_columns(
        raw_variant.get("output_columns"),
        field_prefix,
        header_signature,
    )


def load_config(path: Path | str = DEFAULT_CONFIG_PATH) -> PlatformPreprocessConfig:
    config_path = Path(path)
    try:
        raw = json.loads(config_path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as exc:
        raise PlatformPreprocessConfigError(
            f"Could not load platform preprocessing config: {config_path}"
        ) from exc

    if not isinstance(raw, dict):
        raise PlatformPreprocessConfigError("Platform preprocessing config root must be an object")
    if raw.get("schema_version") != 1:
        raise PlatformPreprocessConfigError("schema_version must be 1")
    header_row = raw.get("header_row", 1)
    if not isinstance(header_row, int) or isinstance(header_row, bool) or header_row < 1:
        raise PlatformPreprocessConfigError("header_row must be a positive integer")
    raw_platforms = raw.get("platforms")
    if not isinstance(raw_platforms, list) or not raw_platforms:
        raise PlatformPreprocessConfigError("platforms must be a non-empty list")

    registered_names: set[str] = set()
    registered_signatures: dict[tuple[str, ...], str] = {}
    platforms: list[PlatformPreprocessDefinition] = []
    for index, raw_platform in enumerate(raw_platforms):
        platform_prefix = f"platforms[{index}]"
        if not isinstance(raw_platform, dict):
            raise PlatformPreprocessConfigError(f"{platform_prefix} must be an object")
        namespace = raw_platform.get("namespace")
        if not isinstance(namespace, str) or not namespace.strip():
            raise PlatformPreprocessConfigError(f"{platform_prefix}.namespace must be a non-empty string")
        namespace = namespace.strip()
        aliases = _require_string_list(
            raw_platform.get("aliases", []),
            f"{platform_prefix}.aliases",
            allow_empty=True,
        )
        names = {namespace.casefold(), *(alias.casefold() for alias in aliases)}
        duplicate_names = registered_names.intersection(names)
        if duplicate_names:
            raise PlatformPreprocessConfigError(
                f"Duplicate platform name in config: {sorted(duplicate_names)[0]}"
            )
        registered_names.update(names)

        raw_variants = raw_platform.get("variants")
        legacy_fields_present = any(
            field_name in raw_platform for field_name in ("header_signature", "output_columns")
        )
        variants: list[tuple[str, tuple[str, ...], tuple[PreprocessOutputColumn, ...]]] = []
        if raw_variants is None:
            header_signature, output_columns = _load_variant(raw_platform, platform_prefix)
            variants.append(("default", header_signature, output_columns))
        else:
            if legacy_fields_present:
                raise PlatformPreprocessConfigError(
                    f"{platform_prefix} must define either legacy header_signature/output_columns or variants"
                )
            if not isinstance(raw_variants, list) or not raw_variants:
                raise PlatformPreprocessConfigError(f"{platform_prefix}.variants must be a non-empty list")
            variant_names: set[str] = set()
            for variant_index, raw_variant in enumerate(raw_variants):
                variant_prefix = f"{platform_prefix}.variants[{variant_index}]"
                if not isinstance(raw_variant, dict):
                    raise PlatformPreprocessConfigError(f"{variant_prefix} must be an object")
                variant_name = raw_variant.get("name")
                if not isinstance(variant_name, str) or not variant_name.strip():
                    raise PlatformPreprocessConfigError(
                        f"{variant_prefix}.name must be a non-empty string"
                    )
                variant_name = variant_name.strip()
                if variant_name.casefold() in variant_names:
                    raise PlatformPreprocessConfigError(
                        f"{platform_prefix}.variants must not contain duplicate names"
                    )
                variant_names.add(variant_name.casefold())
                header_signature, output_columns = _load_variant(raw_variant, variant_prefix)
                variants.append((variant_name, header_signature, output_columns))

        for variant_name, header_signature, output_columns in variants:
            signature_key = tuple(normalize_header(header) for header in header_signature)
            signature_owner = f"{namespace}:{variant_name}"
            existing_owner = registered_signatures.get(signature_key)
            if existing_owner is not None:
                raise PlatformPreprocessConfigError(
                    f"Duplicate header_signature registered by {existing_owner} and {signature_owner}"
                )
            registered_signatures[signature_key] = signature_owner
            platforms.append(
                PlatformPreprocessDefinition(
                    namespace=namespace,
                    aliases=aliases,
                    variant_name=variant_name,
                    header_signature=header_signature,
                    output_columns=output_columns,
                )
            )

    return PlatformPreprocessConfig(
        schema_version=1,
        header_row=header_row,
        platforms=tuple(platforms),
    )


def build_header_lookup(headers: tuple[Any, ...] | list[Any]) -> dict[str, list[int]]:
    lookup: dict[str, list[int]] = {}
    for index, header in enumerate(headers, start=1):
        key = normalize_header(header)
        if key:
            lookup.setdefault(key, []).append(index)
    return lookup


def _signature_matches(
    definition: PlatformPreprocessDefinition,
    headers: tuple[Any, ...] | list[Any],
) -> bool:
    return tuple(normalize_header(header) for header in headers) == tuple(
        normalize_header(header) for header in definition.header_signature
    )


def _find_definitions_by_name(
    platform: str,
    config: PlatformPreprocessConfig,
) -> list[PlatformPreprocessDefinition]:
    key = platform.strip().casefold()
    definitions = [
        definition
        for definition in config.platforms
        if key == definition.namespace.casefold()
        or key in {alias.casefold() for alias in definition.aliases}
    ]
    if not definitions:
        raise PlatformNotDetectedError(
            f"Platform is not registered for preprocessing: {platform!r}"
        )
    if len({definition.namespace for definition in definitions}) > 1:
        raise AmbiguousPlatformError(f"Platform matches multiple preprocessing definitions: {platform!r}")
    return definitions


def detect_platform(
    headers: tuple[Any, ...] | list[Any],
    config: PlatformPreprocessConfig,
    platform: str | None = None,
) -> PlatformPreprocessDefinition:
    if platform is not None:
        definitions = _find_definitions_by_name(platform, config)
        matches = [
            definition
            for definition in definitions
            if _signature_matches(definition, headers)
        ]
        if not matches:
            available_headers = [str(header) for header in headers if normalize_header(header)]
            raise PlatformHeaderSignatureError(
                f"Configured platform signature does not match {definitions[0].namespace!r}. "
                "Expected one of the registered header signatures: "
                f"{[list(definition.header_signature) for definition in definitions]}. "
                f"Available headers: {available_headers}"
            )
        if len(matches) > 1:
            raise AmbiguousPlatformError(
                "Multiple registered variants matched the source headers: "
                f"{[f'{definition.namespace}:{definition.variant_name}' for definition in matches]}"
            )
        return matches[0]

    matches = [
        definition
        for definition in config.platforms
        if _signature_matches(definition, headers)
    ]
    if not matches:
        available_headers = [str(header) for header in headers if normalize_header(header)]
        raise PlatformNotDetectedError(
            "No configured platform signature matched the source headers. "
            f"Available headers: {available_headers}"
        )
    if len(matches) > 1:
        raise AmbiguousPlatformError(
            "Multiple configured platform signatures matched the source headers: "
            f"{[f'{definition.namespace}:{definition.variant_name}' for definition in matches]}"
        )
    return matches[0]


def _trimmed_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    return text or None


def _amazon_review_date(value: Any) -> Any:
    text = _trimmed_text(value)
    if text is None:
        return None
    match = AMAZON_REVIEW_DATE_PATTERN.fullmatch(text)
    if not match:
        return value
    return (
        f"{int(match.group('year')):04d}-"
        f"{int(match.group('month')):02d}-"
        f"{int(match.group('day')):02d}"
    )


def _amazon_star_rating(value: Any) -> Any:
    text = _trimmed_text(value)
    if text is None:
        return None
    match = AMAZON_STAR_RATING_PATTERN.fullmatch(text)
    if not match:
        return value
    rating = float(match.group("rating"))
    if not 1 <= rating <= 5:
        return value
    return match.group("rating")


def _amazon_helpful_count(value: Any) -> Any:
    text = _trimmed_text(value)
    if text is None:
        return None
    match = AMAZON_HELPFUL_COUNT_PATTERN.fullmatch(text)
    if not match:
        return value
    return int(match.group("count"))


def _rakuten_review_date(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = _trimmed_text(value)
    if text is None:
        return None
    match = RAKUTEN_MONTH_DAY_YEAR_PATTERN.fullmatch(text)
    if match is None:
        match = RAKUTEN_YEAR_MONTH_DAY_PATTERN.fullmatch(text)
    if match is None:
        return value
    try:
        return date(
            int(match.group("year")),
            int(match.group("month")),
            int(match.group("day")),
        ).isoformat()
    except ValueError:
        return value


def _rakuten_helpful_count(value: Any) -> Any:
    text = _trimmed_text(value)
    if text is None:
        return None
    match = RAKUTEN_HELPFUL_COUNT_PATTERN.fullmatch(text)
    if not match:
        return value
    return int(match.group("count"))


def _rakuten_user_attribute(value: Any) -> str | None:
    text = _trimmed_text(value)
    if text is None:
        return None
    gender_match = RAKUTEN_GENDER_PATTERN.search(text)
    age_match = RAKUTEN_AGE_PATTERN.search(text)
    parts = [
        match.group(0)
        for match in (gender_match, age_match)
        if match is not None
    ]
    if not parts:
        return None
    return " ".join(re.sub(r"\s+", " ", part) for part in parts)


def _rakuten_display_name(value: Any) -> str | None:
    text = _trimmed_text(value)
    if text is None or text == RAKUTEN_ANONYMOUS_DISPLAY_NAME:
        return None
    return text


def _value_for_column(
    source_values: tuple[Any, ...],
    source_columns: dict[str, int],
    output_column: PreprocessOutputColumn,
) -> Any:
    values = [
        source_values[source_columns[header] - 1]
        if source_columns[header] - 1 < len(source_values)
        else None
        for header in output_column.source_headers
    ]
    if output_column.operation == "copy":
        return values[0]
    if output_column.operation == "join_trimmed":
        parts = [part for value in values if (part := _trimmed_text(value)) is not None]
        return output_column.separator.join(parts) if parts else None
    if output_column.operation == "amazon_review_date":
        return _amazon_review_date(values[0])
    if output_column.operation == "amazon_star_rating":
        return _amazon_star_rating(values[0])
    if output_column.operation == "amazon_helpful_count":
        return _amazon_helpful_count(values[0])
    if output_column.operation == "rakuten_review_date":
        return _rakuten_review_date(values[0])
    if output_column.operation == "rakuten_helpful_count":
        return _rakuten_helpful_count(values[0])
    if output_column.operation == "rakuten_user_attribute":
        return _rakuten_user_attribute(values[0])
    if output_column.operation == "rakuten_display_name":
        return _rakuten_display_name(values[0])
    raise PlatformPreprocessConfigError(
        f"Unsupported preprocessing operation: {output_column.operation!r}"
    )


def _write_output_row(output_sheet: Worksheet, values: list[Any]) -> None:
    output_sheet.append(values)
    row_number = output_sheet.max_row
    for column_index, value in enumerate(values, start=1):
        if isinstance(value, str) and value.startswith("="):
            output_sheet.cell(row=row_number, column=column_index).data_type = "s"


def _validate_batch_input_paths(input_paths: list[Path]) -> list[Path]:
    if not input_paths:
        raise ValueError("At least one input file is required.")

    validated: list[Path] = []
    seen: set[Path] = set()
    for input_path in input_paths:
        resolved = input_path.resolve()
        if resolved.is_dir():
            raise ValueError("Platform-preprocessed merge input must be an explicit file, not a folder.")
        if not is_supported_input_path(resolved):
            raise ValueError(f"{unsupported_input_message()}: {resolved}")
        if resolved.name.startswith("~$"):
            continue
        if not resolved.exists():
            raise FileNotFoundError(resolved)
        if resolved in seen:
            raise ValueError(f"Duplicate input path is not allowed: {resolved}")
        seen.add(resolved)
        validated.append(resolved)

    if not validated:
        raise ValueError("No usable input files were provided.")
    return validated


def _output_headers_for_definitions(
    definitions: list[PlatformPreprocessDefinition],
) -> tuple[str, ...]:
    output_headers = tuple(column.header for column in definitions[0].output_columns)
    for definition in definitions[1:]:
        candidate_headers = tuple(column.header for column in definition.output_columns)
        if candidate_headers != output_headers:
            raise PlatformPreprocessConfigError(
                "Registered platform variants must use one identical ordered preprocessing output schema. "
                f"{definitions[0].namespace}:{definitions[0].variant_name} has {list(output_headers)}, "
                f"but {definition.namespace}:{definition.variant_name} has {list(candidate_headers)}."
            )
    return output_headers


def _source_columns_for_definition(
    headers: tuple[Any, ...] | list[Any],
    definition: PlatformPreprocessDefinition,
) -> dict[str, int]:
    header_lookup = build_header_lookup(headers)
    return {
        header: header_lookup[normalize_header(header)][0]
        for header in definition.header_signature
    }


def preprocess_sheet(
    source_sheet: Worksheet,
    output_sheet: Worksheet,
    config: PlatformPreprocessConfig,
    definition: PlatformPreprocessDefinition,
) -> SheetPreprocessSummary:
    header_values = next(
        source_sheet.iter_rows(
            min_row=config.header_row,
            max_row=config.header_row,
            max_col=source_sheet.max_column,
            values_only=True,
        ),
        None,
    )
    if header_values is None:
        raise PlatformHeaderSignatureError(f"Worksheet has no header row: {source_sheet.title}")
    headers = tuple(header_values)
    detected = detect_platform(headers, config, platform=definition.namespace)
    if (
        detected.namespace != definition.namespace
        or detected.variant_name != definition.variant_name
    ):
        raise PlatformHeaderSignatureError(
            f"Worksheet platform does not match workbook platform: {source_sheet.title}"
        )
    source_columns = _source_columns_for_definition(headers, definition)

    _write_output_row(output_sheet, [column.header for column in definition.output_columns])
    input_rows = 0
    output_rows = 0
    for row_cells in source_sheet.iter_rows(
        min_row=config.header_row + 1,
        max_row=source_sheet.max_row,
        max_col=source_sheet.max_column,
        values_only=False,
    ):
        input_rows += 1
        source_values = tuple(cell.value for cell in row_cells)
        output_values = [
            _value_for_column(source_values, source_columns, output_column)
            for output_column in definition.output_columns
        ]
        _write_output_row(output_sheet, output_values)
        output_rows += 1

    return SheetPreprocessSummary(
        sheet_name=source_sheet.title,
        input_rows=input_rows,
        output_rows=output_rows,
        source_headers=tuple("" if header is None else str(header) for header in headers),
    )


def preprocess_and_merge_workbooks(
    input_paths: list[Path],
    config: PlatformPreprocessConfig,
    output_path: Path,
    *,
    platform: str,
    overwrite: bool = False,
) -> PlatformPreprocessMergeResult:
    """Deterministically preprocess exact registered variants, then merge their common output rows."""
    paths = _validate_batch_input_paths(input_paths)
    definitions = _find_definitions_by_name(platform, config)
    namespace = definitions[0].namespace
    output_headers = _output_headers_for_definitions(definitions)
    output_path = output_path.resolve()
    summary_path = output_path.with_suffix(".summary.json")
    ensure_output_paths_safe(paths, [output_path, summary_path], overwrite=overwrite)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    output_workbook = Workbook()
    output_sheet = output_workbook.active
    output_sheet.title = "总表"
    _write_output_row(output_sheet, list(output_headers))

    data_rows_written = 0
    sheets_processed = 0
    profile_variants: list[str] = []
    file_summaries: list[dict[str, Any]] = []
    try:
        for input_path in paths:
            input_workbook = load_workbook_for_processing(
                input_path,
                read_only=True,
                data_only=False,
            )
            sheet_summaries: list[dict[str, Any]] = []
            file_rows_written = 0
            try:
                if not input_workbook.worksheets:
                    raise PlatformHeaderSignatureError(f"Workbook has no worksheets: {input_path}")
                for source_sheet in input_workbook.worksheets:
                    header_values = next(
                        source_sheet.iter_rows(
                            min_row=config.header_row,
                            max_row=config.header_row,
                            max_col=source_sheet.max_column,
                            values_only=True,
                        ),
                        None,
                    )
                    if header_values is None:
                        raise PlatformHeaderSignatureError(
                            f"Worksheet has no header row: {input_path} / {source_sheet.title}"
                        )
                    headers = tuple(header_values)
                    definition = detect_platform(headers, config, platform=namespace)
                    if tuple(column.header for column in definition.output_columns) != output_headers:
                        raise PlatformPreprocessConfigError(
                            f"Configured output schema mismatch for {namespace}:{definition.variant_name}"
                        )
                    source_columns = _source_columns_for_definition(headers, definition)
                    sheet_rows_written = 0
                    for row_cells in source_sheet.iter_rows(
                        min_row=config.header_row + 1,
                        max_row=source_sheet.max_row,
                        max_col=source_sheet.max_column,
                        values_only=False,
                    ):
                        source_values = tuple(cell.value for cell in row_cells)
                        output_values = [
                            _value_for_column(source_values, source_columns, output_column)
                            for output_column in definition.output_columns
                        ]
                        _write_output_row(output_sheet, output_values)
                        sheet_rows_written += 1

                    if definition.variant_name not in profile_variants:
                        profile_variants.append(definition.variant_name)
                    sheets_processed += 1
                    file_rows_written += sheet_rows_written
                    data_rows_written += sheet_rows_written
                    sheet_summaries.append(
                        {
                            "sheet_name": source_sheet.title,
                            "profile_variant": definition.variant_name,
                            "input_rows": sheet_rows_written,
                            "output_rows": sheet_rows_written,
                            "source_headers": [
                                "" if header is None else str(header) for header in headers
                            ],
                        }
                    )
            finally:
                input_workbook.close()
            file_summaries.append(
                {
                    "file": str(input_path),
                    "data_rows_written": file_rows_written,
                    "sheets": sheet_summaries,
                }
            )

        with atomic_output_path(output_path) as staged_output:
            output_workbook.save(staged_output)
    finally:
        output_workbook.close()

    summary = {
        "mode": "platform_preprocessed_merge",
        "output_path": str(output_path),
        "platform": namespace,
        "profile_variants": profile_variants,
        "header_row": config.header_row,
        "output_headers": list(output_headers),
        "files_processed": len(paths),
        "sheets_processed": sheets_processed,
        "data_rows_written": data_rows_written,
        "files": file_summaries,
    }
    with atomic_output_path(summary_path) as staged_summary:
        staged_summary.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return PlatformPreprocessMergeResult(
        output_xlsx=output_path,
        summary_json=summary_path,
        platform=namespace,
        profile_variants=tuple(profile_variants),
        files_processed=len(paths),
        sheets_processed=sheets_processed,
        data_rows_written=data_rows_written,
    )


def make_output_paths(input_path: Path, output_dir: Path | None) -> tuple[Path, Path]:
    timestamp = datetime.now().strftime("%Y%m%d")
    parent = output_dir if output_dir else input_path.parent
    output_xlsx = parent / f"{timestamp}_{input_path.stem}.platform-preprocessed.xlsx"
    return output_xlsx, output_xlsx.with_suffix(".summary.json")


def preprocess_workbook(
    input_path: Path,
    config: PlatformPreprocessConfig,
    output_dir: Path | None = None,
    output_path: Path | None = None,
    *,
    platform: str | None = None,
    overwrite: bool = True,
) -> PlatformPreprocessResult:
    input_path = input_path.resolve()
    if not is_supported_input_path(input_path):
        raise ValueError(unsupported_input_message())
    if not input_path.exists():
        raise FileNotFoundError(input_path)

    if output_path is None:
        output_xlsx, summary_json = make_output_paths(input_path, output_dir)
    else:
        output_xlsx = output_path.resolve()
        summary_json = output_xlsx.with_suffix(".platform-preprocessed.summary.json")
    ensure_output_paths_safe(
        [input_path],
        [output_xlsx, summary_json],
        overwrite=overwrite,
    )

    input_workbook = load_workbook_for_processing(
        input_path,
        read_only=True,
        data_only=False,
    )
    output_workbook = Workbook()
    sheet_summaries: list[SheetPreprocessSummary] = []
    try:
        if not input_workbook.worksheets:
            raise PlatformHeaderSignatureError("Workbook has no worksheets")
        first_sheet = input_workbook.worksheets[0]
        first_headers = next(
            first_sheet.iter_rows(
                min_row=config.header_row,
                max_row=config.header_row,
                max_col=first_sheet.max_column,
                values_only=True,
            ),
            None,
        )
        if first_headers is None:
            raise PlatformHeaderSignatureError(
                f"Worksheet has no header row: {first_sheet.title}"
            )
        definition = detect_platform(first_headers, config, platform=platform)
        for sheet_index, source_sheet in enumerate(input_workbook.worksheets):
            output_sheet = (
                output_workbook.active
                if sheet_index == 0
                else output_workbook.create_sheet()
            )
            output_sheet.title = source_sheet.title
            sheet_summaries.append(
                preprocess_sheet(source_sheet, output_sheet, config, definition)
            )
        with atomic_output_path(output_xlsx) as staged_output:
            output_workbook.save(staged_output)
    finally:
        input_workbook.close()
        output_workbook.close()

    summary = {
        "input_path": str(input_path),
        "output_xlsx": str(output_xlsx),
        "platform": definition.namespace,
        "profile_variant": definition.variant_name,
        "header_row": config.header_row,
        "sheets_processed": len(sheet_summaries),
        "input_rows": sum(sheet.input_rows for sheet in sheet_summaries),
        "output_rows": sum(sheet.output_rows for sheet in sheet_summaries),
        "output_headers": [column.header for column in definition.output_columns],
        "sheets": [
            {
                "sheet_name": sheet.sheet_name,
                "input_rows": sheet.input_rows,
                "output_rows": sheet.output_rows,
                "source_headers": list(sheet.source_headers),
            }
            for sheet in sheet_summaries
        ],
    }
    with atomic_output_path(summary_json) as staged_summary:
        staged_summary.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2),
            encoding="utf-8",
        )

    return PlatformPreprocessResult(
        input_path=input_path,
        output_xlsx=output_xlsx,
        summary_json=summary_json,
        platform=definition.namespace,
        profile_variant=definition.variant_name,
        sheets_processed=len(sheet_summaries),
        input_rows=sum(sheet.input_rows for sheet in sheet_summaries),
        output_rows=sum(sheet.output_rows for sheet in sheet_summaries),
    )


def parse_args(argv: list[str] | None = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Preprocess a registered platform export using exact configured header signatures."
    )
    parser.add_argument("input_paths", type=Path, nargs="+", help="Input .xlsx/.xlsm/.csv file(s).")
    parser.add_argument("--config", type=Path, default=DEFAULT_CONFIG_PATH)
    parser.add_argument("--output", type=Path, default=None, help="Output .xlsx path.")
    parser.add_argument("--output-dir", type=Path, default=None, help="Output directory.")
    parser.add_argument(
        "--platform",
        default=None,
        help="Optional registered platform name; signature validation remains mandatory.",
    )
    parser.add_argument(
        "--overwrite",
        action="store_true",
        help="Replace confirmed existing outputs.",
    )
    parser.add_argument(
        "--merge-registered-variants",
        action="store_true",
        help="Preprocess multiple exact registered variants into one common merged workbook.",
    )
    return parser.parse_args(argv)


def main(argv: list[str] | None = None) -> int:
    args = parse_args(argv)
    config = load_config(args.config)
    if args.merge_registered_variants:
        if args.output is None:
            raise SystemExit("--merge-registered-variants requires --output")
        if args.platform is None:
            raise SystemExit("--merge-registered-variants requires --platform")
        result = preprocess_and_merge_workbooks(
            args.input_paths,
            config,
            args.output,
            platform=args.platform,
            overwrite=args.overwrite,
        )
        print(f"Platform-preprocessed merged xlsx: {result.output_xlsx}")
        print(f"Summary: {result.summary_json}")
        print(f"Platform: {result.platform}")
        print(f"Profile variants: {', '.join(result.profile_variants)}")
        print(f"Files processed: {result.files_processed}")
        print(f"Sheets processed: {result.sheets_processed}")
        print(f"Data rows written: {result.data_rows_written}")
        return 0
    if len(args.input_paths) != 1:
        raise SystemExit("Provide exactly one input unless --merge-registered-variants is set")
    result = preprocess_workbook(
        args.input_paths[0],
        config,
        output_dir=args.output_dir,
        output_path=args.output,
        platform=args.platform,
        overwrite=args.overwrite,
    )
    print(f"Platform-preprocessed xlsx: {result.output_xlsx}")
    print(f"Summary: {result.summary_json}")
    print(f"Platform: {result.platform}")
    print(f"Profile variant: {result.profile_variant}")
    print(f"Sheets processed: {result.sheets_processed}")
    print(f"Data rows written: {result.output_rows}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
