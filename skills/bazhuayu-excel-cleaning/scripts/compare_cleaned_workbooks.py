from __future__ import annotations

import argparse
import csv
import hashlib
import json
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, date):
        return value.isoformat()
    return str(value)


def row_key(row: list[str]) -> str:
    return json.dumps(row, ensure_ascii=False, separators=(",", ":"))


def short_hash(key: str) -> str:
    return hashlib.sha256(key.encode("utf-8")).hexdigest()[:12]


def read_sheet(path: Path, sheet_name: str | None = None) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[sheet_name] if sheet_name else workbook.worksheets[0]
    rows = [[normalize_cell(cell) for cell in row] for row in worksheet.iter_rows(values_only=True)]
    header = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []
    return {
        "path": str(path),
        "sheet_names": workbook.sheetnames,
        "sheet_name": worksheet.title,
        "max_row": worksheet.max_row,
        "max_column": worksheet.max_column,
        "header": header,
        "data_rows": data_rows,
    }


def build_row_index(rows: list[list[str]]) -> dict[str, list[int]]:
    index: dict[str, list[int]] = defaultdict(list)
    for offset, row in enumerate(rows, start=2):
        index[row_key(row)].append(offset)
    return index


def expand_counter_diff(counter: Counter[str], other: Counter[str]) -> list[str]:
    expanded: list[str] = []
    for key, count in (counter - other).items():
        expanded.extend([key] * count)
    return expanded


def first_rows_for_keys(keys: list[str], index: dict[str, list[int]], limit: int) -> list[dict[str, Any]]:
    result: list[dict[str, Any]] = []
    seen_positions: dict[str, int] = defaultdict(int)
    for key in keys[:limit]:
        position = seen_positions[key]
        row_numbers = index.get(key, [])
        row_number = row_numbers[position] if position < len(row_numbers) else None
        seen_positions[key] += 1
        result.append(
            {
                "row_number": row_number,
                "row_hash": short_hash(key),
                "values": json.loads(key),
            }
        )
    return result


def duplicate_summary(rows: list[list[str]], comment_column: int) -> dict[str, Any]:
    comments = [row[comment_column - 1] if len(row) >= comment_column else "" for row in rows]
    counter = Counter(comments)
    duplicates = {value: count for value, count in counter.items() if value and count > 1}
    return {
        "duplicate_comment_count": sum(count - 1 for count in duplicates.values()),
        "duplicate_comment_groups": len(duplicates),
        "sample_duplicate_comments": [
            {"comment": value, "count": count} for value, count in list(duplicates.items())[:20]
        ],
    }


def compare_workbooks(
    left_path: Path,
    right_path: Path,
    output_dir: Path,
    left_label: str,
    right_label: str,
    comment_column: int,
) -> dict[str, Any]:
    left = read_sheet(left_path)
    right = read_sheet(right_path)
    output_dir.mkdir(parents=True, exist_ok=True)

    left_rows = left["data_rows"]
    right_rows = right["data_rows"]
    left_index = build_row_index(left_rows)
    right_index = build_row_index(right_rows)
    left_counter = Counter(left_index.keys())
    right_counter = Counter(right_index.keys())

    only_left_keys = expand_counter_diff(left_counter, right_counter)
    only_right_keys = expand_counter_diff(right_counter, left_counter)

    left_comments = [
        row[comment_column - 1] if len(row) >= comment_column else "" for row in left_rows
    ]
    right_comments = [
        row[comment_column - 1] if len(row) >= comment_column else "" for row in right_rows
    ]
    only_left_comments = list((Counter(left_comments) - Counter(right_comments)).elements())
    only_right_comments = list((Counter(right_comments) - Counter(left_comments)).elements())

    sequence_mismatches = 0
    first_sequence_mismatches: list[dict[str, Any]] = []
    for idx, (left_row, right_row) in enumerate(zip(left_rows, right_rows), start=2):
        if left_row != right_row:
            sequence_mismatches += 1
            if len(first_sequence_mismatches) < 50:
                first_sequence_mismatches.append(
                    {
                        "row_number": idx,
                        left_label: left_row,
                        right_label: right_row,
                    }
                )

    report = {
        "left_label": left_label,
        "right_label": right_label,
        "left_path": str(left_path),
        "right_path": str(right_path),
        "left_sheet": left["sheet_name"],
        "right_sheet": right["sheet_name"],
        "left_sheet_names": left["sheet_names"],
        "right_sheet_names": right["sheet_names"],
        "left_total_rows_including_header": left["max_row"],
        "right_total_rows_including_header": right["max_row"],
        "left_data_rows": len(left_rows),
        "right_data_rows": len(right_rows),
        "left_columns": left["max_column"],
        "right_columns": right["max_column"],
        "headers_match": left["header"] == right["header"],
        "left_header": left["header"],
        "right_header": right["header"],
        "row_sequence_mismatch_count": sequence_mismatches,
        "whole_row_only_in_left_count": len(only_left_keys),
        "whole_row_only_in_right_count": len(only_right_keys),
        "comment_only_in_left_count": len(only_left_comments),
        "comment_only_in_right_count": len(only_right_comments),
        "left_duplicate_summary": duplicate_summary(left_rows, comment_column),
        "right_duplicate_summary": duplicate_summary(right_rows, comment_column),
        "whole_row_only_in_left_samples": first_rows_for_keys(only_left_keys, left_index, 100),
        "whole_row_only_in_right_samples": first_rows_for_keys(only_right_keys, right_index, 100),
        "comment_only_in_left_samples": only_left_comments[:100],
        "comment_only_in_right_samples": only_right_comments[:100],
        "first_sequence_mismatch_samples": first_sequence_mismatches,
    }

    summary_path = output_dir / "comparison-summary.json"
    summary_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")

    write_diff_csv(output_dir / "only-in-left.csv", left["header"], report["whole_row_only_in_left_samples"])
    write_diff_csv(output_dir / "only-in-right.csv", right["header"], report["whole_row_only_in_right_samples"])
    write_report_xlsx(output_dir / "comparison-report.xlsx", report)

    report["summary_path"] = str(summary_path)
    report["report_xlsx"] = str(output_dir / "comparison-report.xlsx")
    report["only_left_csv"] = str(output_dir / "only-in-left.csv")
    report["only_right_csv"] = str(output_dir / "only-in-right.csv")
    return report


def write_diff_csv(path: Path, header: list[str], rows: list[dict[str, Any]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.writer(file)
        writer.writerow(["source_row", "row_hash", *header])
        for row in rows:
            writer.writerow([row["row_number"], row["row_hash"], *row["values"]])


def write_report_xlsx(path: Path, report: dict[str, Any]) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Summary"
    summary_rows = [
        ["Metric", report["left_label"], report["right_label"], "Notes"],
        ["Sheet name", report["left_sheet"], report["right_sheet"], ""],
        ["Rows incl header", report["left_total_rows_including_header"], report["right_total_rows_including_header"], ""],
        ["Data rows", report["left_data_rows"], report["right_data_rows"], ""],
        ["Columns", report["left_columns"], report["right_columns"], ""],
        ["Headers match", str(report["headers_match"]), str(report["headers_match"]), ""],
        ["Whole-row only count", report["whole_row_only_in_left_count"], report["whole_row_only_in_right_count"], "Compared as full row values"],
        ["Comment-only count", report["comment_only_in_left_count"], report["comment_only_in_right_count"], "Compared by column 3 text"],
        ["Sequence mismatch count", report["row_sequence_mismatch_count"], report["row_sequence_mismatch_count"], "Rows differ at same row number"],
        ["Duplicate comment extra rows", report["left_duplicate_summary"]["duplicate_comment_count"], report["right_duplicate_summary"]["duplicate_comment_count"], "Column 3 duplicates after cleaning"],
    ]
    for row in summary_rows:
        summary.append(row)
    style_sheet(summary)

    add_diff_sheet(workbook, "Only in Left", report["left_header"], report["whole_row_only_in_left_samples"])
    add_diff_sheet(workbook, "Only in Right", report["right_header"], report["whole_row_only_in_right_samples"])
    add_comment_sheet(workbook, "Comment Diff", report)
    add_sequence_sheet(workbook, "Sequence Diff", report)
    workbook.save(path)


def style_sheet(sheet: Any) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for column in sheet.columns:
        letter = column[0].column_letter
        width = min(max(len(str(cell.value or "")) for cell in column) + 2, 60)
        sheet.column_dimensions[letter].width = width
    sheet.freeze_panes = "A2"
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def add_diff_sheet(workbook: Workbook, name: str, header: list[str], rows: list[dict[str, Any]]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(["source_row", "row_hash", *header])
    for row in rows:
        sheet.append([row["row_number"], row["row_hash"], *row["values"]])
    style_sheet(sheet)


def add_comment_sheet(workbook: Workbook, name: str, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(["side", "comment"])
    for comment in report["comment_only_in_left_samples"]:
        sheet.append([report["left_label"], comment])
    for comment in report["comment_only_in_right_samples"]:
        sheet.append([report["right_label"], comment])
    style_sheet(sheet)


def add_sequence_sheet(workbook: Workbook, name: str, report: dict[str, Any]) -> None:
    sheet = workbook.create_sheet(name)
    sheet.append(["row_number", report["left_label"], report["right_label"]])
    for item in report["first_sequence_mismatch_samples"]:
        sheet.append(
            [
                item["row_number"],
                json.dumps(item[report["left_label"]], ensure_ascii=False),
                json.dumps(item[report["right_label"]], ensure_ascii=False),
            ]
        )
    style_sheet(sheet)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare two cleaned Excel workbooks.")
    parser.add_argument("--left", type=Path, required=True)
    parser.add_argument("--right", type=Path, required=True)
    parser.add_argument("--output-dir", type=Path, required=True)
    parser.add_argument("--left-label", default="left")
    parser.add_argument("--right-label", default="right")
    parser.add_argument("--comment-column", type=int, default=3)
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    report = compare_workbooks(
        left_path=args.left,
        right_path=args.right,
        output_dir=args.output_dir,
        left_label=args.left_label,
        right_label=args.right_label,
        comment_column=args.comment_column,
    )
    print(f"Report: {report['report_xlsx']}")
    print(f"Rows: {report['left_label']}={report['left_data_rows']}, {report['right_label']}={report['right_data_rows']}")
    print(
        "Whole-row only: "
        f"{report['left_label']}={report['whole_row_only_in_left_count']}, "
        f"{report['right_label']}={report['whole_row_only_in_right_count']}"
    )
    print(
        "Comment only: "
        f"{report['left_label']}={report['comment_only_in_left_count']}, "
        f"{report['right_label']}={report['comment_only_in_right_count']}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
