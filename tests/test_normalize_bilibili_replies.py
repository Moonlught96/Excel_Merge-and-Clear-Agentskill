from __future__ import annotations

import json
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from tools.strip_bilibili_reply_prefixes import OutputPathConflictError, strip_bilibili_reply_prefixes


class StripBilibiliReplyPrefixesTest(unittest.TestCase):
    def test_closes_input_workbook_when_prefix_processing_fails(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-strip-closes-on-error"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "prefix-stripped.xlsx"
        output_path.unlink(missing_ok=True)
        output_path.with_suffix(".reply-prefix-stripped.summary.json").unlink(missing_ok=True)
        source = Workbook()
        source.active.append(["content"])
        source.active.append(["回复@user：真实评论内容"])
        source.save(input_path)
        source.close()

        opened = load_workbook(input_path, read_only=True, data_only=False)
        opened.close = mock.Mock(wraps=opened.close)
        with mock.patch(
            "tools.strip_bilibili_reply_prefixes.load_workbook_for_processing",
            return_value=opened,
        ), mock.patch(
            "tools.strip_bilibili_reply_prefixes.strip_sheet_reply_prefixes",
            side_effect=RuntimeError("processing failed"),
        ):
            with self.assertRaisesRegex(RuntimeError, "processing failed"):
                strip_bilibili_reply_prefixes(input_path, output_path)

        opened.close.assert_called_once_with()

    def test_requires_explicit_overwrite_for_existing_output(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-strip-existing-output"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "prefix-stripped.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["content"])
        sheet.append(["回复@user：真实评论内容"])
        workbook.save(input_path)
        output_path.write_text("existing", encoding="utf-8")

        with self.assertRaisesRegex(OutputPathConflictError, "already exists"):
            strip_bilibili_reply_prefixes(
                input_path,
                output_path,
                overwrite=False,
            )

        self.assertEqual("existing", output_path.read_text(encoding="utf-8"))

    def test_strips_reply_prefix_without_moving_rows_or_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-strip-bilibili-reply-prefixes"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "merged.xlsx"
        output_path = tmp / "prefix-stripped.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "总表"
        header = ["rpid", "parent_rpid", "username", "content", "like_count", "timestamp", "ip_location"]
        sheet.append(header)
        sheet.append(["root1", "0", "root-user", "主评论内容足够长", "10", "1547698467", "未知"])
        sheet.append(["child1", "root1", "child-user", "回复@root-user：一级回复内容", "2", "1547698468", "未知"])
        sheet.append(["child2", "child1", "grand-user", "回复 @child-user: 二级回复内容", "1", "1547698469", "未知"])
        sheet.append(["child3", "child2", "third-user", "  回复 @阿珍 爱上了阿强：[滑稽]  ", "0", "1547698470", "未知"])
        sheet.append(["plain", "0", "plain-user", "这不是回复@用户：不要修改", "0", "1547698471", "未知"])
        sheet.append(["no-colon", "0", "plain-user", "回复@用户 没有冒号不要修改", "0", "1547698472", "未知"])
        workbook.save(input_path)

        result = strip_bilibili_reply_prefixes(input_path, output_path)

        normalized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(normalized["总表"].iter_rows(values_only=True))

        self.assertEqual(tuple(header), rows[0])
        self.assertEqual(7, len(rows))
        self.assertEqual("主评论内容足够长", rows[1][3])
        self.assertEqual("一级回复内容", rows[2][3])
        self.assertEqual("二级回复内容", rows[3][3])
        self.assertEqual("[滑稽]  ", rows[4][3])
        self.assertEqual("这不是回复@用户：不要修改", rows[5][3])
        self.assertEqual("回复@用户 没有冒号不要修改", rows[6][3])
        self.assertEqual(("child1", "root1", "child-user"), rows[2][:3])

        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        self.assertEqual(6, summary["input_rows"])
        self.assertEqual(6, summary["output_rows"])
        self.assertEqual(3, summary["reply_prefixes_stripped"])
        self.assertEqual(["content"], summary["sheets"][0]["target_headers"])

    def test_preserves_formula_cells_outside_the_content_column(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-strip-preserves-formulas"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "merged.xlsx"
        output_path = tmp / "prefix-stripped.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "总表"
        sheet.append(["content", "like_count", "timestamp"])
        sheet.append(["回复@user：真实评论内容", "=1+1", "1547698468"])
        workbook.save(input_path)

        result = strip_bilibili_reply_prefixes(input_path, output_path)

        normalized = load_workbook(result.output_xlsx, read_only=True, data_only=False)
        self.assertEqual("真实评论内容", normalized["总表"].cell(row=2, column=1).value)
        self.assertEqual("=1+1", normalized["总表"].cell(row=2, column=2).value)

    def test_csv_formula_like_reply_body_remains_text(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-strip-csv-formula-text"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.csv"
        output_path = tmp / "prefix-stripped.xlsx"
        input_path.write_text(
            "content,like_count,timestamp\n回复@user：=1+1,=2+2,1547698468\n",
            encoding="utf-8-sig",
        )

        strip_bilibili_reply_prefixes(input_path, output_path)

        normalized = load_workbook(output_path, read_only=False, data_only=False)
        self.assertEqual("=1+1", normalized.active.cell(row=2, column=1).value)
        self.assertEqual("s", normalized.active.cell(row=2, column=1).data_type)
        self.assertEqual("=2+2", normalized.active.cell(row=2, column=2).value)
        self.assertEqual("s", normalized.active.cell(row=2, column=2).data_type)



if __name__ == "__main__":
    unittest.main()
