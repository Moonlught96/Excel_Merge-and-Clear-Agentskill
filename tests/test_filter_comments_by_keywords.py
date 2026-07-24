from __future__ import annotations

import json
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.filter_comments_by_keywords import (
    KeywordFilterConfigError,
    KeywordTargetHeaderError,
    filter_workbook,
)


class FilterCommentsByKeywordsTest(unittest.TestCase):
    def write_source(self, path: Path, headers: list[str], rows: list[list[object]]) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "总表"
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        workbook.save(path)
        workbook.close()

    def test_keeps_only_rows_matching_confirmed_keywords_case_insensitively(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-twitter-keyword-filter"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "standardized.xlsx"
        output_path = tmp / "twitter-keyword-filtered.xlsx"
        output_path.unlink(missing_ok=True)
        output_path.with_suffix(".keyword-filter.summary.json").unlink(missing_ok=True)
        self.write_source(
            source_path,
            ["评论日期", "评论内容", "哈希ID", "点赞数"],
            [
                ["2026-07-22", "ScreenBar works well.", "hash-a", 3],
                ["2026-07-22", "A generic desk setup.", "hash-b", 1],
                ["2026-07-23", "BENQ makes it brighter.", "hash-c", 2],
            ],
        )

        result = filter_workbook(
            source_path,
            ["screenbar", "benq"],
            output_path=output_path,
            overwrite=False,
        )

        output_book = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        try:
            rows = list(output_book["总表"].iter_rows(values_only=True))
        finally:
            output_book.close()

        self.assertEqual(
            [
                ("评论日期", "评论内容", "哈希ID", "点赞数"),
                ("2026-07-22", "ScreenBar works well.", "hash-a", 3),
                ("2026-07-23", "BENQ makes it brighter.", "hash-c", 2),
            ],
            rows,
        )
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        self.assertEqual(3, summary["input_rows"])
        self.assertEqual(2, summary["kept_rows"])
        self.assertEqual(1, summary["deleted_rows"])
        self.assertNotIn("ScreenBar works well.", result.summary_json.read_text(encoding="utf-8"))

    def test_rejects_blank_keywords_without_writing_output(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-twitter-keyword-filter-blank"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "standardized.xlsx"
        output_path = tmp / "filtered.xlsx"
        self.write_source(source_path, ["评论内容"], [["ScreenBar works well."]])

        with self.assertRaises(KeywordFilterConfigError):
            filter_workbook(source_path, [" ", "\t"], output_path=output_path)

        self.assertFalse(output_path.exists())

    def test_requires_exact_target_header_without_guessing(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-twitter-keyword-filter-header"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "standardized.xlsx"
        self.write_source(source_path, ["正文"], [["ScreenBar works well."]])

        with self.assertRaises(KeywordTargetHeaderError):
            filter_workbook(source_path, ["screenbar"], output_dir=tmp / "out")


if __name__ == "__main__":
    unittest.main()
