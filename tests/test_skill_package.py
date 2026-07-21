from __future__ import annotations

import os
import shutil
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[1]
SKILL_NAME = "product-user-comment-data-merge-cleaning"
SKILL_ROOT = PROJECT_ROOT / "skills" / SKILL_NAME

REFERENCE_FILES = (
    "workflow.md",
    "data-contract.md",
    "header-standardization.md",
    "cleaning-rules.md",
    "naming-and-retention.md",
    "tool-reference.md",
    "extension-policy.md",
)

SCRIPT_FILES = (
    "cleanup_intermediate_outputs.py",
    "clean_excel_comments.py",
    "compare_cleaned_workbooks.py",
    "csv_excel_compat.py",
    "merge_excel_workbooks.py",
    "hash_id_project_store.py",
    "hash_id_pseudonymizer.py",
    "output_file_naming.py",
    "standardize_excel_headers.py",
    "strip_bilibili_reply_prefixes.py",
)

CONFIG_FILES = (
    "comment-cleaner.json",
    "hash-id.json",
    "header-standardizer.json",
)


class SkillPackageTest(unittest.TestCase):
    def test_complete_standard_skill_structure_exists(self) -> None:
        required_files = (
            SKILL_ROOT / "SKILL.md",
            SKILL_ROOT / "agents" / "openai.yaml",
            SKILL_ROOT / "assets" / "workflow-confirmation-template.md",
            SKILL_ROOT / "assets" / "rule-extension-template.md",
            *(SKILL_ROOT / "references" / name for name in REFERENCE_FILES),
            *(SKILL_ROOT / "scripts" / name for name in SCRIPT_FILES),
            *(SKILL_ROOT / "config" / name for name in CONFIG_FILES),
        )

        missing = [str(path.relative_to(PROJECT_ROOT)) for path in required_files if not path.is_file()]
        self.assertEqual([], missing, f"Skill package is missing required files: {missing}")

    def test_rule_extension_template_enforces_identity_alias_safety(self) -> None:
        template = (
            SKILL_ROOT / "assets" / "rule-extension-template.md"
        ).read_text(encoding="utf-8")

        required_constraints = (
            "未经确认的身份别名不得添加。",
            "Evidence must contain headers/schema only or be redacted.",
            "Raw identity values must never be committed.",
            "新增平台或表头别名必须获得用户明确确认和平台专属证据。",
            "`用户身份` 禁止作为身份来源。",
            "评论 ID 和父评论 ID 禁止作为身份来源。",
            "URL 和主页链接禁止作为身份来源。",
            "IP 字段禁止作为身份来源。",
            "来源自带的 `哈希ID` 禁止作为身份来源。",
        )
        for constraint in required_constraints:
            self.assertIn(constraint, template)

    def test_skill_entrypoint_is_concise_and_uses_progressive_disclosure(self) -> None:
        skill = (SKILL_ROOT / "SKILL.md").read_text(encoding="utf-8")

        self.assertRegex(
            skill,
            r"\A---\nname: product-user-comment-data-merge-cleaning\ndescription: Use when ",
        )
        self.assertIn("pseudonymized standardization", skill)
        self.assertNotIn("privacy-safe", skill)
        self.assertIn("# 产品用户评论数据合并与清洗 Skill", skill)
        self.assertIn(
            "为抓取的大量用户评论数据进行文档合并、标准化和清洗工作，并输出为 XLSX 与 CSV 格式文档。",
            skill,
        )
        self.assertIn("目前输入不受限制，支持 Excel 与 CSV 文件。", skill)
        self.assertLessEqual(len(skill.splitlines()), 180)
        for heading in ("## Skill Responsibilities", "## Trigger Scenarios", "## Execution Steps", "## Output Standard"):
            self.assertIn(heading, skill)
        for reference_name in REFERENCE_FILES:
            self.assertIn(f"references/{reference_name}", skill)
        self.assertIn("scripts/merge_excel_workbooks.py", skill)
        self.assertIn("scripts/standardize_excel_headers.py", skill)
        self.assertIn("scripts/clean_excel_comments.py", skill)
        self.assertIn("in a confirmed single-file run, use the original input as the source", skill)
        self.assertIn("Automatic creation of a new protected hash-ID project requires Windows DPAPI", skill)
        self.assertNotIn("tools/clean_excel_comments.py", skill)

    def test_references_preserve_all_confirmed_rule_categories(self) -> None:
        references = {
            name: (SKILL_ROOT / "references" / name).read_text(encoding="utf-8")
            for name in REFERENCE_FILES
        }
        combined = "\n".join(references.values())

        required_rules = (
            "请确认以上产品名、数据来源和文件命名是否正确，并确认是否可以进入合并流程。",
            "是否已经提供并合并完所有需要合并的表格？你确认后我再进行标准化。",
            "是否已经提供完成所有 KOL 清理词？你确认后我再进行清洗。",
            "`评论日期`、`评论内容`、`产品名`、`哈希ID`、`点赞数`、`子评论数/追评数`、`一级评论`、`二级评论`、`三级评论`",
            "`评论日期与产品`",
            "Beijing date (`UTC+8`)",
            "Relative year values output only `YYYY`; relative month values output only `YYYY-MM`.",
            "Chinese comments whose trimmed length is less than or equal to 7 characters are deleted.",
            "Non-Chinese comments with four or fewer words are deleted.",
            "Pure numeric comments keep the legacy seven-character threshold",
            "`一级评论`, `二级评论`, and `三级评论` cells whose trimmed length is less than or equal to 5 characters",
            "Fixed delete words are appended to the original `链接` rule",
            "Chinese, English, Japanese, Korean, Spanish, Thai, and Hindi",
            "Do not use AI",
            "data_only=False",
            "Keep only the final cleaned `.xlsx` and `.csv`",
            "Do not delete original input files.",
        )
        for rule in required_rules:
            self.assertIn(rule, combined)

        self.assertIn("## Multi-File Workflow", references["workflow.md"])
        self.assertIn("## Standard Output Schema", references["header-standardization.md"])
        self.assertIn("## Main Comment Cleaning", references["cleaning-rules.md"])
        self.assertIn("## Filename Standard", references["naming-and-retention.md"])
        self.assertIn("## Command Reference", references["tool-reference.md"])
        self.assertIn("## Locked Base Rules", references["extension-policy.md"])

    def test_bundled_scripts_and_configs_match_project_sources(self) -> None:
        for filename in SCRIPT_FILES:
            project_source = (PROJECT_ROOT / "tools" / filename).read_bytes()
            bundled_source = (SKILL_ROOT / "scripts" / filename).read_bytes()
            self.assertEqual(project_source, bundled_source, f"Bundled script is stale: {filename}")

        for filename in CONFIG_FILES:
            project_source = (PROJECT_ROOT / "config" / filename).read_bytes()
            bundled_source = (SKILL_ROOT / "config" / filename).read_bytes()
            self.assertEqual(project_source, bundled_source, f"Bundled config is stale: {filename}")

    def test_sync_tool_covers_all_bundled_scripts_and_configs(self) -> None:
        sync_tool = (PROJECT_ROOT / "tools" / "sync_skill_bundle.py").read_text(
            encoding="utf-8"
        )

        for filename in (*SCRIPT_FILES, *CONFIG_FILES):
            self.assertIn(f'"{filename}"', sync_tool, f"Sync tool omits: {filename}")
    def test_project_instructions_enforce_standard_skill_packaging(self) -> None:
        agents = (PROJECT_ROOT / "AGENTS.md").read_text(encoding="utf-8")

        self.assertIn("## Agent Skill Packaging Standard", agents)
        self.assertIn("完整的 Skill 文件夹结构", agents)
        self.assertIn("Skill 职责、触发场景、执行步骤、输出标准", agents)
        self.assertIn("`references/`", agents)
        self.assertIn("`scripts/`", agents)
        self.assertIn("`assets/`", agents)
        self.assertIn("单独复制", agents)
        self.assertIn("独立运行", agents)

    def test_copied_skill_runs_without_project_root(self) -> None:
        temp_directory = tempfile.TemporaryDirectory(prefix="standalone-skill-")
        temp_root = Path(temp_directory.name).resolve()
        self.assertFalse(temp_root.is_relative_to(PROJECT_ROOT.resolve()))
        copied_skill = temp_root / SKILL_NAME
        run_root = temp_root / "run"
        run_root.mkdir()
        shutil.copytree(SKILL_ROOT, copied_skill)

        subprocess_env = os.environ.copy()
        for name in (
            "PYTHONPATH",
            "PYTHONHOME",
            "PYTHONSTARTUP",
            "PYTHONUSERBASE",
            "PYTHONINSPECT",
            "PYTHONSAFEPATH",
            "PYTHONPLATLIBDIR",
        ):
            subprocess_env.pop(name, None)
        subprocess_env["PYTHONNOUSERSITE"] = "1"

        probe_code = (
            "from pathlib import Path\n"
            "import sys\n"
            f"project_root = Path({str(PROJECT_ROOT.resolve())!r})\n"
            "resolved_paths = {Path(entry or '.').resolve() for entry in sys.path}\n"
            "assert project_root not in resolved_paths, resolved_paths\n"
            "try:\n"
            "    import tools.hash_id_pseudonymizer\n"
            "except ModuleNotFoundError:\n"
            "    pass\n"
            "else:\n"
            "    raise AssertionError('project-root tools unexpectedly importable')\n"
        )
        subprocess.run(
            [sys.executable, "-c", probe_code],
            cwd=run_root,
            env=subprocess_env,
            check=True,
            capture_output=True,
            text=True,
        )

        input_path = run_root / "input.xlsx"
        standardized_path = run_root / "standardized.xlsx"
        cleaned_path = run_root / "cleaned.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["timestamp", "content", "like_count", "author"])
        sheet.append(["1678870952", "这个显示器挂灯使用体验确实很好", "2", "Portable User"])
        workbook.save(input_path)

        subprocess.run(
            [
                sys.executable,
                str(copied_skill / "scripts" / "standardize_excel_headers.py"),
                str(input_path),
                "--output",
                str(standardized_path),
                "--platform",
                "YouTube",
                "--project-name",
                "portable-test-project",
                "--initialize-project",
                "--project-store",
                str(run_root / "project-store"),
            ],
            cwd=run_root,
            env=subprocess_env,
            check=True,
            capture_output=True,
            text=True,
        )
        subprocess.run(
            [
                sys.executable,
                str(copied_skill / "scripts" / "clean_excel_comments.py"),
                str(standardized_path),
                "--target-header",
                "评论内容",
                "--output",
                str(cleaned_path),
            ],
            cwd=run_root,
            env=subprocess_env,
            check=True,
            text=True,
            capture_output=True,
        )

        self.assertTrue(cleaned_path.is_file())
        self.assertTrue(cleaned_path.with_suffix(".csv").is_file())
        cleaned = load_workbook(cleaned_path, data_only=False)
        self.assertEqual("评论日期", cleaned.active.cell(row=1, column=1).value)
        self.assertEqual("2023-03-15", cleaned.active.cell(row=2, column=1).value)
        self.assertEqual("这个显示器挂灯使用体验确实很好", cleaned.active.cell(row=2, column=2).value)
        hash_id = cleaned.active.cell(row=2, column=4).value
        self.assertIsInstance(hash_id, str)
        self.assertRegex(hash_id, r"^[0-9a-f]{64}$")

        self.assertNotIn("author", [cell.value for cell in cleaned.active[1]])
        temp_directory.cleanup()


if __name__ == "__main__":
    unittest.main()
