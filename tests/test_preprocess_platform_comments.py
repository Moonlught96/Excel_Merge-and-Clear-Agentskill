from __future__ import annotations

import csv
import json
import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.hash_id_pseudonymizer import (
    HashProjectContext,
    hash_display_name,
    hash_user_id,
    load_hash_id_config,
)
import tools.preprocess_platform_comments as preprocess_module
from tools.preprocess_platform_comments import (
    PlatformHeaderSignatureError,
    PlatformNotDetectedError,
    load_config,
    preprocess_workbook,
)
from tools.standardize_excel_headers import load_config as load_header_config
from tools.standardize_excel_headers import standardize_workbook


AMAZON_HEADERS = [
    "标题",
    "标题链接",
    "图片",
    "aprofile_链接",
    "名称",
    "aiconalt",
    "查看",
    "状态",
    "查看1",
    "asizebase",
    "crhelpfultext",
    "asizebase_链接",
    "asizebase2",
]
PREPROCESSED_HEADERS = ("评论日期", "评论内容", "电商平台评分", "点赞数", "名称")
RAKUTEN_PREPROCESSED_HEADERS = (
    "评论日期",
    "评论内容",
    "电商平台评分",
    "用户属性",
    "点赞数",
    "乐天市场昵称",
)
RAKUTEN_REVIEWER_TITLE_BODY_HEADERS = [
    "レビュータイトル",
    "評価",
    "レビュー本文",
    "レビュー投稿者",
    "レビュー投稿日",
    "注文日",
    "レビュアー属性",
    "参考になった数",
]
RAKUTEN_REVIEWER_DATE_BODY_HEADERS = [
    "レビュー投稿者",
    "評価",
    "投稿日",
    "レビュー本文",
    "レビュータイトル",
    "レビュアー属性",
    "参考になった数",
]
RAKUTEN_TITLE_DATE_BODY_REVIEWER_HEADERS = [
    "レビュータイトル",
    "評価",
    "レビュー投稿日",
    "レビュー本文",
    "レビュー投稿者",
    "注文日",
    "レビュアー属性",
    "参考になった数",
]
RAKUTEN_POSTER_TITLE_BODY_HEADERS = [
    "レビュータイトル",
    "評価",
    "レビュー投稿日",
    "投稿者名",
    "レビュー本文",
    "レビュアー属性",
    "参考になった数",
]
RAKUTEN_REVIEWER_NAME_TITLE_CONTENT_HEADERS = [
    "レビュアー名",
    "評価",
    "投稿日",
    "カラー",
    "レビュータイトル",
    "レビュー内容",
    "レビュアー属性",
    "参考になった数",
]
TWITTER_HEADERS = [
    "id",
    "created_at",
    "full_text",
    "media",
    "screen_name",
    "name",
    "profile_image_url",
    "user_id",
    "in_reply_to",
    "retweeted_status",
    "quoted_status",
    "media_tags",
    "favorite_count",
    "retweet_count",
    "bookmark_count",
    "quote_count",
    "reply_count",
    "views_count",
    "favorited",
    "retweeted",
    "bookmarked",
    "url",
    "metadata",
]
TWITTER_PREPROCESSED_HEADERS = (
    "评论日期",
    "评论内容",
    "点赞数",
    "子评论数/追评数",
    "Twitter用户ID",
    "Twitter昵称",
)
STANDARDIZED_HEADERS = (
    "评论日期",
    "评论内容",
    "产品名",
    "电商平台评分",
    "用户属性",
    "哈希ID",
    "点赞数",
    "子评论数/追评数",
    "一级评论",
    "二级评论",
    "三级评论",
)
STANDARDIZED_HASH_ID_INDEX = STANDARDIZED_HEADERS.index("哈希ID")
STANDARDIZED_LIKES_INDEX = STANDARDIZED_HEADERS.index("点赞数")


def write_amazon_source(path: Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(AMAZON_HEADERS)
    sheet.append(
        [
            "初めてのモニターライト",
            "https://example.invalid/title",
            "https://example.invalid/image",
            "https://example.invalid/profile",
            "Ricky",
            "4.0 颗星，最多 5 颗星",
            "2025年9月10日在日本发布评论",
            "已确认购买",
            "本文评论内容",
            "4 个人发现此评论有用",
            "有帮助",
            "https://example.invalid/helpful",
            "报告",
        ]
    )
    sheet.append(
        [
            None,
            "https://example.invalid/title2",
            None,
            None,
            "Ricky",
            "5.0 颗星，最多 5 颗星",
            "2026年4月26日在日本发布评论",
            "已确认购买",
            "只有正文的评论",
            None,
            "有帮助",
            None,
            "报告",
        ]
    )
    workbook.save(path)
    workbook.close()


def write_rakuten_source(path: Path, headers: list[str], rows: list[dict[str, object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sheet1"
    sheet.append(headers)
    for row in rows:
        sheet.append([row.get(header) for header in headers])
    workbook.save(path)
    workbook.close()


def write_twitter_source(path: Path, rows: list[dict[str, object]]) -> None:
    with path.open("w", newline="", encoding="utf-8-sig") as csv_file:
        writer = csv.writer(csv_file)
        writer.writerow(TWITTER_HEADERS)
        for row in rows:
            writer.writerow([row.get(header) for header in TWITTER_HEADERS])


class PreprocessPlatformCommentsTest(unittest.TestCase):
    def setUp(self) -> None:
        self.hash_context = HashProjectContext(
            project_id="project-screenbar",
            project_name="ScreenBar",
            key_version=1,
            key_fingerprint="test-fingerprint",
            secret_key=b"k" * 32,
        )
        self.hash_config = load_hash_id_config()

    def test_preprocesses_twitter_csv_with_exact_signature_and_hashes_account_ids(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-twitter-preprocess"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "twitter-screenbar.csv"
        write_twitter_source(
            source_path,
            [
                {
                    "id": "tweet-1",
                    "created_at": "2026-07-22T17:40:47+08:00",
                    "full_text": "ScreenBar works well.",
                    "screen_name": "screenbar-user-one",
                    "user_id": "1000000000000000001",
                    "favorite_count": "13",
                    "reply_count": "3",
                    "url": "https://x.example.invalid/post/1",
                },
                {
                    "id": "tweet-2",
                    "created_at": "2026-07-23T17:40:47+08:00",
                    "full_text": "BenQ is useful.",
                    "screen_name": "screenbar-user-two",
                    "user_id": "",
                    "favorite_count": "0",
                    "reply_count": "0",
                },
            ],
        )

        preprocessed = preprocess_workbook(
            source_path,
            load_config(),
            output_dir=tmp / "preprocessed",
            platform="Twitter",
        )
        preprocessed_book = load_workbook(preprocessed.output_xlsx, read_only=True, data_only=True)
        try:
            preprocessed_rows = list(preprocessed_book.worksheets[0].iter_rows(values_only=True))
        finally:
            preprocessed_book.close()

        self.assertEqual("twitter", preprocessed.platform)
        self.assertEqual(TWITTER_PREPROCESSED_HEADERS, preprocessed_rows[0])
        self.assertEqual(
            (
                "2026-07-22T17:40:47+08:00",
                "ScreenBar works well.",
                "13",
                "3",
                "1000000000000000001",
                "screenbar-user-one",
            ),
            preprocessed_rows[1],
        )
        self.assertNotIn("url", preprocessed_rows[0])
        self.assertNotIn("screenbar-user-one", preprocessed.summary_json.read_text(encoding="utf-8"))

        standardized = standardize_workbook(
            preprocessed.output_xlsx,
            load_header_config(),
            output_dir=tmp / "standardized",
            platform="twitter",
            hash_context=self.hash_context,
            hash_config=self.hash_config,
            product_name="ScreenBar系列",
        )
        standardized_book = load_workbook(standardized.output_xlsx, read_only=True, data_only=True)
        try:
            standardized_rows = list(standardized_book.worksheets[0].iter_rows(values_only=True))
        finally:
            standardized_book.close()

        self.assertEqual(STANDARDIZED_HEADERS, standardized_rows[0])
        self.assertEqual(
            hash_user_id(
                "1000000000000000001",
                "Twitter",
                self.hash_context,
                self.hash_config,
            ),
            standardized_rows[1][STANDARDIZED_HASH_ID_INDEX],
        )
        self.assertIsNone(standardized_rows[2][STANDARDIZED_HASH_ID_INDEX])
        self.assertEqual("2026-07-22", standardized_rows[1][0])
        self.assertEqual("ScreenBar系列", standardized_rows[1][2])
        self.assertNotIn("Twitter用户ID", standardized_rows[0])
        self.assertNotIn("Twitter昵称", standardized_rows[0])
        self.assertNotIn("screenbar-user-one", str(standardized_rows))

    def test_detects_amazon_signature_and_preprocesses_only_registered_fields(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-amazon-preprocess"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "amazon.xlsx"
        write_amazon_source(source_path)

        result = preprocess_workbook(
            source_path,
            load_config(),
            output_dir=tmp / "out",
        )

        self.assertEqual("amazon", result.platform)
        workbook = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        try:
            rows = list(workbook["Sheet1"].iter_rows(values_only=True))
        finally:
            workbook.close()

        self.assertEqual(PREPROCESSED_HEADERS, rows[0])
        self.assertEqual(
            (
                "2025-09-10",
                "初めてのモニターライト\n\n本文评论内容",
                "4.0",
                4,
                "Ricky",
            ),
            rows[1],
        )
        self.assertEqual(
            ("2026-04-26", "只有正文的评论", "5.0", None, "Ricky"),
            rows[2],
        )

        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        self.assertEqual("amazon", summary["platform"])
        self.assertNotIn("Ricky", result.summary_json.read_text(encoding="utf-8"))

    def test_rejects_an_unmatched_header_signature_without_guessing(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-amazon-preprocess-unmatched"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "unmatched.xlsx"
        workbook = Workbook()
        workbook.active.append(["标题", "名称", "aiconalt", "查看", "asizebase"])
        workbook.active.append(["标题", "昵称", "4.0 颗星，最多 5 颗星", "2025年9月10日", "4"])
        workbook.save(source_path)
        workbook.close()

        with self.assertRaisesRegex(PlatformNotDetectedError, "No configured platform signature matched"):
            preprocess_workbook(
                source_path,
                load_config(),
                output_dir=tmp / "out",
            )

    def test_rejects_a_selected_profile_when_a_required_header_is_duplicated(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-amazon-preprocess-duplicate-header"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "duplicate-header.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([*AMAZON_HEADERS, "标题"])
        sheet.append(
            [
                "标题一",
                None,
                None,
                None,
                "Ricky",
                "4.0 颗星，最多 5 颗星",
                "2025年9月10日在日本发布评论",
                None,
                "正文",
                "4 个人发现此评论有用",
                None,
                None,
                None,
                "标题二",
            ]
        )
        workbook.save(source_path)
        workbook.close()

        with self.assertRaises(PlatformHeaderSignatureError):
            preprocess_workbook(
                source_path,
                load_config(),
                output_dir=tmp / "out",
                platform="amazon",
            )

    def test_rejects_a_sheet_that_only_partially_resembles_amazon_schema(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-amazon-preprocess-extra-header"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "other-platform.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([*AMAZON_HEADERS, "其他平台字段"])
        sheet.append(
            [
                "标题",
                None,
                None,
                None,
                "Ricky",
                "4.0 颗星，最多 5 颗星",
                "2025年9月10日在日本发布评论",
                None,
                "正文",
                "4 个人发现此评论有用",
                None,
                None,
                None,
                "平台专属值",
            ]
        )
        workbook.save(source_path)
        workbook.close()

        with self.assertRaisesRegex(PlatformNotDetectedError, "No configured platform signature matched"):
            preprocess_workbook(
                source_path,
                load_config(),
                output_dir=tmp / "out",
            )

    def test_rejects_a_sheet_with_reordered_amazon_headers(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-amazon-preprocess-reordered-header"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "reordered.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(list(reversed(AMAZON_HEADERS)))
        sheet.append([None] * len(AMAZON_HEADERS))
        workbook.save(source_path)
        workbook.close()

        with self.assertRaisesRegex(PlatformNotDetectedError, "No configured platform signature matched"):
            preprocess_workbook(
                source_path,
                load_config(),
                output_dir=tmp / "out",
            )

    def test_preprocessed_amazon_name_becomes_a_platform_scoped_hash_id(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-amazon-preprocess-standardize"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "amazon.xlsx"
        write_amazon_source(source_path)

        preprocessed = preprocess_workbook(
            source_path,
            load_config(),
            output_dir=tmp / "preprocessed",
        )
        standardized = standardize_workbook(
            preprocessed.output_xlsx,
            load_header_config(),
            output_dir=tmp / "standardized",
            platform=preprocessed.platform,
            hash_context=self.hash_context,
            hash_config=self.hash_config,
        )

        workbook = load_workbook(standardized.output_xlsx, read_only=True, data_only=True)
        try:
            rows = list(workbook["Sheet1"].iter_rows(values_only=True))
        finally:
            workbook.close()

        self.assertEqual(STANDARDIZED_HEADERS, rows[0])
        expected_hash = hash_display_name("Ricky", "amazon", self.hash_context, self.hash_config)
        self.assertEqual(expected_hash, rows[1][STANDARDIZED_HASH_ID_INDEX])
        self.assertEqual(expected_hash, rows[2][STANDARDIZED_HASH_ID_INDEX])
        self.assertEqual("4.0", rows[1][3])
        self.assertEqual(4, rows[1][STANDARDIZED_LIKES_INDEX])
        self.assertNotIn("名称", rows[0])
        self.assertNotIn("Ricky", str(rows))

    def test_preprocesses_each_registered_rakuten_header_variant_without_guessing(self) -> None:
        cases = (
            (
                "reviewer-title-body",
                RAKUTEN_REVIEWER_TITLE_BODY_HEADERS,
                {
                    "レビュータイトル": "标题",
                    "評価": "4.5",
                    "レビュー本文": "正文",
                    "レビュー投稿者": "楽天太郎さん",
                    "レビュー投稿日": "01/23/2026",
                    "注文日": "01/01/2026",
                    "レビュアー属性": "女性, 50代",
                    "参考になった数": "3人",
                },
            ),
            (
                "reviewer-date-body",
                RAKUTEN_REVIEWER_DATE_BODY_HEADERS,
                {
                    "レビュー投稿者": "楽天太郎さん",
                    "評価": "4.5",
                    "投稿日": "01/23/2026",
                    "レビュー本文": "正文",
                    "レビュータイトル": "标题",
                    "レビュアー属性": "女性, 50代",
                    "参考になった数": "3人",
                },
            ),
            (
                "title-date-body-reviewer",
                RAKUTEN_TITLE_DATE_BODY_REVIEWER_HEADERS,
                {
                    "レビュータイトル": "标题",
                    "評価": "4.5",
                    "レビュー投稿日": "01/23/2026",
                    "レビュー本文": "正文",
                    "レビュー投稿者": "楽天太郎さん",
                    "注文日": "01/01/2026",
                    "レビュアー属性": "女性, 50代",
                    "参考になった数": "3人",
                },
            ),
            (
                "poster-title-body",
                RAKUTEN_POSTER_TITLE_BODY_HEADERS,
                {
                    "レビュータイトル": "标题",
                    "評価": "4.5",
                    "レビュー投稿日": "01/23/2026",
                    "投稿者名": "楽天太郎さん",
                    "レビュー本文": "正文",
                    "レビュアー属性": "女性, 50代",
                    "参考になった数": "3人",
                },
            ),
            (
                "reviewer-name-title-content",
                RAKUTEN_REVIEWER_NAME_TITLE_CONTENT_HEADERS,
                {
                    "レビュアー名": "楽天太郎さん",
                    "評価": "4.5",
                    "投稿日": "01/23/2026",
                    "カラー": "ブラック",
                    "レビュータイトル": "标题",
                    "レビュー内容": "正文",
                    "レビュアー属性": "女性, 50代",
                    "参考になった数": "3人",
                },
            ),
        )
        for case_name, headers, row in cases:
            with self.subTest(case_name=case_name):
                tmp = Path.cwd() / ".tmp-tests" / f"case-rakuten-{case_name}"
                tmp.mkdir(parents=True, exist_ok=True)
                source_path = tmp / "rakuten.xlsx"
                write_rakuten_source(source_path, headers, [row])

                result = preprocess_workbook(
                    source_path,
                    load_config(),
                    output_dir=tmp / "out",
                    platform="rakuten",
                )

                self.assertEqual("rakuten", result.platform)
                workbook = load_workbook(result.output_xlsx, read_only=True, data_only=True)
                try:
                    rows = list(workbook["Sheet1"].iter_rows(values_only=True))
                finally:
                    workbook.close()

                self.assertEqual(RAKUTEN_PREPROCESSED_HEADERS, rows[0])
                self.assertEqual(
                    (
                        "2026-01-23",
                        "标题\n\n正文",
                        "4.5",
                        "女性 50代",
                        3,
                        "楽天太郎さん",
                    ),
                    rows[1],
                )

    def test_rakuten_keeps_only_gender_and_age_and_skips_anonymous_buyer_hashes(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-rakuten-attributes-and-hash"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "rakuten.xlsx"
        write_rakuten_source(
            source_path,
            RAKUTEN_REVIEWER_TITLE_BODY_HEADERS,
            [
                {
                    "レビュータイトル": "标题一",
                    "評価": "5",
                    "レビュー本文": "正文一",
                    "レビュー投稿者": "購入者さん",
                    "レビュー投稿日": "11/16/2025",
                    "注文日": "11/04/2025",
                    "レビュアー属性": "実用品・普段使い｜自分用｜はじめて",
                    "参考になった数": "",
                },
                {
                    "レビュータイトル": "标题二",
                    "評価": "4",
                    "レビュー本文": "正文二",
                    "レビュー投稿者": "楽天太郎さん",
                    "レビュー投稿日": "11/17/2025",
                    "注文日": "11/05/2025",
                    "レビュアー属性": "男性 50代｜自分用",
                    "参考になった数": "1人",
                },
                {
                    "レビュータイトル": "标题三",
                    "評価": "5",
                    "レビュー本文": "正文三",
                    "レビュー投稿者": "楽天次郎さん",
                    "レビュー投稿日": "11/18/2025",
                    "注文日": "11/06/2025",
                    "レビュアー属性": "男性, 70代以上",
                    "参考になった数": "2人",
                },
            ],
        )

        preprocessed = preprocess_workbook(
            source_path,
            load_config(),
            output_dir=tmp / "preprocessed",
            platform="乐天市场",
        )
        preprocessed_book = load_workbook(preprocessed.output_xlsx, read_only=True, data_only=True)
        try:
            preprocessed_rows = list(preprocessed_book["Sheet1"].iter_rows(values_only=True))
        finally:
            preprocessed_book.close()

        self.assertEqual(None, preprocessed_rows[1][5])
        self.assertEqual(None, preprocessed_rows[1][3])
        self.assertEqual("男性 50代", preprocessed_rows[2][3])
        self.assertEqual("楽天太郎さん", preprocessed_rows[2][5])
        self.assertEqual("男性 70代以上", preprocessed_rows[3][3])

        standardized = standardize_workbook(
            preprocessed.output_xlsx,
            load_header_config(),
            output_dir=tmp / "standardized",
            platform=preprocessed.platform,
            hash_context=self.hash_context,
            hash_config=self.hash_config,
        )
        standardized_book = load_workbook(standardized.output_xlsx, read_only=True, data_only=True)
        try:
            standardized_rows = list(standardized_book["Sheet1"].iter_rows(values_only=True))
        finally:
            standardized_book.close()

        self.assertEqual(STANDARDIZED_HEADERS, standardized_rows[0])
        self.assertIsNone(standardized_rows[1][STANDARDIZED_HASH_ID_INDEX])
        self.assertEqual(
            hash_display_name("楽天太郎さん", "rakuten", self.hash_context, self.hash_config),
            standardized_rows[2][STANDARDIZED_HASH_ID_INDEX],
        )
        self.assertEqual("男性 50代", standardized_rows[2][4])
        self.assertEqual("男性 70代以上", standardized_rows[3][4])
        self.assertNotIn("購入者さん", str(standardized_rows))

    def test_rejects_rakuten_header_variants_with_extra_or_reordered_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-rakuten-invalid-signature"
        tmp.mkdir(parents=True, exist_ok=True)
        source_path = tmp / "rakuten-invalid.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append([*RAKUTEN_REVIEWER_TITLE_BODY_HEADERS, "额外字段"])
        sheet.append([None] * (len(RAKUTEN_REVIEWER_TITLE_BODY_HEADERS) + 1))
        workbook.save(source_path)
        workbook.close()

        with self.assertRaisesRegex(PlatformHeaderSignatureError, "Configured platform signature does not match"):
            preprocess_workbook(
                source_path,
                load_config(),
                output_dir=tmp / "out",
                platform="rakuten",
            )

    def test_preprocesses_and_merges_registered_rakuten_variants_in_input_order(self) -> None:
        processor = getattr(preprocess_module, "preprocess_and_merge_workbooks", None)
        self.assertIsNotNone(
            processor,
            "Mixed registered platform variants need a deterministic preprocessed merge entry point.",
        )

        tmp = Path.cwd() / ".tmp-tests" / "case-rakuten-mixed-variant-merge"
        tmp.mkdir(parents=True, exist_ok=True)
        first_path = tmp / "first.xlsx"
        second_path = tmp / "second.xlsx"
        output_path = tmp / "merged.xlsx"
        output_path.unlink(missing_ok=True)
        output_path.with_suffix(".summary.json").unlink(missing_ok=True)
        write_rakuten_source(
            first_path,
            RAKUTEN_REVIEWER_TITLE_BODY_HEADERS,
            [
                {
                    "レビュータイトル": "标题一",
                    "評価": "4",
                    "レビュー本文": "正文一",
                    "レビュー投稿者": "楽天一郎さん",
                    "レビュー投稿日": "01/23/2026",
                    "注文日": "01/01/2026",
                    "レビュアー属性": "男性, 30代",
                    "参考になった数": "2人",
                }
            ],
        )
        write_rakuten_source(
            second_path,
            RAKUTEN_REVIEWER_NAME_TITLE_CONTENT_HEADERS,
            [
                {
                    "レビュアー名": "楽天二郎さん",
                    "評価": "5",
                    "投稿日": "2026/02/02",
                    "カラー": "黒",
                    "レビュータイトル": "标题二",
                    "レビュー内容": "正文二",
                    "レビュアー属性": "女性, 40代",
                    "参考になった数": "1人",
                }
            ],
        )

        result = processor(
            [first_path, second_path],
            load_config(),
            output_path,
            platform="rakuten",
            overwrite=False,
        )

        merged = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        try:
            rows = list(merged["总表"].iter_rows(values_only=True))
        finally:
            merged.close()

        self.assertEqual(RAKUTEN_PREPROCESSED_HEADERS, rows[0])
        self.assertEqual(
            [
                ("2026-01-23", "标题一\n\n正文一", "4", "男性 30代", 2, "楽天一郎さん"),
                ("2026-02-02", "标题二\n\n正文二", "5", "女性 40代", 1, "楽天二郎さん"),
            ],
            rows[1:],
        )
        self.assertEqual(2, result.files_processed)
        self.assertEqual(2, result.data_rows_written)

    def test_mixed_variant_merge_rejects_unregistered_headers_without_partial_output(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-rakuten-mixed-variant-merge-reject"
        tmp.mkdir(parents=True, exist_ok=True)
        valid_path = tmp / "valid.xlsx"
        invalid_path = tmp / "invalid.xlsx"
        output_path = tmp / "merged.xlsx"
        output_path.unlink(missing_ok=True)
        output_path.with_suffix(".summary.json").unlink(missing_ok=True)
        write_rakuten_source(
            valid_path,
            RAKUTEN_REVIEWER_TITLE_BODY_HEADERS,
            [{header: None for header in RAKUTEN_REVIEWER_TITLE_BODY_HEADERS}],
        )
        invalid_headers = [*RAKUTEN_REVIEWER_NAME_TITLE_CONTENT_HEADERS, "未登记字段"]
        write_rakuten_source(
            invalid_path,
            invalid_headers,
            [{header: None for header in invalid_headers}],
        )

        with self.assertRaises(PlatformHeaderSignatureError):
            preprocess_module.preprocess_and_merge_workbooks(
                [valid_path, invalid_path],
                load_config(),
                output_path,
                platform="rakuten",
                overwrite=False,
            )

        self.assertFalse(output_path.exists())
        self.assertFalse(output_path.with_suffix(".summary.json").exists())


if __name__ == "__main__":
    unittest.main()
