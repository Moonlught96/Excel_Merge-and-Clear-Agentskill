import csv
import io
import json
import os
from pathlib import Path
import subprocess
import sys
import tempfile
import unittest
from unittest.mock import patch

from openpyxl import load_workbook

from tools.reddit_free_csv import FreeComment, FreeRedditExport
from tools.reddit_saved_html import HtmlComment, SavedRedditHtml
from tools.output_path_safety import OutputPathConflictError
from tools import reconstruct_reddit_comments
from tools.reconstruct_reddit_comments import (
    OUTPUT_HEADERS,
    reconstruct_rows,
    write_outputs,
)
from tools.reddit_json_text_merge import JSON_TEXT_OUTPUT_HEADERS


class ReconstructRedditRowsTests(unittest.TestCase):
    def test_module_imports_from_direct_script_directory_context(self) -> None:
        tools_directory = Path(__file__).resolve().parents[1] / "tools"

        completed = subprocess.run(
            [sys.executable, "-c", "import reconstruct_reddit_comments"],
            cwd=tools_directory,
            capture_output=True,
            text=True,
            check=False,
        )

        self.assertEqual(0, completed.returncode, completed.stderr)

    def free(self, *comments: FreeComment, post_id: str = "post1") -> FreeRedditExport:
        return FreeRedditExport(
            title="Exact title",
            body="Exact body",
            url="https://www.reddit.com/r/test/comments/post1/example/",
            post_id=post_id,
            comments=comments,
        )

    def comment(
        self,
        comment_id: str,
        *,
        author: str = "free-author",
        time: str = "free-time",
        text: str = "free-comment",
        url: str | None = None,
    ) -> FreeComment:
        return FreeComment(
            author=author,
            time=time,
            comment=text,
            comment_url=url or f"https://reddit.com/comment/{comment_id}/",
            comment_id=comment_id,
        )

    def html(
        self,
        *comments: HtmlComment,
        post_id: str = "post1",
        post_author: str = "html-author",
        post_score: str = "42",
        post_comment_count: str = "2",
    ) -> SavedRedditHtml:
        return SavedRedditHtml(
            post_id=post_id,
            post_author=post_author,
            post_score=post_score,
            post_comment_count=post_comment_count,
            comments={comment.comment_id: comment for comment in comments},
        )

    def html_comment(
        self,
        comment_id: str,
        *,
        parent_id: str = "post1",
        level: int | None = 0,
        score: str = "7",
    ) -> HtmlComment:
        return HtmlComment(
            comment_id=comment_id,
            parent_id=parent_id,
            thread_level=level,
            score=score,
        )

    def test_output_headers_have_exact_order_and_rows_have_no_extra_fields(self) -> None:
        expected = (
            "Title",
            "Post Body",
            "Post URL",
            "Post Author",
            "Post Score",
            "Post Comment Count",
            "Author",
            "Time",
            "Score",
            "Thread Level",
            "Is Reply",
            "Comment",
            "Comment URL",
            "Comment ID",
            "Parent ID",
        )
        rows = reconstruct_rows(
            self.free(self.comment("c1")),
            self.html(self.html_comment("c1")),
        )

        self.assertEqual(expected, OUTPUT_HEADERS)
        self.assertEqual(15, len(OUTPUT_HEADERS))
        self.assertEqual(list(expected), list(rows[0]))

    def test_preserves_source_order_and_marks_root_and_replies(self) -> None:
        free = self.free(self.comment("second"), self.comment("first"))
        html = self.html(
            self.html_comment("first", level=2, parent_id="parent"),
            self.html_comment("second", level=0),
        )

        rows = reconstruct_rows(free, html)

        self.assertEqual(["second", "first"], [row["Comment ID"] for row in rows])
        self.assertEqual(["No", "Yes"], [row["Is Reply"] for row in rows])
        self.assertEqual([0, 2], [row["Thread Level"] for row in rows])

    def test_only_exact_zero_thread_level_is_not_a_reply(self) -> None:
        rows = reconstruct_rows(
            self.free(self.comment("synthetic-negative")),
            self.html(self.html_comment("synthetic-negative", level=-1)),
        )

        self.assertEqual("Yes", rows[0]["Is Reply"])

    def test_blank_comment_score_is_allowed_and_preserved(self) -> None:
        rows = reconstruct_rows(
            self.free(self.comment("c1")),
            self.html(self.html_comment("c1", score="")),
        )

        self.assertEqual("", rows[0]["Score"])

    def test_nonempty_html_post_values_including_zero_override_fallbacks(self) -> None:
        rows = reconstruct_rows(
            self.free(self.comment("c1")),
            self.html(
                self.html_comment("c1"),
                post_author="html",
                post_score="0",
                post_comment_count="0",
            ),
            post_author="wrong",
            post_score="wrong",
            post_comment_count="wrong",
        )

        self.assertEqual("html", rows[0]["Post Author"])
        self.assertEqual("0", rows[0]["Post Score"])
        self.assertEqual("0", rows[0]["Post Comment Count"])

    def test_explicit_post_values_are_fallbacks_for_empty_html_values(self) -> None:
        rows = reconstruct_rows(
            self.free(self.comment("c1")),
            self.html(
                self.html_comment("c1"),
                post_author="",
                post_score="",
                post_comment_count="",
            ),
            post_author="fallback-author",
            post_score="fallback-score",
            post_comment_count="fallback-count",
        )

        self.assertEqual("fallback-author", rows[0]["Post Author"])
        self.assertEqual("fallback-score", rows[0]["Post Score"])
        self.assertEqual("fallback-count", rows[0]["Post Comment Count"])

    def test_each_missing_required_post_field_names_needed_cli_value(self) -> None:
        cases = (
            ("Post Author", "post_author", {"post_author": ""}),
            ("Post Score", "post_score", {"post_score": ""}),
            (
                "Post Comment Count",
                "post_comment_count",
                {"post_comment_count": ""},
            ),
        )
        for field_name, cli_name, html_override in cases:
            with self.subTest(field_name=field_name):
                html_values = {
                    "post_author": "author",
                    "post_score": "score",
                    "post_comment_count": "count",
                }
                html_values.update(html_override)
                with self.assertRaisesRegex(
                    ValueError, rf"{field_name}.*{cli_name}"
                ):
                    reconstruct_rows(
                        self.free(self.comment("c1")),
                        self.html(self.html_comment("c1"), **html_values),
                    )

    def test_whitespace_only_html_post_fields_are_missing(self) -> None:
        cases = (
            ("Post Author", "post_author"),
            ("Post Score", "post_score"),
            ("Post Comment Count", "post_comment_count"),
        )
        for field_name, field_key in cases:
            with self.subTest(field_name=field_name):
                html_values = {
                    "post_author": "author",
                    "post_score": "score",
                    "post_comment_count": "count",
                }
                html_values[field_key] = " \t "
                with self.assertRaisesRegex(ValueError, field_name):
                    reconstruct_rows(
                        self.free(self.comment("c1")),
                        self.html(self.html_comment("c1"), **html_values),
                    )

    def test_whitespace_only_explicit_post_fallbacks_are_missing(self) -> None:
        cases = (
            ("Post Author", "post_author"),
            ("Post Score", "post_score"),
            ("Post Comment Count", "post_comment_count"),
        )
        for field_name, field_key in cases:
            with self.subTest(field_name=field_name):
                html_values = {
                    "post_author": "author",
                    "post_score": "score",
                    "post_comment_count": "count",
                }
                html_values[field_key] = ""
                fallback_values = {field_key: " \r\n "}
                with self.assertRaisesRegex(ValueError, field_name):
                    reconstruct_rows(
                        self.free(self.comment("c1")),
                        self.html(self.html_comment("c1"), **html_values),
                        **fallback_values,
                    )

    def test_whitespace_html_uses_valid_fallbacks_without_stripping_them(self) -> None:
        rows = reconstruct_rows(
            self.free(self.comment("c1")),
            self.html(
                self.html_comment("c1"),
                post_author=" ",
                post_score="\t",
                post_comment_count="\r\n",
            ),
            post_author=" fallback author ",
            post_score=" 7 ",
            post_comment_count=" 1 ",
        )

        self.assertEqual(" fallback author ", rows[0]["Post Author"])
        self.assertEqual(" 7 ", rows[0]["Post Score"])
        self.assertEqual(" 1 ", rows[0]["Post Comment Count"])

    def test_post_id_mismatch_names_both_ids(self) -> None:
        with self.assertRaisesRegex(ValueError, r"free-id.*html-id"):
            reconstruct_rows(
                self.free(self.comment("c1"), post_id="free-id"),
                self.html(self.html_comment("c1"), post_id="html-id"),
            )

    def test_one_missing_html_comment_is_reported(self) -> None:
        with self.assertRaisesRegex(ValueError, r"Missing HTML comments.*absent"):
            reconstruct_rows(self.free(self.comment("absent")), self.html())

    def test_all_missing_html_comment_ids_are_reported(self) -> None:
        with self.assertRaises(ValueError) as caught:
            reconstruct_rows(
                self.free(self.comment("missing1"), self.comment("missing2")),
                self.html(),
            )

        message = str(caught.exception)
        self.assertIn("Missing HTML comments", message)
        self.assertIn("missing1", message)
        self.assertIn("missing2", message)

    def test_empty_parent_and_none_depth_are_both_reported_as_invalid_hierarchy(self) -> None:
        free = self.free(self.comment("no-parent"), self.comment("no-depth"))
        html = self.html(
            self.html_comment("no-parent", parent_id=""),
            self.html_comment("no-depth", level=None),
        )

        with self.assertRaises(ValueError) as caught:
            reconstruct_rows(free, html)

        message = str(caught.exception)
        self.assertIn("Invalid hierarchy", message)
        self.assertIn("no-parent", message)
        self.assertIn("no-depth", message)

    def test_missing_node_and_invalid_hierarchy_share_one_error(self) -> None:
        free = self.free(self.comment("missing"), self.comment("invalid"))
        html = self.html(self.html_comment("invalid", parent_id=""))

        with self.assertRaises(ValueError) as caught:
            reconstruct_rows(free, html)

        message = str(caught.exception)
        self.assertIn("Missing HTML comments", message)
        self.assertIn("missing", message)
        self.assertIn("Invalid hierarchy", message)
        self.assertIn("invalid", message)

    def test_join_is_exactly_by_comment_id_not_similar_content_or_author(self) -> None:
        free = self.free(
            self.comment("wanted", author="same", text="same text")
        )
        html = self.html(self.html_comment("different"))

        with self.assertRaises(ValueError) as caught:
            reconstruct_rows(free, html)

        self.assertIn("wanted", str(caught.exception))
        self.assertNotIn("different", str(caught.exception))

    def test_preserves_all_free_content_fields_exactly(self) -> None:
        free_comment = self.comment(
            "c1",
            author="=AUTHOR()",
            time="2026-01-01\n12:34",
            text="first line\n=HYPERLINK(\"x\")",
            url="=FORMULA-LIKE-URL",
        )

        row = reconstruct_rows(
            self.free(free_comment),
            self.html(self.html_comment("c1")),
        )[0]

        self.assertEqual("Exact title", row["Title"])
        self.assertEqual("Exact body", row["Post Body"])
        self.assertEqual(
            "https://www.reddit.com/r/test/comments/post1/example/",
            row["Post URL"],
        )
        self.assertEqual("=AUTHOR()", row["Author"])
        self.assertEqual("2026-01-01\n12:34", row["Time"])
        self.assertEqual("first line\n=HYPERLINK(\"x\")", row["Comment"])
        self.assertEqual("=FORMULA-LIKE-URL", row["Comment URL"])


class RedditOutputTests(unittest.TestCase):
    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary_directory.cleanup)
        self.directory = Path(self.temporary_directory.name)
        self.input_csv = self.directory / "input.csv"
        self.input_html = self.directory / "input.html"
        self.input_csv.write_bytes(b"original csv input")
        self.input_html.write_bytes(b"original html input")
        self.output_xlsx = self.directory / "output.xlsx"
        self.output_csv = self.directory / "output.csv"

    def row(self, **overrides: str | int) -> dict[str, str | int]:
        values: dict[str, str | int] = {
            "Title": "A title",
            "Post Body": "A body",
            "Post URL": "https://reddit.com/r/test/comments/post1/",
            "Post Author": "post-author",
            "Post Score": "10",
            "Post Comment Count": "1",
            "Author": "comment-author",
            "Time": "2026-07-23",
            "Score": "5",
            "Thread Level": 0,
            "Is Reply": "No",
            "Comment": "A comment",
            "Comment URL": "https://reddit.com/comment/c1/",
            "Comment ID": "c1",
            "Parent ID": "post1",
        }
        values.update(overrides)
        return values

    def json_text_row(self, **overrides: str | int) -> dict[str, str | int]:
        values: dict[str, str | int] = {
            "Title": "A title",
            "Post Body": "A body",
            "Post Author": "post-author",
            "Post Time": "8 hours ago",
            "Post Score": 10,
            "Post Comment Count": 1,
            "Author": "comment-author",
            "Time": "2026-07-23",
            "Score": 5,
            "Thread Level": 0,
            "Is Reply": "No",
            "Comment": "A comment",
            "Comment ID": "c1",
            "Parent ID": "post1",
        }
        values.update(overrides)
        return values

    def write(self, rows: list[dict[str, str | int]], *, overwrite: bool = False) -> None:
        write_outputs(
            rows,
            headers=OUTPUT_HEADERS,
            input_paths=(self.input_csv, self.input_html),
            output_xlsx=self.output_xlsx,
            output_csv=self.output_csv,
            overwrite=overwrite,
        )

    def csv_rows(self) -> list[list[str]]:
        with self.output_csv.open("r", encoding="utf-8-sig", newline="") as handle:
            return list(csv.reader(handle))

    def test_writes_matching_exact_headers_row_count_and_values(self) -> None:
        rows = [self.row(), self.row(**{"Comment ID": "c2", "Thread Level": 2})]

        self.write(rows)

        workbook = load_workbook(self.output_xlsx, data_only=False)
        sheet = workbook.active
        xlsx_rows = list(sheet.iter_rows(values_only=True))
        csv_rows = self.csv_rows()
        self.assertEqual("Reddit Comments", sheet.title)
        self.assertEqual(list(OUTPUT_HEADERS), list(xlsx_rows[0]))
        self.assertEqual(list(OUTPUT_HEADERS), csv_rows[0])
        self.assertEqual(len(rows) + 1, len(xlsx_rows))
        self.assertEqual(len(rows) + 1, len(csv_rows))
        for row_number, source_row in enumerate(rows, start=1):
            expected = [source_row[header] for header in OUTPUT_HEADERS]
            self.assertEqual(expected, list(xlsx_rows[row_number]))
            self.assertEqual([str(value) for value in expected], csv_rows[row_number])

    def test_writes_json_text_headers_integer_scores_and_formula_like_text(self) -> None:
        json_path = self.directory / "export.json"
        page_text_path = self.directory / "page.txt"
        json_path.write_text("{}", encoding="utf-8")
        page_text_path.write_text("page", encoding="utf-8")
        rows = [
            {
                "Title": "Title",
                "Post Body": "=FORMULA-LIKE-JSON-TEXT",
                "Post Author": "poster",
                "Post Time": "8 hours ago",
                "Post Score": 12,
                "Post Comment Count": 2,
                "Author": "commenter",
                "Time": "2026-07-23",
                "Score": 7,
                "Thread Level": 0,
                "Is Reply": "No",
                "Comment": "Comment",
                "Comment ID": "c1",
                "Parent ID": "p1",
            }
        ]

        write_outputs(
            rows,
            headers=JSON_TEXT_OUTPUT_HEADERS,
            input_paths=(json_path, page_text_path),
            output_xlsx=self.output_xlsx,
            output_csv=self.output_csv,
            overwrite=False,
        )

        sheet = load_workbook(self.output_xlsx, data_only=False).active
        self.assertEqual(JSON_TEXT_OUTPUT_HEADERS, list(sheet.values)[0])
        self.assertEqual(12, sheet.cell(2, 5).value)
        self.assertEqual(7, sheet.cell(2, 9).value)
        self.assertEqual("s", sheet.cell(2, 2).data_type)
        self.assertEqual("=FORMULA-LIKE-JSON-TEXT", sheet.cell(2, 2).value)
        csv_rows = self.csv_rows()
        self.assertEqual(list(JSON_TEXT_OUTPUT_HEADERS), csv_rows[0])
        self.assertEqual("12", csv_rows[1][4])
        self.assertEqual("7", csv_rows[1][8])

    def test_exact_xlsx_integer_boundaries_round_trip_in_both_outputs(self) -> None:
        maximum = 2**53 - 1
        minimum = -maximum
        rows = [
            self.json_text_row(
                **{
                    "Post Score": maximum,
                    "Post Comment Count": maximum,
                    "Score": minimum,
                    "Thread Level": maximum,
                }
            )
        ]

        write_outputs(
            rows,
            headers=JSON_TEXT_OUTPUT_HEADERS,
            input_paths=(self.input_csv, self.input_html),
            output_xlsx=self.output_xlsx,
            output_csv=self.output_csv,
            overwrite=False,
        )

        sheet_values = list(load_workbook(self.output_xlsx).active.values)[1]
        csv_values = self.csv_rows()[1]
        for header, expected in (
            ("Post Score", maximum),
            ("Post Comment Count", maximum),
            ("Score", minimum),
            ("Thread Level", maximum),
        ):
            column = JSON_TEXT_OUTPUT_HEADERS.index(header)
            self.assertEqual(expected, sheet_values[column])
            self.assertEqual(str(expected), csv_values[column])

    def test_out_of_range_integers_are_rejected_before_any_output_mutation(
        self,
    ) -> None:
        old_xlsx = b"old xlsx"
        old_csv = b"old csv"
        numeric_headers = (
            "Post Score",
            "Post Comment Count",
            "Score",
            "Thread Level",
        )
        for index, header in enumerate(numeric_headers):
            with self.subTest(header=header):
                self.output_xlsx.write_bytes(old_xlsx)
                self.output_csv.write_bytes(old_csv)
                private_value = 9007199254740993 + index
                with self.assertRaisesRegex(
                    ValueError,
                    rf"^XLSX integer is outside exact range at data row 1, column {header}$",
                ) as caught:
                    write_outputs(
                        [self.json_text_row(**{header: private_value})],
                        headers=JSON_TEXT_OUTPUT_HEADERS,
                        input_paths=(self.input_csv, self.input_html),
                        output_xlsx=self.output_xlsx,
                        output_csv=self.output_csv,
                        overwrite=True,
                    )
                self.assertNotIn(str(private_value), str(caught.exception))
                self.assertEqual(old_xlsx, self.output_xlsx.read_bytes())
                self.assertEqual(old_csv, self.output_csv.read_bytes())
                self.assertEqual(
                    {"input.csv", "input.html", "output.xlsx", "output.csv"},
                    {path.name for path in self.directory.iterdir()},
                )

    def test_csv_has_utf8_bom_and_round_trips_special_content(self) -> None:
        special = '中文 😀, "quoted"\nsecond line'
        self.write([self.row(**{"Comment": special})])

        self.assertTrue(self.output_csv.read_bytes().startswith(b"\xef\xbb\xbf"))
        self.assertEqual(special, self.csv_rows()[1][OUTPUT_HEADERS.index("Comment")])
        workbook = load_workbook(self.output_xlsx)
        self.assertEqual(
            special,
            workbook.active.cell(2, OUTPUT_HEADERS.index("Comment") + 1).value,
        )

    def test_formula_like_xlsx_values_are_text(self) -> None:
        markers = ("=SUM(1,2)", "+123", "-456", "@mention")
        rows = [
            self.row(**{"Comment ID": f"c{index}", "Comment": value})
            for index, value in enumerate(markers)
        ]

        self.write(rows)

        sheet = load_workbook(self.output_xlsx, data_only=False).active
        column = OUTPUT_HEADERS.index("Comment") + 1
        for row_number, expected in enumerate(markers, start=2):
            cell = sheet.cell(row_number, column)
            self.assertEqual(expected, cell.value)
            self.assertEqual("s", cell.data_type)

    def test_xlsx_maximum_length_string_round_trips_in_both_outputs(self) -> None:
        maximum_length = "界" * 32_767

        self.write([self.row(**{"Comment": maximum_length})])

        comment_column = OUTPUT_HEADERS.index("Comment")
        self.assertEqual(
            maximum_length,
            load_workbook(self.output_xlsx).active.cell(
                2, comment_column + 1
            ).value,
        )
        self.assertEqual(maximum_length, self.csv_rows()[1][comment_column])

    def test_xlsx_overlength_string_is_rejected_before_outputs_change(self) -> None:
        overlength = "private-value-" + ("界" * 32_768)
        self.output_xlsx.write_bytes(b"old xlsx")
        self.output_csv.write_bytes(b"old csv")

        with self.assertRaisesRegex(
            ValueError,
            r"data row 1.*Comment",
        ) as caught:
            self.write(
                [self.row(**{"Comment": overlength})],
                overwrite=True,
            )

        self.assertNotIn("private-value", str(caught.exception))
        self.assertEqual(b"old xlsx", self.output_xlsx.read_bytes())
        self.assertEqual(b"old csv", self.output_csv.read_bytes())
        self.assertEqual(
            {"input.csv", "input.html", "output.xlsx", "output.csv"},
            {path.name for path in self.directory.iterdir()},
        )

    def test_missing_optional_score_is_written_as_blank_in_both_outputs(self) -> None:
        self.write([self.row(**{"Score": ""})])

        sheet = load_workbook(self.output_xlsx).active
        score_column = OUTPUT_HEADERS.index("Score") + 1
        self.assertIsNone(sheet.cell(2, score_column).value)
        self.assertEqual("", self.csv_rows()[1][score_column - 1])

    def test_existing_outputs_are_rejected_without_overwrite_and_unchanged(self) -> None:
        self.output_xlsx.write_bytes(b"old xlsx")
        self.output_csv.write_bytes(b"old csv")

        with self.assertRaises(OutputPathConflictError):
            self.write([self.row()])

        self.assertEqual(b"old xlsx", self.output_xlsx.read_bytes())
        self.assertEqual(b"old csv", self.output_csv.read_bytes())

    def test_overwrite_true_updates_both_outputs(self) -> None:
        self.output_xlsx.write_bytes(b"old xlsx")
        self.output_csv.write_bytes(b"old csv")

        self.write([self.row(**{"Comment": "replacement"})], overwrite=True)

        self.assertEqual(
            "replacement",
            load_workbook(self.output_xlsx).active.cell(
                2, OUTPUT_HEADERS.index("Comment") + 1
            ).value,
        )
        self.assertEqual(
            "replacement",
            self.csv_rows()[1][OUTPUT_HEADERS.index("Comment")],
        )

    def test_input_output_alias_is_rejected_even_with_overwrite(self) -> None:
        with self.assertRaises(OutputPathConflictError):
            write_outputs(
                [self.row()],
                headers=OUTPUT_HEADERS,
                input_paths=(self.input_csv, self.input_html),
                output_xlsx=self.input_csv,
                output_csv=self.output_csv,
                overwrite=True,
            )

        self.assertEqual(b"original csv input", self.input_csv.read_bytes())
        self.assertFalse(self.output_csv.exists())

    def test_duplicate_output_paths_are_rejected(self) -> None:
        with self.assertRaises(OutputPathConflictError):
            write_outputs(
                [self.row()],
                headers=OUTPUT_HEADERS,
                input_paths=(self.input_csv, self.input_html),
                output_xlsx=self.output_xlsx,
                output_csv=self.output_xlsx,
                overwrite=False,
            )

        self.assertFalse(self.output_xlsx.exists())

    def test_swapped_output_suffixes_are_rejected_before_writes(self) -> None:
        swapped_xlsx = self.directory / "swapped.csv"
        swapped_csv = self.directory / "swapped.xlsx"

        with self.assertRaises(ValueError):
            write_outputs(
                [self.row()],
                headers=OUTPUT_HEADERS,
                input_paths=(self.input_csv, self.input_html),
                output_xlsx=swapped_xlsx,
                output_csv=swapped_csv,
                overwrite=False,
            )

        self.assertFalse(swapped_xlsx.exists())
        self.assertFalse(swapped_csv.exists())
        self.assertEqual(
            {"input.csv", "input.html"},
            {path.name for path in self.directory.iterdir()},
        )

    def test_wrong_output_suffixes_are_rejected_before_writes(self) -> None:
        cases = (
            (self.directory / "wrong.xlsm", self.output_csv),
            (self.output_xlsx, self.directory / "wrong.txt"),
        )
        for wrong_xlsx, wrong_csv in cases:
            with self.subTest(
                output_xlsx=wrong_xlsx.name,
                output_csv=wrong_csv.name,
            ):
                with self.assertRaises(ValueError):
                    write_outputs(
                        [self.row()],
                        headers=OUTPUT_HEADERS,
                        input_paths=(self.input_csv, self.input_html),
                        output_xlsx=wrong_xlsx,
                        output_csv=wrong_csv,
                        overwrite=False,
                    )
                self.assertFalse(wrong_xlsx.exists())
                self.assertFalse(wrong_csv.exists())
                self.assertEqual(
                    {"input.csv", "input.html"},
                    {path.name for path in self.directory.iterdir()},
                )

    def test_output_suffix_roles_are_case_insensitive(self) -> None:
        uppercase_xlsx = self.directory / "output.XLSX"
        uppercase_csv = self.directory / "output.CSV"

        write_outputs(
            [self.row()],
            headers=OUTPUT_HEADERS,
            input_paths=(self.input_csv, self.input_html),
            output_xlsx=uppercase_xlsx,
            output_csv=uppercase_csv,
            overwrite=False,
        )

        self.assertTrue(uppercase_xlsx.is_file())
        self.assertTrue(uppercase_csv.is_file())

    def test_csv_staging_failure_preserves_existing_outputs_and_cleans_stages(
        self,
    ) -> None:
        self.output_xlsx.write_bytes(b"old xlsx")
        self.output_csv.write_bytes(b"old csv")

        with patch.object(
            reconstruct_reddit_comments,
            "_write_csv",
            side_effect=RuntimeError("simulated CSV failure"),
        ):
            with self.assertRaisesRegex(RuntimeError, "simulated CSV failure"):
                self.write([self.row()], overwrite=True)

        self.assertEqual(b"old xlsx", self.output_xlsx.read_bytes())
        self.assertEqual(b"old csv", self.output_csv.read_bytes())
        self.assertEqual(
            {"input.csv", "input.html", "output.xlsx", "output.csv"},
            {path.name for path in self.directory.iterdir()},
        )

    def test_success_leaves_no_extra_files(self) -> None:
        self.write([self.row()])

        self.assertEqual(
            {"input.csv", "input.html", "output.xlsx", "output.csv"},
            {path.name for path in self.directory.iterdir()},
        )

    def test_preheld_output_reservation_blocks_writer_without_changes(self) -> None:
        reservation = self.output_xlsx.with_name(
            f".{self.output_xlsx.name}.reddit-output.lock"
        )
        reservation.write_bytes(b"held by another writer")
        try:
            with self.assertRaises(OutputPathConflictError):
                self.write([self.row()])

            self.assertFalse(self.output_xlsx.exists())
            self.assertFalse(self.output_csv.exists())
            self.assertEqual(b"held by another writer", reservation.read_bytes())
            self.assertEqual(
                {"input.csv", "input.html", reservation.name},
                {path.name for path in self.directory.iterdir()},
            )
        finally:
            reservation.unlink(missing_ok=True)

    def test_final_replace_failure_rolls_back_both_outputs_without_residue(
        self,
    ) -> None:
        real_replace = os.replace
        targets = {self.output_xlsx.resolve(), self.output_csv.resolve()}
        for preexisting in (False, True):
            for failure_position in (1, 2):
                with self.subTest(
                    preexisting=preexisting,
                    failure_position=failure_position,
                ):
                    self.output_xlsx.unlink(missing_ok=True)
                    self.output_csv.unlink(missing_ok=True)
                    old_xlsx = b"old xlsx bytes"
                    old_csv = b"old csv bytes"
                    if preexisting:
                        self.output_xlsx.write_bytes(old_xlsx)
                        self.output_csv.write_bytes(old_csv)
                    final_replace_count = 0
                    failure_injected = False

                    def fail_selected_final_replace(
                        source: str | bytes | os.PathLike[str] | os.PathLike[bytes],
                        destination: str
                        | bytes
                        | os.PathLike[str]
                        | os.PathLike[bytes],
                    ) -> None:
                        nonlocal final_replace_count, failure_injected
                        if (
                            not failure_injected
                            and Path(destination).resolve() in targets
                        ):
                            final_replace_count += 1
                            if final_replace_count == failure_position:
                                failure_injected = True
                                raise OSError(
                                    f"simulated final replace {failure_position}"
                                )
                        real_replace(source, destination)

                    with patch.object(
                        os,
                        "replace",
                        side_effect=fail_selected_final_replace,
                    ):
                        with self.assertRaisesRegex(
                            OSError,
                            f"simulated final replace {failure_position}",
                        ):
                            self.write([self.row()], overwrite=preexisting)

                    if preexisting:
                        self.assertEqual(old_xlsx, self.output_xlsx.read_bytes())
                        self.assertEqual(old_csv, self.output_csv.read_bytes())
                    else:
                        self.assertFalse(self.output_xlsx.exists())
                        self.assertFalse(self.output_csv.exists())
                    expected_names = {"input.csv", "input.html"}
                    if preexisting:
                        expected_names.update({"output.xlsx", "output.csv"})
                    self.assertEqual(
                        expected_names,
                        {path.name for path in self.directory.iterdir()},
                    )

    def test_backup_preparation_failure_keeps_existing_outputs_unchanged(
        self,
    ) -> None:
        old_xlsx = b"old xlsx bytes"
        old_csv = b"old csv bytes"
        self.output_xlsx.write_bytes(old_xlsx)
        self.output_csv.write_bytes(old_csv)
        real_copyfile = reconstruct_reddit_comments.shutil.copyfile
        for failure_position in (1, 2):
            with self.subTest(failure_position=failure_position):
                self.output_xlsx.write_bytes(old_xlsx)
                self.output_csv.write_bytes(old_csv)
                copy_count = 0

                def fail_selected_backup(
                    source: str | os.PathLike[str],
                    destination: str | os.PathLike[str],
                ) -> str:
                    nonlocal copy_count
                    copy_count += 1
                    if copy_count == failure_position:
                        raise OSError(
                            f"simulated backup failure {failure_position}"
                        )
                    return real_copyfile(source, destination)

                with patch.object(
                    reconstruct_reddit_comments.shutil,
                    "copyfile",
                    side_effect=fail_selected_backup,
                ):
                    with self.assertRaisesRegex(
                        OSError,
                        f"simulated backup failure {failure_position}",
                    ):
                        self.write([self.row()], overwrite=True)

                self.assertEqual(old_xlsx, self.output_xlsx.read_bytes())
                self.assertEqual(old_csv, self.output_csv.read_bytes())
                self.assertEqual(
                    {"input.csv", "input.html", "output.xlsx", "output.csv"},
                    {path.name for path in self.directory.iterdir()},
                )

    def test_restore_failure_retains_backup_and_reports_commit_and_rollback(
        self,
    ) -> None:
        old_xlsx = b"old xlsx bytes"
        old_csv = b"old csv bytes"
        self.output_xlsx.write_bytes(old_xlsx)
        self.output_csv.write_bytes(old_csv)
        real_replace = os.replace
        backup_path: Path | None = None

        def fail_commit_then_restore(
            source: str | bytes | os.PathLike[str] | os.PathLike[bytes],
            destination: str | bytes | os.PathLike[str] | os.PathLike[bytes],
        ) -> None:
            nonlocal backup_path
            source_path = Path(source)
            destination_path = Path(destination).resolve()
            if (
                destination_path == self.output_csv.resolve()
                and "reddit-stage" in source_path.name
            ):
                raise OSError("simulated second commit failure")
            if (
                destination_path == self.output_xlsx.resolve()
                and "reddit-backup" in source_path.name
            ):
                backup_path = source_path
                raise OSError("simulated restore failure")
            real_replace(source, destination)

        with patch.object(os, "replace", side_effect=fail_commit_then_restore):
            with self.assertRaises(RuntimeError) as caught:
                self.write(
                    [self.row(**{"Comment": "private row content"})],
                    overwrite=True,
                )

        self.assertIsNotNone(backup_path)
        assert backup_path is not None
        self.assertTrue(backup_path.is_file())
        self.assertEqual(old_xlsx, backup_path.read_bytes())
        self.assertNotEqual(old_xlsx, self.output_xlsx.read_bytes())
        self.assertEqual(old_csv, self.output_csv.read_bytes())
        message = str(caught.exception)
        self.assertIn("simulated second commit failure", message)
        self.assertIn("simulated restore failure", message)
        self.assertIn(str(backup_path), message)
        self.assertNotIn("private row content", message)
        self.assertEqual(
            {
                "input.csv",
                "input.html",
                "output.xlsx",
                "output.csv",
                backup_path.name,
            },
            {path.name for path in self.directory.iterdir()},
        )
        backup_path.unlink()

    def test_new_target_unlink_failure_reports_surviving_target(self) -> None:
        real_replace = os.replace
        real_unlink = Path.unlink

        def fail_second_commit(
            source: str | bytes | os.PathLike[str] | os.PathLike[bytes],
            destination: str | bytes | os.PathLike[str] | os.PathLike[bytes],
        ) -> None:
            if (
                Path(destination).resolve() == self.output_csv.resolve()
                and "reddit-stage" in Path(source).name
            ):
                raise OSError("simulated second commit failure")
            real_replace(source, destination)

        def fail_new_target_unlink(
            path: Path,
            missing_ok: bool = False,
        ) -> None:
            if path.resolve() == self.output_xlsx.resolve():
                raise OSError("simulated unlink rollback failure")
            real_unlink(path, missing_ok=missing_ok)

        with patch.object(os, "replace", side_effect=fail_second_commit):
            with patch.object(Path, "unlink", new=fail_new_target_unlink):
                with self.assertRaises(RuntimeError) as caught:
                    self.write([self.row()])

        self.assertTrue(self.output_xlsx.is_file())
        self.assertFalse(self.output_csv.exists())
        message = str(caught.exception)
        self.assertIn("simulated second commit failure", message)
        self.assertIn("simulated unlink rollback failure", message)
        self.assertIn(str(self.output_xlsx.resolve()), message)
        self.assertEqual(
            {"input.csv", "input.html", "output.xlsx"},
            {path.name for path in self.directory.iterdir()},
        )

    def test_all_committed_targets_are_rollback_attempted_after_failures(
        self,
    ) -> None:
        if not hasattr(reconstruct_reddit_comments, "_rollback_committed_outputs"):
            self.fail("missing committed-output rollback helper")
        rollback = reconstruct_reddit_comments._rollback_committed_outputs
        first_target = self.output_xlsx
        second_target = self.output_csv
        first_backup = self.directory / ".first.reddit-backup.xlsx"
        first_target.write_bytes(b"new first")
        first_backup.write_bytes(b"old first")
        second_target.write_bytes(b"new second")
        real_unlink = Path.unlink
        attempted: list[Path] = []

        def fail_restore(
            source: str | bytes | os.PathLike[str] | os.PathLike[bytes],
            destination: str | bytes | os.PathLike[str] | os.PathLike[bytes],
        ) -> None:
            attempted.append(Path(destination).resolve())
            raise OSError("simulated restore failure")

        def fail_unlink(path: Path, missing_ok: bool = False) -> None:
            if path.resolve() == second_target.resolve():
                attempted.append(path.resolve())
                raise OSError("simulated unlink failure")
            real_unlink(path, missing_ok=missing_ok)

        with patch.object(os, "replace", side_effect=fail_restore):
            with patch.object(Path, "unlink", new=fail_unlink):
                failures, retained_backups = rollback(
                    [second_target, first_target],
                    {first_target: first_backup},
                )

        self.assertEqual(
            [first_target.resolve(), second_target.resolve()],
            attempted,
        )
        self.assertEqual(2, len(failures))
        self.assertEqual({first_backup}, retained_backups)
        self.assertTrue(first_backup.is_file())
        self.assertTrue(second_target.is_file())

    def test_partial_backup_copy_is_removed_without_changing_outputs(self) -> None:
        old_xlsx = b"old xlsx bytes"
        old_csv = b"old csv bytes"
        self.output_xlsx.write_bytes(old_xlsx)
        self.output_csv.write_bytes(old_csv)

        def write_partial_backup_then_fail(
            source: str | os.PathLike[str],
            destination: str | os.PathLike[str],
        ) -> str:
            Path(destination).write_bytes(b"partial backup bytes")
            raise OSError("simulated partial backup failure")

        with patch.object(
            reconstruct_reddit_comments.shutil,
            "copyfile",
            side_effect=write_partial_backup_then_fail,
        ):
            with self.assertRaisesRegex(
                OSError,
                "simulated partial backup failure",
            ):
                self.write([self.row()], overwrite=True)

        self.assertEqual(old_xlsx, self.output_xlsx.read_bytes())
        self.assertEqual(old_csv, self.output_csv.read_bytes())
        self.assertEqual(
            {"input.csv", "input.html", "output.xlsx", "output.csv"},
            {path.name for path in self.directory.iterdir()},
        )


class RedditJsonPageTextCliTests(unittest.TestCase):
    SCRIPT = Path(__file__).resolve().parents[1] / "tools" / "reconstruct_reddit_comments.py"

    def setUp(self) -> None:
        self.temporary_directory = tempfile.TemporaryDirectory()
        self.addCleanup(self.temporary_directory.cleanup)
        self.directory = Path(self.temporary_directory.name)
        self.json_path = self.directory / "export.json"
        self.page_text_path = self.directory / "page.txt"
        self.output_xlsx = self.directory / "result.xlsx"
        self.output_csv = self.directory / "result.csv"

    def write_json(self, *, content: str = "SECRET-COMMENT-1") -> None:
        payload = {
            "meta": {"completeness": "complete", "collectedCommentCount": 2,
                     "reportedByApi": 3, "discrepancy": 1, "failedMore": 0,
                     "failedNodes": [], "failedReasons": [], "failedDetails": []},
            "post": {"id": "p1", "subreddit": "python", "title": "SECRET-TITLE",
                     "content": "SECRET-BODY", "author": "SECRET-AUTHOR", "num_comments": 3},
            "comments": [
                {"id": "c1", "parent_id": "p1", "content": content, "depth": 0,
                 "username": "alpha", "date": "exact-time-1", "created_utc": 1},
                {"id": "c2", "parent_id": "c1", "content": "SECRET-COMMENT-2", "depth": 1,
                 "username": "beta", "date": "exact-time-2", "created_utc": 2},
            ],
        }
        self.json_path.write_text(json.dumps(payload), encoding="utf-8")

    def write_collapsed_automoderator_json(self) -> None:
        payload = {
            "meta": {
                "completeness": "complete",
                "collectedCommentCount": 3,
                "reportedByApi": 3,
                "discrepancy": 0,
                "failedMore": 0,
                "failedNodes": [],
                "failedReasons": [],
                "failedDetails": [],
            },
            "post": {
                "id": "postidk3m7",
                "subreddit": "python",
                "title": "FixtureTitleQx7",
                "content": "FixtureBodyRz8",
                "author": "PostAuthorF8",
                "num_comments": 3,
            },
            "comments": [
                {
                    "id": "autoidp8v2",
                    "parent_id": "postidk3m7",
                    "content": "AutoBodyH2j4",
                    "depth": 0,
                    "username": "AutoModerator",
                    "date": "exact-time-0",
                    "created_utc": 0,
                },
                {
                    "id": "rootidq4w6",
                    "parent_id": "postidk3m7",
                    "content": "RootBodyK5n9",
                    "depth": 0,
                    "username": "RootAuthorD6",
                    "date": "exact-time-1",
                    "created_utc": 1,
                },
                {
                    "id": "replyidr9x1",
                    "parent_id": "rootidq4w6",
                    "content": "ReplyBodyL6p3",
                    "depth": 1,
                    "username": "ReplyAuthorE7",
                    "date": "exact-time-2",
                    "created_utc": 2,
                },
            ],
        }
        self.json_path.write_text(json.dumps(payload), encoding="utf-8")

    def write_collapsed_automoderator_page(self) -> None:
        first = "\n".join(
            (
                "RootAuthorD6",
                "\u20228\u5c0f\u65f6\u524d",
                "RootBodyK5n9",
                "",
                "\u8d5e\u540c",
                "1",
                "\u53cd\u5bf9",
                "\u56de\u590d",
                "\u5956\u52b1",
                "\u5206\u4eab",
            )
        )
        second = "\n".join(
            (
                "ReplyAuthorE7",
                "\u20228\u5c0f\u65f6\u524d",
                "ReplyBodyL6p3",
                "",
                "\u8d5e\u540c\u6295\u7968",
                "\u53cd\u5bf9",
                "\u56de\u590d",
                "\u5956\u52b1",
                "\u5206\u4eab",
            )
        )
        banner = "\n".join(
            (
                "AutoModerator",
                "\u8fd9\u662f\u81ea\u52a8\u5316\u8d26\u6237\u3002",
                "\u7248\u4e3b",
                "4\u5929\u524d",
            )
        )
        self.page_text_path.write_text(
            "\n".join(
                (
                    "Reddit",
                    "r/python",
                    "u/PostAuthorF8",
                    "PostAuthorF8 \u5934\u50cf",
                    "8\u5c0f\u65f6\u524d",
                    "FixtureTitleQx7",
                    "\u6b63\u6587",
                    "\u8d5e\u540c",
                    "99",
                    "\u53cd\u5bf9",
                    "3",
                    "\u8f6c\u5230\u8bc4\u8bba",
                    "\u8bc4\u8bba\u533a\u57df",
                    banner,
                    first,
                    second,
                )
            ),
            encoding="utf-8",
        )

    def write_page(
        self,
        *,
        author: str = "alpha",
        content: str = "SECRET-COMMENT-1",
        score: str = "99",
        count: str = "3",
        reverse: bool = False,
    ) -> None:
        first = "\n".join((author, "\u20228\u5c0f\u65f6\u524d", content, "", "\u8d5e\u540c", "3", "\u53cd\u5bf9", "\u56de\u590d", "\u5956\u52b1", "\u5206\u4eab"))
        second = "\n".join(("beta", "\u20228\u5c0f\u65f6\u524d", "SECRET-COMMENT-2", "", "\u8d5e\u540c\u6295\u7968", "\u53cd\u5bf9", "\u56de\u590d", "\u5956\u52b1", "\u5206\u4eab"))
        comments = (second, first) if reverse else (first, second)
        self.page_text_path.write_text("\n".join((
            "Reddit", "r/python", "u/SECRET-AUTHOR", "SECRET-AUTHOR \u5934\u50cf",
            "8\u5c0f\u65f6\u524d", "SECRET-TITLE", "\u6b63\u6587", "\u8d5e\u540c", score, "\u53cd\u5bf9", count,
            "\u8f6c\u5230\u8bc4\u8bba", "\u8bc4\u8bba\u533a\u57df", *comments,
        )), encoding="utf-8")

    def run_cli(self, *args: str) -> subprocess.CompletedProcess[str]:
        return subprocess.run(
            [sys.executable, "tools/reconstruct_reddit_comments.py", "--json", str(self.json_path),
             "--page-text", str(self.page_text_path), "--output-xlsx", str(self.output_xlsx),
             "--output-csv", str(self.output_csv), *args],
            cwd=Path.cwd(), text=True, capture_output=True, check=False,
        )

    def assert_safe_failure(
        self,
        completed: subprocess.CompletedProcess[str],
        *page_only_secrets: str,
    ) -> None:
        self.assertNotEqual(0, completed.returncode)
        self.assertNotIn("Traceback", completed.stderr)
        for secret in (
            "SECRET-TITLE",
            "SECRET-BODY",
            "SECRET-AUTHOR",
            "SECRET-COMMENT",
            "c1",
            "c2",
            *page_only_secrets,
        ):
            self.assertNotIn(secret, completed.stdout + completed.stderr)
        self.assertFalse(self.output_xlsx.exists())
        self.assertFalse(self.output_csv.exists())
        self.assertEqual(
            [],
            [
                path.name
                for path in self.directory.iterdir()
                if path.suffix.lower() in (".xlsx", ".csv")
            ],
        )
        self.assertEqual([], [
            path.name for path in self.directory.iterdir()
            if any(marker in path.name for marker in ("reddit-stage", "reddit-backup", ".lock"))
        ])

    def test_json_page_text_happy_path_writes_safe_summary(self) -> None:
        self.write_json()
        self.write_page()

        completed = self.run_cli()

        self.assertEqual(0, completed.returncode, completed.stderr)
        self.assertEqual(
            [
                f"Reddit JSON input: {self.json_path.resolve()}",
                f"Reddit page text input: {self.page_text_path.resolve()}",
                f"XLSX output: {self.output_xlsx.resolve()}",
                f"CSV output: {self.output_csv.resolve()}",
                "JSON comment count: 2",
                "Page comment match count: 2",
                "Missing comment score count: 1",
                "Unavailable reported comment gap: 1",
            ],
            completed.stdout.splitlines(),
        )
        for secret in (
            "SECRET-TITLE",
            "SECRET-BODY",
            "SECRET-AUTHOR",
            "SECRET-COMMENT",
            "c1",
            "c2",
        ):
            self.assertNotIn(secret, completed.stdout + completed.stderr)
        sheet = load_workbook(self.output_xlsx, data_only=False).active
        self.assertEqual(list(JSON_TEXT_OUTPUT_HEADERS), list(next(sheet.values)))
        self.assertEqual(99, sheet.cell(2, 5).value)
        self.assertEqual(3, sheet.cell(2, 9).value)
        self.assertEqual("c2", sheet.cell(3, 13).value)
        self.assertIsNone(sheet.cell(3, 9).value)

    def test_collapsed_automoderator_cli_outputs_only_verified_comments(
        self,
    ) -> None:
        self.write_collapsed_automoderator_json()
        self.write_collapsed_automoderator_page()

        completed = self.run_cli()

        self.assertEqual(0, completed.returncode, completed.stderr)
        self.assertIn("JSON comment count: 3", completed.stdout)
        self.assertIn("Page comment match count: 2", completed.stdout)
        sheet = load_workbook(self.output_xlsx, data_only=False).active
        self.assertEqual(3, sheet.max_row)
        self.assertEqual(2, sheet.cell(2, 6).value)
        self.assertEqual(2, sheet.cell(3, 6).value)
        self.assertEqual(
            ["rootidq4w6", "replyidr9x1"],
            [sheet.cell(row, 13).value for row in range(2, 4)],
        )
        protected_tokens = (
            "FixtureTitleQx7",
            "FixtureBodyRz8",
            "PostAuthorF8",
            "postidk3m7",
            "AutoBodyH2j4",
            "AutoModerator",
            "autoidp8v2",
            "RootBodyK5n9",
            "RootAuthorD6",
            "rootidq4w6",
            "ReplyBodyL6p3",
            "ReplyAuthorE7",
            "replyidr9x1",
        )
        self.assertEqual(
            len(protected_tokens),
            len(set(protected_tokens)),
        )
        for token in protected_tokens:
            self.assertFalse(
                any(
                    token != other and token in other
                    for other in protected_tokens
                ),
                f"protected token is contained in another token: {token}",
            )
            self.assertNotIn(token, completed.stdout + completed.stderr)

    def test_overwrite_rejection_preserves_pair_and_overwrite_succeeds(self) -> None:
        self.write_json()
        self.write_page()
        self.assertEqual(0, self.run_cli().returncode)
        old_xlsx, old_csv = self.output_xlsx.read_bytes(), self.output_csv.read_bytes()

        rejected = self.run_cli()

        self.assertNotEqual(0, rejected.returncode)
        self.assertNotIn("Traceback", rejected.stderr)
        self.assertEqual(old_xlsx, self.output_xlsx.read_bytes())
        self.assertEqual(old_csv, self.output_csv.read_bytes())
        self.assertEqual(0, self.run_cli("--overwrite").returncode)

    def test_invalid_json_page_mismatches_and_control_character_fail_safely(self) -> None:
        scenarios = (
            ("invalid-json", lambda: self.json_path.write_text('{"title":"SECRET-TITLE"', encoding="utf-8"), ()),
            ("author", lambda: (self.write_json(), self.write_page(author="PAGE-ONLY-SECRET-AUTHOR")), ("PAGE-ONLY-SECRET-AUTHOR",)),
            ("content", lambda: (self.write_json(), self.write_page(content="PAGE-ONLY-SECRET-CONTENT")), ("PAGE-ONLY-SECRET-CONTENT",)),
            ("order", lambda: (self.write_json(), self.write_page(reverse=True)), ()),
            ("count", lambda: (self.write_json(), self.write_page(count="2")), ()),
            ("control", lambda: (self.write_json(content="SECRET-COMMENT-1\u0001"), self.write_page(content="SECRET-COMMENT-1\u0001")), ()),
        )
        for name, arrange, page_only_secrets in scenarios:
            with self.subTest(name=name):
                self.json_path = self.directory / f"{name}.json"
                self.page_text_path = self.directory / f"{name}.txt"
                self.output_xlsx = self.directory / f"{name}.xlsx"
                self.output_csv = self.directory / f"{name}.csv"
                arrange()
                completed = self.run_cli()
                self.assert_safe_failure(completed, *page_only_secrets)
                if name == "control":
                    self.assertIn("unsupported control character", completed.stderr)

    def test_out_of_range_page_integers_fail_cli_safely_without_mutation(
        self,
    ) -> None:
        old_xlsx = b"old xlsx"
        old_csv = b"old csv"
        cases = (
            "9007199254740993",
            "8" * 400,
        )
        for index, private_value in enumerate(cases):
            with self.subTest(digits=len(private_value)):
                self.json_path = self.directory / f"numeric-{index}.json"
                self.page_text_path = self.directory / f"numeric-{index}.txt"
                self.output_xlsx = self.directory / f"numeric-{index}.xlsx"
                self.output_csv = self.directory / f"numeric-{index}.csv"
                self.write_json()
                self.write_page(score=private_value)
                self.output_xlsx.write_bytes(old_xlsx)
                self.output_csv.write_bytes(old_csv)

                completed = self.run_cli("--overwrite")

                self.assertNotEqual(0, completed.returncode)
                self.assertNotIn("Traceback", completed.stderr)
                self.assertNotIn(private_value, completed.stdout + completed.stderr)
                self.assertIn(
                    "post score outside exact XLSX integer range",
                    completed.stderr,
                )
                self.assertEqual(old_xlsx, self.output_xlsx.read_bytes())
                self.assertEqual(old_csv, self.output_csv.read_bytes())
                self.assertEqual(
                    [],
                    [
                        path.name
                        for path in self.directory.iterdir()
                        if any(
                            marker in path.name
                            for marker in (
                                "reddit-stage",
                                "reddit-backup",
                                ".lock",
                            )
                        )
                    ],
                )

    def test_help_has_only_json_page_text_workflow_options(self) -> None:
        completed = subprocess.run(
            [sys.executable, "tools/reconstruct_reddit_comments.py", "--help"],
            cwd=Path.cwd(), text=True, capture_output=True, check=False,
        )

        self.assertEqual(0, completed.returncode)
        for option in ("--json", "--page-text", "--output-xlsx", "--output-csv", "--overwrite"):
            self.assertIn(option, completed.stdout)
        for option in ("--free-csv", "--html", "--post-author", "--post-score", "--post-comment-count"):
            self.assertNotIn(option, completed.stdout)

    def test_formal_cli_import_does_not_require_legacy_parsers(self) -> None:
        completed = subprocess.run(
            [
                sys.executable,
                "-c",
                (
                    "import sys; "
                    "sys.modules['tools.reddit_free_csv'] = None; "
                    "sys.modules['tools.reddit_saved_html'] = None; "
                    "import tools.reconstruct_reddit_comments"
                ),
            ],
            cwd=Path.cwd(),
            text=True,
            capture_output=True,
            check=False,
        )

        self.assertEqual(0, completed.returncode, completed.stderr)

    def test_path_resolution_error_is_concise(self) -> None:
        stderr = io.StringIO()
        with patch.object(Path, "resolve", side_effect=OSError("simulated path resolution failure")):
            with patch.object(sys, "stderr", stderr):
                result = reconstruct_reddit_comments.main([
                    "--json", "export.json", "--page-text", "page.txt",
                    "--output-xlsx", "result.xlsx", "--output-csv", "result.csv",
                ])

        self.assertEqual(1, result)
        self.assertEqual("Error: simulated path resolution failure\n", stderr.getvalue())
        self.assertNotIn("Traceback", stderr.getvalue())


if __name__ == "__main__":
    unittest.main()
