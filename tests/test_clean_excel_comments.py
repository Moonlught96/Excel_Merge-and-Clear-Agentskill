from __future__ import annotations

import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from tools.clean_excel_comments import (
    DEFAULT_CONFIG_PATH,
    CleanerConfig,
    ExternalCleanerConfigError,
    FinalizedCleaningAuditArtifactsError,
    clean_workbook,
    load_config,
    main,
    parse_args,
    should_delete_comment,
)


class CleanExcelCommentsTest(unittest.TestCase):
    def test_cli_requires_explicit_standard_comment_header_and_platform(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-cleaner-explicit-comment-column"
        tmp.mkdir(parents=True, exist_ok=True)

        with self.assertRaises(SystemExit):
            parse_args(
                [
                    str(tmp / "source.xlsx"),
                    "--output",
                    str(tmp / "cleaned.xlsx"),
                    "--platform",
                    "B站",
                ]
            )

        with self.assertRaises(SystemExit):
            parse_args(
                [
                    str(tmp / "source.xlsx"),
                    "--output",
                    str(tmp / "cleaned.xlsx"),
                    "--target-header",
                    "产品名",
                    "--platform",
                    "B站",
                ]
            )

        args = parse_args(
            [
                str(tmp / "source.xlsx"),
                "--output",
                str(tmp / "cleaned.xlsx"),
                "--target-header",
                "评论内容",
                "--platform",
                "B站",
            ]
        )
        self.assertEqual("评论内容", args.target_header)
        self.assertEqual("B站", args.platform)

    def test_twitter_uses_base_url_and_spam_rules(self) -> None:
        twitter_config = load_config(DEFAULT_CONFIG_PATH, platform="Twitter")

        self.assertIsNone(twitter_config.platform_profile)
        self.assertEqual(
            "评论包含固定删除词: https://",
            should_delete_comment(
                "This is a normal ScreenBar review https://t.co/example",
                set(),
                twitter_config,
                (),
            ),
        )
        self.assertEqual(
            "评论包含固定删除词: link in bio",
            should_delete_comment(
                "This is a normal ScreenBar review, link in bio for a coupon",
                set(),
                twitter_config,
                (),
            ),
        )

    def test_cli_rejects_external_or_temporary_cleaner_config(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-cleaner-external-config"
        tmp.mkdir(parents=True, exist_ok=True)
        external_config = tmp / "comment-cleaner-temporary.json"
        external_config.write_text(
            DEFAULT_CONFIG_PATH.read_text(encoding="utf-8"),
            encoding="utf-8",
        )

        with self.assertRaisesRegex(ExternalCleanerConfigError, "External or temporary"):
            main(
                [
                    str(tmp / "source.xlsx"),
                    "--output",
                    str(tmp / "cleaned.xlsx"),
                    "--target-header",
                    "评论内容",
                    "--platform",
                    "B站",
                    "--config",
                    str(external_config),
                ]
            )

    def test_refuses_to_regenerate_a_deleted_cleaning_log_at_finalized_path(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-cleaner-finalized-log-restoration"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "cleaned.xlsx"
        deletion_log_path = output_path.with_suffix(".deletions.csv")

        workbook = Workbook()
        workbook.active.append(["评论内容"])
        workbook.active.append(["这是需要保留的完整评论内容"])
        workbook.save(input_path)
        workbook.close()
        output_path.write_text("finalized-cleaned-output", encoding="utf-8")
        deletion_log_path.unlink(missing_ok=True)

        with self.assertRaisesRegex(FinalizedCleaningAuditArtifactsError, "Refusing to regenerate"):
            clean_workbook(
                input_path,
                CleanerConfig(target_header="评论内容"),
                (),
                output_path=output_path,
                overwrite=True,
                overwrite_confirmations=(output_path,),
            )

        self.assertEqual("finalized-cleaned-output", output_path.read_text(encoding="utf-8"))
        self.assertFalse(deletion_log_path.exists())

    def test_closes_input_workbook_when_cleaning_fails(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-clean-closes-on-error"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "cleaned.xlsx"
        source = Workbook()
        source.active.append(["评论内容"])
        source.active.append(["评论内容足够完整"])
        source.save(input_path)
        source.close()

        opened = load_workbook(input_path, read_only=False, data_only=False)
        opened.close = mock.Mock(wraps=opened.close)
        with mock.patch(
            "tools.clean_excel_comments.load_workbook_for_processing",
            return_value=opened,
        ), mock.patch(
            "tools.clean_excel_comments.clean_sheet",
            side_effect=RuntimeError("processing failed"),
        ):
            with self.assertRaisesRegex(RuntimeError, "processing failed"):
                clean_workbook(
                    input_path,
                    CleanerConfig(target_header="评论内容"),
                    (),
                    output_path=output_path,
                )

        opened.close.assert_called_once_with()

    def test_case_insensitive_latin_fixed_word_requires_word_boundaries(self) -> None:
        config = CleanerConfig(
            delete_contains_texts=(),
            delete_contains_case_insensitive_texts=("test",),
        )

        self.assertIsNone(
            should_delete_comment(
                "This monitor light uses TESTV for custom firmware",
                set(),
                config,
                (),
            )
        )
        self.assertEqual(
            "评论包含固定删除词: test",
            should_delete_comment(
                "This monitor light is only a TEST comment",
                set(),
                config,
                (),
            ),
        )

    def test_chinese_comments_do_not_use_latin_fixed_delete_words(self) -> None:
        config = CleanerConfig(
            delete_contains_texts=(),
            delete_contains_case_insensitive_texts=("test", "加v"),
        )

        self.assertIsNone(
            should_delete_comment(
                "这条中文评论提到 test 但仍然是正常产品说明",
                set(),
                config,
                (),
            )
        )
        self.assertEqual(
            "评论包含固定删除词: 加v",
            should_delete_comment(
                "这条中文广告要求用户加V领取优惠资料",
                set(),
                config,
                (),
            ),
        )

    def test_canonical_chinese_fixed_terms_do_not_delete_normal_comments(self) -> None:
        config = load_config(DEFAULT_CONFIG_PATH)
        removed_ambiguous_terms = ("无", "第一", "略", "测试", "来了", "路过")
        normal_comments = (
            "无炫光设计让我很满意",
            "第一次购买屏幕挂灯效果很好",
            "价格略贵但整体品质不错",
            "感应到人来了就会自动亮灯",
            "测试使用两周后光线非常舒服",
            "路过时人体感应会自动亮灯",
        )

        for term in removed_ambiguous_terms:
            self.assertNotIn(term, config.delete_contains_texts)
            self.assertNotIn(term, CleanerConfig().delete_contains_texts)

        for comment in normal_comments:
            self.assertIsNone(should_delete_comment(comment, set(), config, ()))

    def test_contextual_latin_words_do_not_delete_normal_reviews(self) -> None:
        config = load_config(DEFAULT_CONFIG_PATH)
        contextual_terms = ("first", "test", "how much", "price?")
        normal_reviews = (
            "This is my first ScreenBar and it works great every evening.",
            "This test report explains why the lamp works well for my desk.",
            "How much does this ScreenBar cost in Japan after shipping today?",
            "Price? The ScreenBar has excellent build quality and useful light.",
        )

        for term in contextual_terms:
            self.assertNotIn(term, config.delete_contains_texts)
            self.assertNotIn(term, config.delete_contains_case_insensitive_texts)
            self.assertNotIn(term, CleanerConfig().delete_contains_texts)
            self.assertNotIn(term, CleanerConfig().delete_contains_case_insensitive_texts)

        for comment in normal_reviews:
            self.assertIsNone(should_delete_comment(comment, set(), config, ()))

    def test_fixed_delete_words_are_isolated_by_script_group(self) -> None:
        config = CleanerConfig(
            delete_contains_texts=("https://",),
            delete_contains_case_insensitive_texts=("test", "テスト", "테스트"),
        )

        self.assertIsNone(
            should_delete_comment(
                "これは test を含む通常の製品レビューです",
                set(),
                config,
                (),
            )
        )
        self.assertEqual(
            "评论包含固定删除词: テスト",
            should_delete_comment(
                "これはテスト投稿なので削除してください",
                set(),
                config,
                (),
            ),
        )
        self.assertIsNone(
            should_delete_comment(
                "이 한국어 제품 설명에는 test 문자가 있습니다",
                set(),
                config,
                (),
            )
        )
        self.assertEqual(
            "评论包含固定删除词: https://",
            should_delete_comment(
                "这条中文评论包含网址 https://example.com",
                set(),
                config,
                (),
            ),
        )

    def test_japanese_with_kanji_does_not_use_chinese_length_threshold(self) -> None:
        config = CleanerConfig(
            delete_contains_texts=(),
            delete_contains_case_insensitive_texts=(),
        )

        self.assertIsNone(should_delete_comment("照明最高です", set(), config, ()))
        self.assertEqual(
            "非中文无空格短文本长度小于等于 4",
            should_delete_comment("最高です", set(), config, ()),
        )

    def test_korean_with_hanja_does_not_use_chinese_length_threshold(self) -> None:
        config = CleanerConfig(
            delete_contains_texts=(),
            delete_contains_case_insensitive_texts=(),
        )

        self.assertIsNone(should_delete_comment("漢字좋아요", set(), config, ()))

    def test_clean_workbook_matches_rpa_rules_without_like_column(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-rules"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "dirty.xlsx"
        output_dir = tmp / "out"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["id", "name", "评论内容"])
        sheet.append([1, "keep", "正常评论内容很完整"])
        sheet.append([2, "short", "abcd"])
        sheet.append([3, "placeholder", "该用户未填写评价内容"])
        sheet.append([4, "clean-word", "这里包含KOL清理词1"])
        sheet.append([5, "duplicate-old", "重复评论内容很完整"])
        sheet.append([6, "duplicate-new", "重复评论内容很完整"])
        sheet.append([7, "link", "这里包含链接"])

        second = workbook.create_sheet("SheetB")
        second.append(["id", "name", "评论内容"])
        second.append([1, "same-text-different-sheet", "重复评论内容很完整"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(),
            clean_words=("KOL清理词1", "KOL清理词2"),
            output_dir=output_dir,
        )

        cleaned = load_workbook(result.output_xlsx)
        cleaned_sheet = cleaned["SheetA"]
        remaining_comments = [
            cleaned_sheet.cell(row=row_number, column=3).value
            for row_number in range(2, cleaned_sheet.max_row + 1)
        ]

        self.assertEqual(["正常评论内容很完整", "重复评论内容很完整"], remaining_comments)
        self.assertEqual(5, result.rows_deleted)
        self.assertTrue(result.output_csv and result.output_csv.exists())
        self.assertTrue(result.deletion_log_csv.exists())
        self.assertTrue(result.summary_json.exists())

    def test_clean_workbook_rejects_legacy_numeric_target_column_fallback(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-reject-legacy-target-column"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "legacy.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["id", "name", "comment"])
        sheet.append([1, "user", "This is a normal review with enough detail."])
        workbook.save(input_path)
        workbook.close()

        with self.assertRaisesRegex(ValueError, "numeric target_column fallback"):
            clean_workbook(
                input_path=input_path,
                config=CleanerConfig(target_column=3, target_header=None),
                clean_words=(),
                output_path=tmp / "cleaned.xlsx",
            )

    def test_duplicate_main_comments_keep_highest_like_count_then_last_tie(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-duplicate-main-comments-max-likes"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "duplicates.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论内容", "点赞数", "标记"])
        sheet.append(["同一条完整评论内容", 10, "highest-like"])
        sheet.append(["同一条完整评论内容", 2, "lower-like-later"])
        sheet.append(["另一条完整评论内容", 5, "tie-earlier"])
        sheet.append(["另一条完整评论内容", 5, "tie-later"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx)
        rows = list(cleaned["SheetA"].iter_rows(min_row=2, values_only=True))
        self.assertEqual(
            [
                ("同一条完整评论内容", 10, "highest-like"),
                ("另一条完整评论内容", 5, "tie-later"),
            ],
            rows,
        )
        self.assertEqual(2, result.rows_deleted)

    def test_clean_words_are_optional(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-optional-clean-words"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "dirty.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["id", "name", "评论内容"])
        sheet.append([1, "keep", "这里包含KOL清理词1但没有传入清理词"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx)
        self.assertEqual(2, cleaned["SheetA"].max_row)
        self.assertEqual(0, result.rows_deleted)

    def test_clean_workbook_accepts_confirmed_output_filename(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-confirmed-output-filename"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"
        output_path = tmp / "20260707_ScreenBar Halo2_淘宝评论数据_清洗后总表.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容"])
        sheet.append(["2026/07/07", "这是一条正常评论"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_path=output_path,
        )

        self.assertEqual(output_path.resolve(), result.output_xlsx)
        self.assertEqual(output_path.with_suffix(".csv").resolve(), result.output_csv)
        self.assertTrue(output_path.exists())

    def test_clean_workbook_rejects_output_path_that_would_overwrite_input_file(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-clean-output-conflict"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "dirty.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["id", "name", "comment"])
        sheet.append([1, "keep", "正常评论内容很完整"])
        workbook.save(input_path)

        with self.assertRaisesRegex(ValueError, "new workbook path"):
            clean_workbook(
                input_path=input_path,
                config=CleanerConfig(),
                clean_words=(),
                output_path=input_path,
            )

        original = load_workbook(input_path, read_only=True, data_only=True)
        self.assertEqual("正常评论内容很完整", original.active.cell(row=2, column=3).value)

    def test_clean_workbook_rejects_derived_csv_path_that_would_overwrite_csv_input(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-clean-csv-sidecar-conflict"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.csv"
        output_path = tmp / "source.xlsx"
        output_path.unlink(missing_ok=True)
        original_bytes = "id,name,comment\n1,keep,正常评论内容很完整\n".encode("utf-8-sig")
        input_path.write_bytes(original_bytes)

        with self.assertRaisesRegex(ValueError, "derived output path"):
            clean_workbook(
                input_path=input_path,
                config=CleanerConfig(),
                clean_words=(),
                output_path=output_path,
            )

        self.assertEqual(original_bytes, input_path.read_bytes())
        self.assertFalse(output_path.exists())

    def test_requires_explicit_overwrite_for_any_existing_clean_output(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-clean-existing-output"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"
        output_path = tmp / "cleaned.xlsx"
        output_csv = output_path.with_suffix(".csv")
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容"])
        sheet.append(["2026-07-22", "这是一条足够完整的产品评论"])
        workbook.save(input_path)
        output_csv.write_text("existing", encoding="utf-8")

        with self.assertRaisesRegex(ValueError, "already exists"):
            clean_workbook(
                input_path=input_path,
                config=CleanerConfig(target_header="评论内容"),
                clean_words=(),
                output_path=output_path,
                overwrite=False,
            )

        self.assertEqual("existing", output_csv.read_text(encoding="utf-8"))
        self.assertFalse(output_path.exists())

    def test_comments_with_seven_or_fewer_characters_are_deleted(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-length-limit"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "dirty.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["id", "name", "评论内容"])
        sheet.append([1, "delete", "1234567"])
        sheet.append([2, "keep", "12345678"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx)
        cleaned_sheet = cleaned["SheetA"]
        remaining_comments = [
            cleaned_sheet.cell(row=row_number, column=3).value
            for row_number in range(2, cleaned_sheet.max_row + 1)
        ]
        self.assertEqual(["12345678"], remaining_comments)
        self.assertEqual(1, result.rows_deleted)

    def test_clean_workbook_can_target_comment_by_header_after_standardization(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-target-header"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论时间", "评论内容", "点赞数"])
        sheet.append(["2026/01/05", "1234567", 10])
        sheet.append(["2026/01/06", "12345678", 9])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_column=99, target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))

        self.assertEqual(("评论时间", "评论内容", "点赞数"), rows[0])
        self.assertEqual([("2026/01/06", "12345678", 9)], rows[1:])
        self.assertEqual(1, result.rows_deleted)

    def test_clean_workbook_accepts_csv_input_through_compatibility_layer(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-clean-csv-input"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.csv"
        input_path.write_text(
            "评论日期,评论内容,点赞数\n"
            "2026/01/05,1234567,10\n"
            "2026/01/06,12345678,9\n",
            encoding="utf-8-sig",
        )

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_column=99, target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["standardized"].iter_rows(values_only=True))

        self.assertEqual(("评论日期", "评论内容", "点赞数"), rows[0])
        self.assertEqual([("2026/01/06", "12345678", "9")], rows[1:])
        self.assertEqual(1, result.rows_deleted)
        self.assertTrue(result.output_csv and result.output_csv.exists())

    def test_deletes_expanded_fixed_contains_words(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-expanded-fixed-contains"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数"])
        sheet.append(["2026/01/01", "这是正常评论内容足够长", 1])
        sheet.append(["2026/01/02", "为了金币随便写几个字", 1])
        sheet.append(["2026/01/03", "请加V领取内部券", 1])
        sheet.append(["2026/01/04", "点击链接 https://example.com", 1])
        sheet.append(["2026/01/05", "暂无评价，完成任务", 1])
        sheet.append(["2026/01/06", "这里包含原有固定词链接", 1])
        sheet.append(["2026/01/07", "蹲一个求分享", 1])
        sheet.append(["2026/01/08", "互赞交朋友", 1])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))

        self.assertEqual([("2026/01/01", "这是正常评论内容足够长", 1)], rows[1:])
        self.assertEqual(7, result.rows_deleted)

    def test_deletes_multilingual_fixed_spam_words(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-multilingual-fixed-contains"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数"])
        sheet.append(["2026/01/01", "This monitor light works great", 1])
        sheet.append(["2026/01/02", "LINK IN BIO for discount code", 1])
        sheet.append(["2026/01/03", "Please DM me on Telegram", 1])
        sheet.append(["2026/01/04", "内容なしです", 1])
        sheet.append(["2026/01/05", "링크 확인 부탁드립니다", 1])
        sheet.append(["2026/01/06", "FIRST!!!", 1])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))

        self.assertEqual([("2026/01/01", "This monitor light works great", 1)], rows[1:])
        self.assertEqual(5, result.rows_deleted)

    def test_deletes_random_alnum_without_chinese_but_keeps_normal_english_phrase(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-random-alnum-without-chinese"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数"])
        sheet.append(["2026/01/01", "good product works well today", 1])
        sheet.append(["2026/01/01", "good product works well", 1])
        sheet.append(["2026/01/02", "jdsklafjskl", 1])
        sheet.append(["2026/01/03", "123123123", 1])
        sheet.append(["2026/01/04", "abc123xyz9", 1])
        sheet.append(["2026/01/05", "中文夹着 jdsklafjskl 不按纯英文数字堆砌删除", 1])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))

        self.assertEqual(
            [
                ("2026/01/01", "good product works well today", 1),
                ("2026/01/05", "中文夹着 jdsklafjskl 不按纯英文数字堆砌删除", 1),
            ],
            rows[1:],
        )
        self.assertEqual(4, result.rows_deleted)

    def test_deletes_non_chinese_comments_with_four_or_fewer_words(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-non-chinese-short-word-limit"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["date", "评论内容", "likes"])
        sheet.append(["2026/01/01", "This works very well", 1])
        sheet.append(["2026/01/02", "This monitor light works very well", 1])
        sheet.append(["2026/01/03", "muy buen producto", 1])
        sheet.append(["2026/01/04", "esta lámpara funciona muy bien", 1])
        sheet.append(["2026/01/05", "\u0e42\u0e04\u0e21\u0e44\u0e1f\u0e15\u0e31\u0e49\u0e07\u0e42\u0e15\u0e4a\u0e30\u0e19\u0e35\u0e49\u0e14\u0e35\u0e21\u0e32\u0e01", 1])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))

        self.assertEqual(
            [
                ("2026/01/02", "This monitor light works very well", 1),
                ("2026/01/04", "esta lámpara funciona muy bien", 1),
                ("2026/01/05", "\u0e42\u0e04\u0e21\u0e44\u0e1f\u0e15\u0e31\u0e49\u0e07\u0e42\u0e15\u0e4a\u0e30\u0e19\u0e35\u0e49\u0e14\u0e35\u0e21\u0e32\u0e01", 1),
            ],
            rows[1:],
        )
        self.assertEqual(2, result.rows_deleted)

    def test_deletes_thai_hindi_and_spanish_fixed_spam_words(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-thai-hindi-spanish-fixed-contains"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["date", "评论内容", "likes"])
        sheet.append(["2026/01/01", "This monitor light works great on my desk", 1])
        sheet.append(["2026/01/02", "pásame el link por favor", 1])
        sheet.append(["2026/01/03", "\u0e02\u0e2d\u0e25\u0e34\u0e07\u0e01\u0e4c\u0e2a\u0e34\u0e19\u0e04\u0e49\u0e32\u0e2b\u0e19\u0e48\u0e2d\u0e22", 1])
        sheet.append(["2026/01/04", "\u0932\u093f\u0902\u0915 \u092d\u0947\u091c\u094b \u092a\u094d\u0932\u0940\u091c", 1])
        sheet.append(["2026/01/05", "\u0e42\u0e04\u0e21\u0e44\u0e1f\u0e19\u0e35\u0e49\u0e43\u0e0a\u0e49\u0e07\u0e32\u0e19\u0e14\u0e35\u0e21\u0e32\u0e01", 1])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))

        self.assertEqual(
            [
                ("2026/01/01", "This monitor light works great on my desk", 1),
                ("2026/01/05", "\u0e42\u0e04\u0e21\u0e44\u0e1f\u0e19\u0e35\u0e49\u0e43\u0e0a\u0e49\u0e07\u0e32\u0e19\u0e14\u0e35\u0e21\u0e32\u0e01", 1),
            ],
            rows[1:],
        )
        self.assertEqual(3, result.rows_deleted)

    def test_duplicate_subcomments_are_cleared_without_deleting_main_comment_rows(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-duplicate-subcomments"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数", "子评论数/追评数", "一级评论", "二级评论", "三级评论"])
        sheet.append(["2026/01/01", "主评论内容A足够长", 1, 0, "重复子评论甲", "", ""])
        sheet.append(["2026/01/02", "主评论内容B足够长", 1, 0, "重复子评论甲", "", ""])
        sheet.append(["2026/01/03", "主评论内容C足够长", 1, 0, "", "重复子评论乙", ""])
        sheet.append(["2026/01/04", "主评论内容D足够长", 1, 0, "", "重复子评论乙", ""])
        sheet.append(["2026/01/05", "主评论内容E足够长", 1, 0, "", "", "重复子评论丙"])
        sheet.append(["2026/01/06", "主评论内容F足够长", 1, 0, "", "", "重复子评论丙"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))
        main_comments = [row[1] for row in rows[1:]]
        subcomment_values = [row[4:7] for row in rows[1:]]

        self.assertEqual(
            [
                "主评论内容A足够长",
                "主评论内容B足够长",
                "主评论内容C足够长",
                "主评论内容D足够长",
                "主评论内容E足够长",
                "主评论内容F足够长",
            ],
            main_comments,
        )
        self.assertEqual(
            [
                (None, None, None),
                ("重复子评论甲", None, None),
                (None, None, None),
                (None, "重复子评论乙", None),
                (None, None, None),
                (None, None, "重复子评论丙"),
            ],
            subcomment_values,
        )
        self.assertEqual(0, result.rows_deleted)
        self.assertEqual(3, result.cells_cleared)

    def test_short_subcomments_are_cleared_without_deleting_main_comment_rows(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-short-subcomments"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数", "子评论数/追评数", "一级评论", "二级评论", "三级评论"])
        sheet.append(["2026/01/01", "主评论内容A足够长", 1, 0, " 12345 ", "123456", "短评"])
        sheet.append(["2026/01/02", "主评论内容B足够长", 1, 0, "保留子评论内容", "二级保留内容", "三级保留内容"])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )

        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned["SheetA"].iter_rows(values_only=True))
        main_comments = [row[1] for row in rows[1:]]
        subcomment_values = [row[4:7] for row in rows[1:]]

        self.assertEqual(["主评论内容A足够长", "主评论内容B足够长"], main_comments)
        self.assertEqual(
            [
                (None, "123456", None),
                ("保留子评论内容", "二级保留内容", "三级保留内容"),
            ],
            subcomment_values,
        )
        self.assertEqual(0, result.rows_deleted)
        self.assertEqual(2, result.cells_cleared)

    def test_cleaning_preserves_hash_id_column_values(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-preserve-hash-id"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "standardized.xlsx"
        hash_id = "a" * 64

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "产品名", "哈希ID", "点赞数"])
        sheet.append(["2026-07-16", "这是一条足够长的正常评论内容", "ScreenBar", hash_id, 3])
        workbook.save(input_path)

        result = clean_workbook(
            input_path=input_path,
            config=CleanerConfig(target_header="评论内容"),
            clean_words=(),
            output_dir=tmp / "out",
        )
        cleaned = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(cleaned.active.iter_rows(values_only=True))
        self.assertEqual(hash_id, rows[1][3])
        csv_text = result.output_csv.read_text(encoding="utf-8-sig")
        self.assertIn(hash_id, csv_text)

if __name__ == "__main__":
    unittest.main()
