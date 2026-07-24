from __future__ import annotations

import json
import unittest
from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.hash_id_pseudonymizer import (
    HashProjectContext,
    hash_display_name,
    load_hash_id_config,
)
from tools.standardize_excel_headers import (
    HeaderNotFoundError,
    MissingHashContextError,
    UnsafeIdentityValueError,
    UnsafeUserIdValueError,
    OutputPathConflictError,
    load_config,
    standardize_workbook,
)


EXPECTED_HEADER = [
    "评论日期",
    "评论内容",
    "产品名",
    "电商平台评分",
    "用户属性",
    "哈希ID",
    "点赞数",
    "子评论数/追评数",
    "一级评论",
    "二级评论",
    "三级评论",
]
HASH_ID_INDEX = EXPECTED_HEADER.index("哈希ID")
LIKES_INDEX = EXPECTED_HEADER.index("点赞数")


def expected_standardized_row(
    comment_date: object,
    comment_content: object,
    product_name: object = None,
    *,
    score: object = None,
    user_attribute: object = None,
    hash_id: object = None,
    likes: object = None,
    subcomment_count: object = None,
    level_one: object = None,
    level_two: object = None,
    level_three: object = None,
) -> tuple[object, ...]:
    return (
        comment_date,
        comment_content,
        product_name,
        score,
        user_attribute,
        hash_id,
        likes,
        subcomment_count,
        level_one,
        level_two,
        level_three,
    )


class StandardizeExcelHeadersTest(unittest.TestCase):
    def setUp(self) -> None:
        self.hash_context = HashProjectContext(
            project_id="project-screenbar",
            project_name="ScreenBar",
            key_version=1,
            key_fingerprint="test-fingerprint",
            secret_key=b"k" * 32,
        )
        self.hash_config = load_hash_id_config()

    def standardize_with_hash(
        self,
        input_path: Path,
        output_dir: Path,
        platform: str,
    ):
        return standardize_workbook(
            input_path,
            load_config(),
            output_dir=output_dir,
            platform=platform,
            hash_context=self.hash_context,
            hash_config=self.hash_config,
        )

    def read_standardized_rows(
        self,
        output_path: Path,
        sheet_name: str | None = None,
    ) -> list[tuple[object, ...]]:
        workbook = load_workbook(output_path, read_only=True, data_only=True)
        try:
            sheet = workbook[sheet_name] if sheet_name is not None else workbook.active
            return list(sheet.iter_rows(values_only=True))
        finally:
            workbook.close()

    def test_csv_formula_like_values_remain_text_after_standardization(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-csv-formula-text"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.csv"
        output_path = tmp / "standardized.xlsx"
        input_path.write_text(
            "timestamp,content,like_count\n1678870952,=1+1,=2+2\n",
            encoding="utf-8-sig",
        )

        standardize_workbook(
            input_path,
            load_config(),
            output_path=output_path,
            today=date(2026, 7, 21),
        )

        standardized = load_workbook(output_path, read_only=False, data_only=False)
        self.assertEqual("=1+1", standardized.active.cell(row=2, column=2).value)
        self.assertEqual("s", standardized.active.cell(row=2, column=2).data_type)
        self.assertEqual("=2+2", standardized.active.cell(row=2, column=LIKES_INDEX + 1).value)
        self.assertEqual("s", standardized.active.cell(row=2, column=LIKES_INDEX + 1).data_type)

    def test_compact_yyyymmdd_date_is_not_treated_as_unix_timestamp(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-compact-date"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["date", "comment", "likes"])
        sheet.append([20260709, "This is a sufficiently detailed product comment", 3])
        workbook.save(input_path)

        standardize_workbook(
            input_path,
            load_config(),
            output_path=output_path,
            today=date(2026, 7, 22),
        )

        rows = self.read_standardized_rows(output_path)
        self.assertEqual("2026-07-09", rows[1][0])

    def test_requires_explicit_overwrite_for_existing_output(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-existing-output"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "standardized.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["date", "comment", "likes"])
        sheet.append(["2026-07-22", "This is a detailed product comment", 1])
        workbook.save(input_path)
        output_path.write_text("existing", encoding="utf-8")

        with self.assertRaisesRegex(OutputPathConflictError, "already exists"):
            standardize_workbook(
                input_path,
                load_config(),
                output_path=output_path,
                overwrite=False,
            )

        self.assertEqual("existing", output_path.read_text(encoding="utf-8"))

    def test_standardizes_header_order_and_removes_risk_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-headers"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        output_dir = tmp / "out"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["IP地址", "用户名称", "评论内容", "评论时间", "点赞数", "子评论数", "一级评论内容", "引用的评论内容", "其它列"])
        sheet.append(["广东", "SHANE", "评论 A", "2026/01/05", 3, 2, "一级 A", "二级 A", "extra"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, output_dir, "taobao")

        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(("2026/01/05", "评论 A", None), rows[1][:3])
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual((3, 2, "一级 A", "二级 A", None), rows[1][LIKES_INDEX:])
        self.assertTrue(result.summary_json.exists())

        original = load_workbook(input_path, read_only=True, data_only=True)
        original_header = next(original["SheetA"].iter_rows(max_row=1, values_only=True))
        self.assertEqual(
            ("IP地址", "用户名称", "评论内容", "评论时间", "点赞数", "子评论数", "一级评论内容", "引用的评论内容", "其它列"),
            original_header,
        )

    def test_copies_confirmed_profile_columns_and_keeps_them_blank_when_absent(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-profile-columns"
        tmp.mkdir(parents=True, exist_ok=True)
        populated_input_path = tmp / "populated.xlsx"
        blank_input_path = tmp / "blank.xlsx"

        populated_workbook = Workbook()
        populated_sheet = populated_workbook.active
        populated_sheet.append(["评论日期", "评论内容", "电商平台评分", "性别", "年龄", "点赞数"])
        populated_sheet.append(["2026-07-24", "完整评论内容", 5, "女", 28, 3])
        populated_workbook.save(populated_input_path)

        blank_workbook = Workbook()
        blank_sheet = blank_workbook.active
        blank_sheet.append(["评论日期", "评论内容", "点赞数"])
        blank_sheet.append(["2026-07-24", "完整评论内容", 3])
        blank_workbook.save(blank_input_path)

        populated_result = standardize_workbook(
            populated_input_path,
            load_config(),
            output_dir=tmp / "populated-out",
        )
        blank_result = standardize_workbook(
            blank_input_path,
            load_config(),
            output_dir=tmp / "blank-out",
        )

        populated_rows = self.read_standardized_rows(populated_result.output_xlsx)
        blank_rows = self.read_standardized_rows(blank_result.output_xlsx)

        self.assertEqual(tuple(EXPECTED_HEADER), populated_rows[0])
        self.assertEqual((5, "女 28"), populated_rows[1][3:5])
        self.assertEqual(tuple(EXPECTED_HEADER), blank_rows[0])
        self.assertEqual((None, None), blank_rows[1][3:5])

    def test_merges_gender_and_age_into_the_single_user_attribute_column(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-user-attribute"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "性别", "年龄", "点赞数"])
        sheet.append(["2026-07-24", "完整评论内容 A", "女", 28, 1])
        sheet.append(["2026-07-24", "完整评论内容 B", "男", None, 2])
        sheet.append(["2026-07-24", "完整评论内容 C", None, 40, 3])
        sheet.append(["2026-07-24", "完整评论内容 D", None, None, 4])
        workbook.save(input_path)

        standardize_workbook(
            input_path,
            load_config(),
            output_path=output_path,
        )

        rows = self.read_standardized_rows(output_path)
        user_attribute_index = rows[0].index("用户属性")
        self.assertNotIn("性别", rows[0])
        self.assertNotIn("年龄", rows[0])
        self.assertEqual("女 28", rows[1][user_attribute_index])
        self.assertEqual("男", rows[2][user_attribute_index])
        self.assertEqual("40", rows[3][user_attribute_index])
        self.assertIsNone(rows[4][user_attribute_index])

    def test_combined_user_attribute_formula_like_value_remains_text(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-user-attribute-formula"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "性别", "年龄", "点赞数"])
        sheet.append(["2026-07-24", "完整评论内容", "=1+1", None, 1])
        workbook.save(input_path)

        standardize_workbook(
            input_path,
            load_config(),
            output_path=output_path,
        )

        standardized = load_workbook(output_path, read_only=False, data_only=False)
        try:
            user_attribute_column = list(
                next(standardized.active.iter_rows(min_row=1, max_row=1, values_only=True))
            ).index("用户属性") + 1
            cell = standardized.active.cell(row=2, column=user_attribute_column)
            self.assertEqual("=1+1", cell.value)
            self.assertEqual("s", cell.data_type)
        finally:
            standardized.close()

    def test_direct_user_attribute_has_row_level_priority_over_gender_and_age(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-direct-user-attribute"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "用户属性", "性别", "年龄", "点赞数"])
        sheet.append(["2026-07-24", "完整评论内容 A", "已登记属性", "女", 28, 1])
        sheet.append(["2026-07-24", "完整评论内容 B", "", "男", 40, 2])
        workbook.save(input_path)

        standardize_workbook(
            input_path,
            load_config(),
            output_path=output_path,
        )

        rows = self.read_standardized_rows(output_path)
        user_attribute_index = rows[0].index("用户属性")
        self.assertEqual("已登记属性", rows[1][user_attribute_index])
        self.assertEqual("男 40", rows[2][user_attribute_index])

    def test_rejects_missing_required_header_without_guessing(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-missing-header"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "子评论数", "一级评论", "二级评论"])
        sheet.append(["2026/01/05", "评论 A", 2, "一级 A", "二级 A"])
        workbook.save(input_path)

        with self.assertRaisesRegex(HeaderNotFoundError, "Available headers"):
            standardize_workbook(input_path, load_config(), output_dir=tmp / "out")

    def test_splits_taobao_comment_date_and_product_into_two_standard_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-taobao-date-product-header"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期与产品", "评论内容", "点赞数", "子评论数", "一级评论", "二级评论", "三级评论", "昵称", "IP属地"])
        sheet.append(["2026年6月21日 已购：触摸开关 / 黑", "评论 A", 3, 2, "一级 A", "二级 A", "三级 A", "user", "广东"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "taobao")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(
            expected_standardized_row(
                "2026年6月21日",
                "评论 A",
                "触摸开关 / 黑",
                likes=3,
                subcomment_count=2,
                level_one="一级 A",
                level_two="二级 A",
                level_three="三级 A",
            ),
            rows[1],
        )

    def test_maps_direct_product_header_to_product_name(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-direct-product-header"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "购买产品", "点赞数", "追评数", "一级评论", "二级评论", "三级评论"])
        sheet.append(["2026/06/21", "评论 A", "无线护眼屏幕挂灯", 3, 2, "一级 A", "二级 A", "三级 A"])
        workbook.save(input_path)

        result = standardize_workbook(
            input_path,
            load_config(),
            output_dir=tmp / "out",
            product_name="ScreenBar系列",
        )
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(
            expected_standardized_row(
                "2026/06/21",
                "评论 A",
                "无线护眼屏幕挂灯",
                likes=3,
                subcomment_count=2,
                level_one="一级 A",
                level_two="二级 A",
                level_three="三级 A",
            ),
            rows[1],
        )

    def test_fills_confirmed_product_name_when_source_product_column_is_absent(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-confirmed-product-name"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        output_path = tmp / "standardized.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数"])
        sheet.append(["2026-07-24", "完整评论内容", 3])
        workbook.save(input_path)

        standardize_workbook(
            input_path,
            load_config(),
            output_path=output_path,
            product_name="ScreenBar系列",
        )

        rows = self.read_standardized_rows(output_path, "SheetA")
        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(
            expected_standardized_row(
                "2026-07-24",
                "完整评论内容",
                "ScreenBar系列",
                likes=3,
            ),
            rows[1],
        )

    def test_maps_confirmed_taobao_aliases_to_standard_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-confirmed-taobao-aliases"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["用户名", "评论日期与产品", "评论内容", "评论数", "点赞量", "追评"])
        sheet.append(["user1", "2026年2月12日已购：浅灰色[MA270U(27英寸4K)]", "评论 A", 5, 3, "追评 A"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "taobao")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(("2026年2月12日", "评论 A", "浅灰色[MA270U(27英寸4K)]"), rows[1][:3])
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual((3, 5, "追评 A", None, None), rows[1][LIKES_INDEX:])

    def test_maps_confirmed_english_comment_aliases_and_drops_id_risk_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-confirmed-english-comment-aliases"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["rpid", "parent_rpid", "username", "content", "like_count", "timestamp", "ip_location", "子评论数"])
        sheet.append([156365080, 0, "user1", "评论 A", 2, 1678870952, "未知", 0])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "bilibili")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(("2023-03-15", "评论 A", None), rows[1][:3])
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual((2, 0, None, None, None), rows[1][LIKES_INDEX:])

        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        self.assertEqual(
            ["rpid", "parent_rpid", "username", "ip_location"],
            summary["sheets"][0]["configured_drop_headers_found"],
        )

    def test_keeps_subcomment_count_output_column_blank_when_source_header_is_missing(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-missing-subcomment-count"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["timestamp", "content", "like_count", "username", "ip_location"])
        sheet.append(["1678870952", "评论 A", "2", "user1", "未知"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "bilibili")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(("2023-03-15", "评论 A", None), rows[1][:3])
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual(("2", None, None, None, None), rows[1][LIKES_INDEX:])

    def test_standardizes_csv_input_with_same_header_mapping_rules(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-csv-input"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.csv"
        input_path.write_text(
            "timestamp,content,like_count,子评论数,username,ip_location\n"
            "1678870952,评论 A,2,0,user1,未知\n",
            encoding="utf-8-sig",
        )

        result = self.standardize_with_hash(input_path, tmp / "out", "bilibili")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["raw"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(("2023-03-15", "评论 A", None), rows[1][:3])
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual(("2", "0", None, None, None), rows[1][LIKES_INDEX:])

    def test_maps_tiktok_comment_exporter_aliases_without_ai(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-tiktok-aliases"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.csv"
        input_path.write_text(
            "id,uniqueId,text,diggCount,replyCommentTotal,createTime\n"
            "1001,user_a,This monitor light works great,4,7,1678870952\n",
            encoding="utf-8-sig",
        )

        result = standardize_workbook(input_path, load_config(), output_dir=tmp / "out")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["raw"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(
            expected_standardized_row(
                "2023-03-15",
                "This monitor light works great",
                likes="4",
                subcomment_count="7",
            ),
            rows[1],
        )

        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        self.assertEqual(["id", "uniqueId"], summary["sheets"][0]["configured_drop_headers_found"])

    def test_maps_confirmed_tiktok_chinese_exporter_aliases_without_ai(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-tiktok-chinese-aliases"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.csv"
        input_path.write_text(
            "评论ID,回复哪个评论,用户身份,用户名,昵称,评论,评论时间,Digg Count,Author Digged,回复数,固定到顶部,用户主页\n"
            "1001,0,user,user_a,Nick,This monitor light works great,1678870952,4,false,7,false,https://example.com/user\n",
            encoding="utf-8-sig",
        )

        result = self.standardize_with_hash(input_path, tmp / "out", "tiktok")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["raw"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(("2023-03-15", "This monitor light works great", None), rows[1][:3])
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual(("4", "7", None, None, None), rows[1][LIKES_INDEX:])

    def test_maps_confirmed_tiktok_chinese_datetime_without_time_output(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-tiktok-chinese-datetime"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.csv"
        input_path.write_text(
            "评论,评论时间,Digg Count,回复数\n"
            "This monitor light works great,2026/7/7 08:24:05,4,7\n",
            encoding="utf-8-sig",
        )

        result = standardize_workbook(input_path, load_config(), output_dir=tmp / "out")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["raw"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(
            expected_standardized_row(
                "2026-07-07",
                "This monitor light works great",
                likes="4",
                subcomment_count="7",
            ),
            rows[1],
        )

    def test_maps_youtube_comment_aliases_and_iso_time_to_beijing_date(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-youtube-aliases"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["authorDisplayName", "commentText", "likeCount", "replyCount", "publishedAt", "videoId", "replyText"])
        sheet.append(["User A", "Great light for my desk", 12, 3, "2026-07-08T23:30:00Z", "abc123", "Thanks for sharing"])
        workbook.save(input_path)

        result = standardize_workbook(input_path, load_config(), output_dir=tmp / "out")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(
            expected_standardized_row(
                "2026-07-09",
                "Great light for my desk",
                likes=12,
                subcomment_count=3,
                level_one="Thanks for sharing",
            ),
            rows[1],
        )

    def test_maps_youtube_relative_published_at_to_beijing_year_or_month(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-youtube-relative-published-at"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["published_at", "text", "likes", "reply_count"])
        sheet.append(["1年前", "One year old comment", 1, 0])
        sheet.append(["2年前", "Two years old comment", 2, 0])
        sheet.append(["9个月前", "Nine months old comment", 3, 0])
        sheet.append(["4个月前", "Four months old comment", 4, 0])
        workbook.save(input_path)

        result = standardize_workbook(
            input_path,
            load_config(),
            output_dir=tmp / "out",
            today=date(2026, 7, 9),
        )
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual(expected_standardized_row("2025", "One year old comment", likes=1, subcomment_count=0), rows[1])
        self.assertEqual(expected_standardized_row("2024", "Two years old comment", likes=2, subcomment_count=0), rows[2])
        self.assertEqual(expected_standardized_row("2025-10", "Nine months old comment", likes=3, subcomment_count=0), rows[3])
        self.assertEqual(expected_standardized_row("2026-03", "Four months old comment", likes=4, subcomment_count=0), rows[4])

    def test_hashes_verified_youtube_user_id_without_exposing_raw_id_in_summary(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-youtube-user-id"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["publishedAt", "commentText", "likeCount", "author_channel_id"])
        sheet.append(["2026-07-08T23:30:00Z", "Great light for my desk", 12, "UC-secret-user"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "YouTube")
        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=True)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))

        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        summary_text = result.summary_json.read_text(encoding="utf-8")
        self.assertNotIn("UC-secret-user", summary_text)
        summary = json.loads(summary_text)
        self.assertEqual("youtube", summary["hash_id"]["platform"])
        self.assertEqual("author_channel_id", summary["sheets"][0]["hash_id"]["source_header"])
        self.assertEqual("account_id", summary["sheets"][0]["hash_id"]["identity_type"])
        self.assertEqual(1, summary["sheets"][0]["hash_id"]["hashed_count"])

    def test_hash_summary_records_reason_only_when_no_identity_header_exists(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-no-registered-identity-header"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        without_identity = workbook.active
        without_identity.title = "WithoutIdentity"
        without_identity.append(
            ["publishedAt", "commentText", "likeCount", "profileUrl"]
        )
        without_identity.append(
            [
                "2026-07-08",
                "Comment without a registered identity header",
                1,
                "https://example.invalid/raw-profile-secret",
            ]
        )
        with_identity = workbook.create_sheet("WithIdentity")
        with_identity.append(["publishedAt", "commentText", "likeCount", "author"])
        with_identity.append(
            [
                "2026-07-09",
                "Comment with a registered identity header",
                2,
                "raw-name-secret",
            ]
        )
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "YouTube")
        summary_text = result.summary_json.read_text(encoding="utf-8")
        self.assertNotIn("raw-profile-secret", summary_text)
        self.assertNotIn("raw-name-secret", summary_text)
        summary = json.loads(summary_text)
        sheets = {sheet["sheet_name"]: sheet for sheet in summary["sheets"]}

        self.assertEqual(
            "no_registered_identity_header",
            sheets["WithoutIdentity"]["hash_id"]["reason"],
        )
        self.assertNotIn("reason", sheets["WithIdentity"]["hash_id"])

    def test_hashes_xiaohongshu_user_id_stably_across_runs(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-xiaohongshu-user-id"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "点赞数", "用户ID"])
        sheet.append(["2026-07-08", "评论内容足够完整", 3, "xhs-user-001"])
        workbook.save(input_path)

        first = self.standardize_with_hash(input_path, tmp / "first", "小红书")
        second = self.standardize_with_hash(input_path, tmp / "second", "xiaohongshu")
        first_rows = list(load_workbook(first.output_xlsx, read_only=True, data_only=True).active.iter_rows(values_only=True))
        second_rows = list(load_workbook(second.output_xlsx, read_only=True, data_only=True).active.iter_rows(values_only=True))
        self.assertEqual(first_rows[1][HASH_ID_INDEX], second_rows[1][HASH_ID_INDEX])

    def test_hashes_bilibili_username_as_display_name_without_exposing_raw_name(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-bilibili-display-name"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"
        sheet.append(["timestamp", "content", "like_count", "username"])
        sheet.append(["1678870952", "评论内容足够完整一", 4, "same-secret-name"])
        sheet.append(["1678870953", "评论内容足够完整二", 5, "same-secret-name"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "B站")
        rows = self.read_standardized_rows(result.output_xlsx, "Comments")
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual(rows[1][HASH_ID_INDEX], rows[2][HASH_ID_INDEX])
        self.assertNotIn("username", rows[0])
        self.assertNotIn("same-secret-name", str(rows))
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertEqual("display_name", hash_summary["identity_type"])
        self.assertEqual("username", hash_summary["source_header"])
        self.assertEqual(2, hash_summary["hashed_count"])
        self.assertEqual(0, hash_summary["blank_count"])

    def test_youtube_account_id_column_wins_globally_without_row_fallback(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-youtube-account-id-priority"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["publishedAt", "commentText", "likeCount", "author_channel_id", "author"])
        sheet.append(["2026-07-08", "First useful comment", 1, "UC-stable", "Name A"])
        sheet.append(["2026-07-09", "Second useful comment", 2, None, "Name B"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "youtube")
        rows = self.read_standardized_rows(result.output_xlsx)
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertIsNone(rows[2][HASH_ID_INDEX])
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertEqual("account_id", hash_summary["identity_type"])
        self.assertEqual("author_channel_id", hash_summary["source_header"])
        self.assertEqual(1, hash_summary["hashed_count"])
        self.assertEqual(1, hash_summary["blank_count"])

    def test_youtube_uses_author_when_account_id_column_is_entirely_blank(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-youtube-blank-account-id-fallback"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["publishedAt", "commentText", "likeCount", "author_channel_id", "author"])
        sheet.append(["2026-07-08", "First useful comment", 1, None, "Same Name"])
        sheet.append(["2026-07-09", "Second useful comment", 2, "  ", "Same Name"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "youtube")
        rows = self.read_standardized_rows(result.output_xlsx)
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual(rows[1][HASH_ID_INDEX], rows[2][HASH_ID_INDEX])
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertEqual("display_name", hash_summary["identity_type"])
        self.assertEqual("author", hash_summary["source_header"])
        self.assertEqual(2, hash_summary["hashed_count"])
        self.assertEqual(0, hash_summary["blank_count"])

    def test_youtube_uses_author_only_when_account_id_column_is_absent(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-youtube-author-fallback"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["publishedAt", "commentText", "likeCount", "author"])
        sheet.append(["2026-07-08", "Useful comment text", 1, "Visible Name"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "YouTube")
        rows = self.read_standardized_rows(result.output_xlsx)
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertEqual("author", hash_summary["source_header"])
        self.assertEqual("display_name", hash_summary["identity_type"])

    def test_youtube_author_name_reuses_author_display_name_hash_domain(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-youtube-author-name-fallback"
        tmp.mkdir(parents=True, exist_ok=True)
        author_path = tmp / "author.xlsx"
        author_name_path = tmp / "author-name.xlsx"

        for path, identity_header in (
            (author_path, "author"),
            (author_name_path, "author_name"),
        ):
            workbook = Workbook()
            sheet = workbook.active
            sheet.append(["publishedAt", "commentText", "likeCount", identity_header])
            sheet.append(["2026-07-08", "Useful comment text", 1, "Same Visible Name"])
            workbook.save(path)

        first = self.standardize_with_hash(author_path, tmp / "author-out", "YouTube")
        second = self.standardize_with_hash(
            author_name_path,
            tmp / "author-name-out",
            "YouTube",
        )
        first_rows = self.read_standardized_rows(first.output_xlsx)
        second_rows = self.read_standardized_rows(second.output_xlsx)
        self.assertRegex(second_rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        self.assertEqual(first_rows[1][HASH_ID_INDEX], second_rows[1][HASH_ID_INDEX])
        summary = json.loads(second.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertEqual("author_name", hash_summary["source_header"])
        self.assertEqual("display_name", hash_summary["identity_type"])

    def test_xiaohongshu_uses_display_name_when_user_id_column_is_absent(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-xhs-display-name-fallback"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "点赞数", "用户名称"])
        sheet.append(["2026-07-08", "评论内容足够完整", 3, "小红书用户"])
        workbook.save(input_path)

        result = self.standardize_with_hash(input_path, tmp / "out", "小红书")
        rows = self.read_standardized_rows(result.output_xlsx)
        self.assertRegex(rows[1][HASH_ID_INDEX], r"^[0-9a-f]{64}$")
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertEqual("用户名称", hash_summary["source_header"])
        self.assertEqual("display_name", hash_summary["identity_type"])

    def test_tiktok_prefers_username_and_uses_nickname_only_when_username_column_is_absent(self) -> None:
        cases = [
            ("username-present", ["用户名", "昵称"], ["priority-name", "fallback-name"], "用户名"),
            ("nickname-only", ["昵称"], ["fallback-name"], "昵称"),
        ]
        for case_name, identity_headers, identity_values, expected_header in cases:
            with self.subTest(case=case_name):
                tmp = Path.cwd() / ".tmp-tests" / f"case-hash-tiktok-{case_name}"
                tmp.mkdir(parents=True, exist_ok=True)
                input_path = tmp / "raw.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["评论", "评论时间", "Digg Count", *identity_headers])
                sheet.append(["Long enough comment", "1678870952", 4, *identity_values])
                workbook.save(input_path)

                result = self.standardize_with_hash(input_path, tmp / "out", "TikTok")
                rows = self.read_standardized_rows(result.output_xlsx)
                selected_value = identity_values[identity_headers.index(expected_header)]
                expected_hash = hash_display_name(
                    selected_value,
                    "TikTok",
                    self.hash_context,
                    self.hash_config,
                )
                self.assertEqual(expected_hash, rows[1][HASH_ID_INDEX])
                if case_name == "username-present":
                    fallback_hash = hash_display_name(
                        "fallback-name",
                        "TikTok",
                        self.hash_context,
                        self.hash_config,
                    )
                    self.assertNotEqual(fallback_hash, rows[1][HASH_ID_INDEX])
                summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
                hash_summary = summary["sheets"][0]["hash_id"]
                self.assertEqual(expected_header, hash_summary["source_header"])
                self.assertEqual("display_name", hash_summary["identity_type"])

    def test_taobao_and_jd_select_configured_display_name_priority(self) -> None:
        cases = [
            ("taobao", ["用户名称", "用户名"], ["priority-name", "other-name"], "用户名称"),
            ("jd", ["用户名"], ["jd-name"], "用户名"),
        ]
        for platform, identity_headers, identity_values, expected_header in cases:
            with self.subTest(platform=platform):
                tmp = Path.cwd() / ".tmp-tests" / f"case-hash-{platform}-display-name"
                tmp.mkdir(parents=True, exist_ok=True)
                input_path = tmp / "raw.xlsx"
                workbook = Workbook()
                sheet = workbook.active
                sheet.append(["评论日期", "评论内容", "点赞数", *identity_headers])
                sheet.append(["2026-07-08", "评论内容足够完整", 3, *identity_values])
                workbook.save(input_path)

                result = self.standardize_with_hash(input_path, tmp / "out", platform)
                rows = self.read_standardized_rows(result.output_xlsx)
                selected_value = identity_values[identity_headers.index(expected_header)]
                expected_hash = hash_display_name(
                    selected_value,
                    platform,
                    self.hash_context,
                    self.hash_config,
                )
                self.assertEqual(expected_hash, rows[1][HASH_ID_INDEX])
                if platform == "taobao":
                    fallback_hash = hash_display_name(
                        "other-name",
                        platform,
                        self.hash_context,
                        self.hash_config,
                    )
                    self.assertNotEqual(fallback_hash, rows[1][HASH_ID_INDEX])
                summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
                hash_summary = summary["sheets"][0]["hash_id"]
                self.assertEqual(expected_header, hash_summary["source_header"])
                self.assertEqual("display_name", hash_summary["identity_type"])

    def test_supplied_platform_ignores_identity_header_from_other_platform(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-cross-platform-identity-header"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"
        sheet.append(["评论", "评论时间", "Digg Count", "author"])
        sheet.append(["Long enough comment", "1678870952", 4, "youtube-only-name"])
        workbook.save(input_path)

        result = standardize_workbook(
            input_path,
            load_config(),
            output_dir=tmp / "out",
            platform="TikTok",
            hash_config=self.hash_config,
        )

        rows = self.read_standardized_rows(result.output_xlsx, "Comments")
        self.assertIsNone(rows[1][HASH_ID_INDEX])
        summary = json.loads(result.summary_json.read_text(encoding="utf-8"))
        hash_summary = summary["sheets"][0]["hash_id"]
        self.assertIsNone(hash_summary["identity_type"])
        self.assertIsNone(hash_summary["source_header"])

    def test_registered_real_user_id_requires_platform_and_project_context(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-hash-context-required"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["publishedAt", "commentText", "likeCount", "author_channel_id"])
        sheet.append(["2026-07-08T23:30:00Z", "Great light for my desk", 12, "UC-secret-user"])
        workbook.save(input_path)

        with self.assertRaises(MissingHashContextError):
            standardize_workbook(input_path, load_config(), output_dir=tmp / "without-context")

    def test_registered_author_without_platform_still_requires_context(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-author-context-required"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"
        sheet.append(["publishedAt", "commentText", "likeCount", "author"])
        sheet.append(["2026-07-08", "Useful comment text", 1, "SECRET-AUTHOR"])
        workbook.save(input_path)

        with self.assertRaises(MissingHashContextError) as raised:
            standardize_workbook(input_path, load_config(), output_dir=tmp / "out")
        message = str(raised.exception)
        self.assertIn("Comments", message)
        self.assertIn("author", message)
        self.assertNotIn("SECRET-AUTHOR", message)

    def test_registered_display_name_requires_context_without_exposing_raw_value(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-display-name-context-required"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"
        sheet.append(["timestamp", "content", "like_count", "username"])
        sheet.append(["1678870952", "评论内容足够完整", 4, "SECRET-DISPLAY-NAME"])
        workbook.save(input_path)

        with self.assertRaises(MissingHashContextError) as raised:
            standardize_workbook(input_path, load_config(), output_dir=tmp / "without-context")
        message = str(raised.exception)
        self.assertIn("Comments", message)
        self.assertIn("username", message)
        self.assertIn("identity value/column", message)
        self.assertNotIn("SECRET-DISPLAY-NAME", message)

    def test_invalid_identity_error_is_safe_and_legacy_compatible(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-invalid-hash-user-id"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Comments"
        sheet.append(["publishedAt", "commentText", "likeCount", "author_channel_id"])
        sheet.append(["2026-07-08T23:30:00Z", "Great light for my desk", 12, "=SECRET_RAW_ID"])
        workbook.save(input_path)

        self.assertTrue(issubclass(UnsafeIdentityValueError, UnsafeUserIdValueError))
        self.assertIs(UnsafeIdentityValueError, UnsafeUserIdValueError)
        with self.assertRaises(UnsafeIdentityValueError) as raised:
            self.standardize_with_hash(input_path, tmp / "out", "youtube")
        self.assertIsInstance(raised.exception, UnsafeUserIdValueError)
        message = str(raised.exception)
        self.assertIn("identity value/column", message)
        self.assertNotIn("user ID", message)
        self.assertIn("Comments", message)
        self.assertIn("row 2", message)
        self.assertIn("author_channel_id", message)
        self.assertIn("account_id", message)
        self.assertIn("InvalidUserIdError", message)
        self.assertNotIn("SECRET_RAW_ID", message)

    def test_preserves_formula_cells_in_retained_columns(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-standardize-preserves-formulas"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "raw.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "SheetA"
        sheet.append(["评论日期", "评论内容", "点赞数"])
        sheet.append(["2026-07-08", "评论内容足够完整", "=1+1"])
        workbook.save(input_path)

        result = standardize_workbook(input_path, load_config(), output_dir=tmp / "out")

        standardized = load_workbook(result.output_xlsx, read_only=True, data_only=False)
        rows = list(standardized["SheetA"].iter_rows(values_only=True))
        self.assertEqual(tuple(EXPECTED_HEADER), rows[0])
        self.assertEqual("=1+1", rows[1][LIKES_INDEX])



if __name__ == "__main__":
    unittest.main()
