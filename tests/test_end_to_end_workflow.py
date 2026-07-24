from __future__ import annotations

import csv
import unittest
from contextlib import nullcontext
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.clean_excel_comments import CleanerConfig, clean_workbook
from tools.cleanup_intermediate_outputs import cleanup_intermediate_outputs
from tools.hash_id_pseudonymizer import HashProjectContext, load_hash_id_config
from tools.merge_excel_workbooks import merge_workbooks
from tools.standardize_excel_headers import load_config, standardize_workbook
from tools.strip_bilibili_reply_prefixes import strip_bilibili_reply_prefixes


EXPECTED_HEADERS = (
    "评论日期",
    "评论内容",
    "产品名",
    "电商平台评分",
    "性别",
    "年龄",
    "哈希ID",
    "点赞数",
    "子评论数/追评数",
    "一级评论",
    "二级评论",
    "三级评论",
)
HASH_ID_INDEX = EXPECTED_HEADERS.index("哈希ID")
LIKES_INDEX = EXPECTED_HEADERS.index("点赞数")


def write_source(path: Path, rows: list[list[object]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "评论"
    sheet.append(["rpid", "parent_rpid", "username", "content", "like_count", "timestamp", "ip_location", "replyCount"])
    for row in rows:
        sheet.append(row)
    workbook.save(path)


class EndToEndWorkflowTest(unittest.TestCase):
    def test_merge_standardize_clean_and_cleanup_preserves_inputs_and_keeps_only_final_outputs(self) -> None:
        base = Path.cwd() / ".tmp-tests"
        base.mkdir(parents=True, exist_ok=True)

        test_directory = base / "case-end-to-end-workflow"
        test_directory.mkdir(parents=True, exist_ok=True)
        for existing_file in test_directory.iterdir():
            if existing_file.is_file():
                existing_file.unlink()
        with nullcontext(test_directory) as temporary_directory:
            tmp = Path(temporary_directory)
            first = tmp / "source1.xlsx"
            second = tmp / "source2.xlsx"
            merged = tmp / "20260710_product_B站评论数据_合并总表.xlsx"
            stripped = tmp / "20260710_product_B站评论数据_回复前缀已清理.xlsx"
            standardized = tmp / "20260710_product_B站评论数据_标准化总表.xlsx"
            cleaned = tmp / "20260710_product_B站评论数据_清洗后总表.xlsx"

            write_source(
                first,
                [["1", "0", "user1", "这是一条足够长的正常主评论内容", "=1+1", "1547698467", "北京", "1"]],
            )
            write_source(
                second,
                [["2", "1", "user1", "回复@user1：这是另一条足够长的正常回复内容", "3", "1547698468", "上海", "0"]],
            )
            original_bytes = {first: first.read_bytes(), second: second.read_bytes()}
            hash_context = HashProjectContext(
                project_id="project-screenbar-e2e",
                project_name="ScreenBar十周年专案",
                key_version=1,
                key_fingerprint="test-fingerprint",
                secret_key=b"e" * 32,
            )
            expected_hash = "cad6b15fb4291d29dc08a17581aec2fa0726ea09b1786dd1ef5bf95b5b224b45"

            merge_result = merge_workbooks([first, second], merged)
            strip_result = strip_bilibili_reply_prefixes(merged, stripped)
            standardize_result = standardize_workbook(
                stripped,
                load_config(),
                output_path=standardized,
                platform="B站",
                hash_context=hash_context,
                hash_config=load_hash_id_config(),
            )
            clean_result = clean_workbook(
                standardized,
                CleanerConfig(target_header="评论内容"),
                clean_words=(),
                output_path=cleaned,
            )

            self.assertTrue(clean_result.output_csv and clean_result.output_csv.exists())
            cleaned_workbook = load_workbook(cleaned, read_only=True, data_only=False)
            try:
                cleaned_rows = list(cleaned_workbook["总表"].iter_rows(values_only=True))
            finally:
                cleaned_workbook.close()
            self.assertEqual(EXPECTED_HEADERS, cleaned_rows[0])
            self.assertEqual("=1+1", cleaned_rows[1][LIKES_INDEX])
            self.assertEqual("这是另一条足够长的正常回复内容", cleaned_rows[2][1])
            first_hash = cleaned_rows[1][HASH_ID_INDEX]
            second_hash = cleaned_rows[2][HASH_ID_INDEX]
            self.assertRegex(first_hash, r"^[0-9a-f]{64}$")
            self.assertRegex(second_hash, r"^[0-9a-f]{64}$")
            self.assertEqual(first_hash, second_hash)
            self.assertEqual(expected_hash, first_hash)
            self.assertEqual(expected_hash, second_hash)
            for row in cleaned_rows[1:]:
                self.assertNotIn("user1", row)
                self.assertNotIn("北京", row)
                self.assertNotIn("上海", row)

            with clean_result.output_csv.open("r", encoding="utf-8-sig", newline="") as csv_file:
                csv_rows = list(csv.reader(csv_file))
            self.assertEqual(list(EXPECTED_HEADERS), csv_rows[0])
            self.assertEqual([expected_hash, expected_hash], [row[HASH_ID_INDEX] for row in csv_rows[1:]])
            csv_cells = [cell for row in csv_rows for cell in row]
            self.assertNotIn("user1", csv_cells)
            self.assertNotIn("北京", csv_cells)
            self.assertNotIn("上海", csv_cells)

            intermediate_paths = [
                merge_result.output_path,
                merge_result.summary_path,
                strip_result.output_xlsx,
                strip_result.summary_json,
                standardize_result.output_xlsx,
                standardize_result.summary_json,
                clean_result.deletion_log_csv,
                clean_result.summary_json,
            ]
            cleanup_result = cleanup_intermediate_outputs(
                intermediate_paths=intermediate_paths,
                protected_paths=[first, second, clean_result.output_xlsx, clean_result.output_csv],
                summary_path=None,
            )

            self.assertEqual(8, cleanup_result.files_deleted)
            self.assertIsNone(cleanup_result.summary_json)
            self.assertTrue(all(not path.exists() for path in intermediate_paths))
            self.assertEqual(original_bytes[first], first.read_bytes())
            self.assertEqual(original_bytes[second], second.read_bytes())
            self.assertEqual(
                sorted([first.name, second.name, cleaned.name]),
                sorted(path.name for path in tmp.glob("*.xlsx")),
            )
            self.assertEqual(
                [clean_result.output_csv.name],
                [path.name for path in tmp.glob("*.csv")],
            )
            self.assertEqual([], list(tmp.glob("*.json")))


if __name__ == "__main__":
    unittest.main()
