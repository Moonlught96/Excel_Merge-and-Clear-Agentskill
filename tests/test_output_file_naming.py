from __future__ import annotations

import io
import unittest
from datetime import datetime
from pathlib import Path
from unittest import mock

from openpyxl import Workbook, load_workbook

from tools.output_file_naming import (
    build_naming_plan,
    product_candidates_from_workbook,
    write_json_document,
)


class OutputFileNamingTest(unittest.TestCase):
    def test_json_output_reconfigures_gbk_stream_for_emoji_filenames(self) -> None:
        raw = io.BytesIO()
        stream = io.TextIOWrapper(raw, encoding="gbk")

        write_json_document({"filename": "comments 🤯.xlsx"}, stream)
        stream.flush()

        self.assertIn("comments 🤯.xlsx", raw.getvalue().decode("utf-8"))
        stream.detach()

    def test_builds_confirmable_names_from_filename_before_merge(self) -> None:
        plan = build_naming_plan(
            [Path("1783327727ScreenBar Halo2淘宝评论数据.xlsx")],
            today=datetime(2026, 7, 7, 9, 30),
        )

        self.assertEqual("20260707", plan.date_text)
        self.assertEqual("ScreenBar Halo2", plan.product_name)
        self.assertEqual("淘宝评论数据", plan.data_source)
        self.assertEqual([], plan.missing_fields)
        self.assertEqual(
            "20260707_ScreenBar Halo2_淘宝评论数据_合并总表.xlsx",
            plan.filenames["merge"],
        )
        self.assertEqual(
            "20260707_ScreenBar Halo2_淘宝评论数据_标准化总表.xlsx",
            plan.filenames["standardized"],
        )
        self.assertEqual(
            "20260707_ScreenBar Halo2_淘宝评论数据_清洗后总表.xlsx",
            plan.filenames["cleaned"],
        )
        self.assertEqual(
            "20260707_ScreenBar Halo2_淘宝评论数据_清洗后总表.csv",
            plan.filenames["cleaned_csv"],
        )

    def test_builds_names_from_parent_folder_and_source_path_when_filename_is_generic(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/专案/26_AI叙事策略锚定/02_收集的数据/未清洗数据/"
                    "社媒_小红书/明基MA/【靓号】根据链接采集笔记评论-20260317-153317.xlsx"
                )
            ],
            today=datetime(2026, 7, 7, 9, 30),
        )

        self.assertEqual("明基MA", plan.product_name)
        self.assertEqual("小红书评论数据", plan.data_source)
        self.assertEqual("20260707_明基MA_小红书评论数据_合并总表.xlsx", plan.filenames["merge"])

    def test_asks_user_when_product_is_ambiguous(self) -> None:
        plan = build_naming_plan(
            [
                Path("1783326608ScreenBar Pro淘宝评论数据.xlsx"),
                Path("1783327727ScreenBar Halo2淘宝评论数据.xlsx"),
            ],
            today=datetime(2026, 7, 7, 9, 30),
        )

        self.assertIn("product_name", plan.missing_fields)
        self.assertEqual(["ScreenBar Halo2", "ScreenBar Pro"], plan.product_candidates)
        self.assertEqual({}, plan.filenames)

    def test_can_use_product_from_workbook_when_filename_and_path_do_not_have_it(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-output-name-workbook-product"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "淘宝评论数据.xlsx"

        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["评论日期", "评论内容", "购买产品"])
        sheet.append(["2026/07/07", "评论内容", "ScreenBar Halo2"])
        sheet.append(["2026/07/07", "评论内容2", "ScreenBar Halo2"])
        workbook.save(input_path)

        plan = build_naming_plan([input_path], today=datetime(2026, 7, 7, 9, 30))

        self.assertEqual("ScreenBar Halo2", plan.product_name)
        self.assertEqual("淘宝评论数据", plan.data_source)
        self.assertEqual("20260707_ScreenBar Halo2_淘宝评论数据_合并总表.xlsx", plan.filenames["merge"])

    def test_can_use_product_from_csv_when_filename_and_path_do_not_have_it(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-output-name-csv-product"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "淘宝评论数据.csv"
        input_path.write_text(
            "评论日期,评论内容,购买产品\n"
            "2026/07/07,评论内容,ScreenBar Halo2\n",
            encoding="utf-8-sig",
        )

        plan = build_naming_plan([input_path], today=datetime(2026, 7, 7, 9, 30))

        self.assertEqual("ScreenBar Halo2", plan.product_name)
        self.assertEqual("淘宝评论数据", plan.data_source)
        self.assertEqual("20260707_ScreenBar Halo2_淘宝评论数据_合并总表.xlsx", plan.filenames["merge"])

    def test_detects_tiktok_and_youtube_sources_from_paths(self) -> None:
        tiktok_plan = build_naming_plan(
            [Path("D:/专案/ScreenBar十周年专案/产品数据/Tiktok/TTCommentExporter-7371053582810008864-51-comments-replies.csv")],
            product_name="ScreenBar系列",
            today=datetime(2026, 7, 7, 9, 30),
        )
        youtube_plan = build_naming_plan(
            [Path("D:/专案/ScreenBar十周年专案/产品数据/youtube数据/长视频评论/yt-comments-2026-07-08-15-35-54.xlsx")],
            product_name="ScreenBar系列",
            today=datetime(2026, 7, 7, 9, 30),
        )

        self.assertEqual("TikTok评论数据", tiktok_plan.data_source)
        self.assertEqual("YouTube评论数据", youtube_plan.data_source)
        self.assertEqual("20260707_ScreenBar系列_TikTok评论数据_合并总表.xlsx", tiktok_plan.filenames["merge"])
        self.assertEqual("20260707_ScreenBar系列_YouTube评论数据_合并总表.xlsx", youtube_plan.filenames["merge"])

    def test_detects_twitter_source_and_preprocessing_route(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/Twitter/"
                    "twitter-screenbar搜索结果-1784787464581.csv"
                )
            ],
            product_name="ScreenBar系列",
            today=datetime(2026, 7, 24, 9, 30),
        )

        self.assertEqual("Twitter评论数据", plan.data_source)
        self.assertEqual("twitter", plan.preprocessing_profile)
        self.assertEqual(
            "20260724_ScreenBar系列_Twitter评论数据_合并总表.xlsx",
            plan.filenames["merge"],
        )

    def test_detects_amazon_source_from_registered_path_keywords(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/2017_ScreenBar/亚马逊日本/"
                    "亚马逊_screenbar_4星最热.xlsx"
                )
            ],
            product_name="ScreenBar",
            today=datetime(2026, 7, 24, 9, 30),
        )

        self.assertEqual("亚马逊日本评论数据", plan.data_source)
        self.assertEqual(
            "20260724_ScreenBar_亚马逊日本评论数据_合并总表.xlsx",
            plan.filenames["merge"],
        )

    def test_distinguishes_amazon_region_in_display_source_name(self) -> None:
        japan_plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/2017_ScreenBar/亚马逊日本/"
                    "亚马逊_screenbar_4星最热.xlsx"
                )
            ],
            product_name="2017_ScreenBar",
            today=datetime(2026, 7, 24, 9, 30),
        )
        us_plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/2017_ScreenBar/亚马逊美国/"
                    "amazon_screenbar_recent.xlsx"
                )
            ],
            product_name="2017_ScreenBar",
            today=datetime(2026, 7, 24, 9, 30),
        )

        self.assertEqual("亚马逊日本评论数据", japan_plan.data_source)
        self.assertEqual("亚马逊美国评论数据", us_plan.data_source)
        self.assertEqual(
            "20260724_2017_ScreenBar_亚马逊日本评论数据_合并总表.xlsx",
            japan_plan.filenames["merge"],
        )
        self.assertEqual(
            "20260724_2017_ScreenBar_亚马逊美国评论数据_合并总表.xlsx",
            us_plan.filenames["merge"],
        )

    def test_uses_amazon_region_parent_for_product_and_exposes_preprocessing_route(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/2017_ScreenBar/亚马逊日本/"
                    "亚马逊_screenbar_4星最热.xlsx"
                )
            ],
            today=datetime(2026, 7, 24, 9, 30),
        )

        self.assertEqual("2017_ScreenBar", plan.product_name)
        self.assertEqual("亚马逊日本评论数据", plan.data_source)
        self.assertEqual("amazon", plan.preprocessing_profile)
        self.assertEqual([], plan.missing_fields)
        self.assertEqual(
            "20260724_2017_ScreenBar_亚马逊日本评论数据_合并总表.xlsx",
            plan.filenames["merge"],
        )

    def test_detects_rakuten_market_and_exposes_its_preprocessing_route(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/乐天市场/"
                    "乐天市场_ScreenBar 1.xlsx.xlsx"
                )
            ],
            today=datetime(2026, 7, 24, 9, 30),
        )

        self.assertEqual("ScreenBar 1", plan.product_name)
        self.assertEqual("乐天市场评论数据", plan.data_source)
        self.assertEqual("rakuten", plan.preprocessing_profile)
        self.assertEqual([], plan.missing_fields)
        self.assertEqual(
            "20260724_ScreenBar 1_乐天市场评论数据_合并总表.xlsx",
            plan.filenames["merge"],
        )

    def test_detects_youtube_shorts_before_generic_youtube_source(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/专案/ScreenBar十周年专案/产品数据/youtube数据/Shorts/"
                    "comments This new screenbar is so good 🤯.xlsx"
                )
            ],
            product_name="ScreenBar系列",
            today=datetime(2026, 7, 22, 9, 30),
        )

        self.assertEqual("YouTube Shorts评论数据", plan.data_source)
        self.assertEqual(
            "20260722_ScreenBar系列_YouTube Shorts评论数据_合并总表.xlsx",
            plan.filenames["merge"],
        )

    def test_uses_bilibili_parent_folder_as_product_and_preserves_release_year(self) -> None:
        plan = build_naming_plan(
            [
                Path(
                    "D:/project/产品数据/2017_ScreenBar/B站/"
                    "comments_BV1kt411h7i1_0ffa5872.csv"
                )
            ],
            today=datetime(2026, 7, 8, 9, 30),
        )

        self.assertEqual("2017_ScreenBar", plan.product_name)
        self.assertEqual("B站评论数据", plan.data_source)
        self.assertEqual("20260708_2017_ScreenBar_B站评论数据_合并总表.xlsx", plan.filenames["merge"])

    def test_does_not_treat_platform_exporter_names_as_product_names(self) -> None:
        tiktok_plan = build_naming_plan(
            [Path("D:/专案/ScreenBar十周年专案/产品数据/Tiktok/TTCommentExporter-7371053582810008864-51-comments-replies.csv")],
            today=datetime(2026, 7, 7, 9, 30),
        )

        self.assertEqual("TikTok评论数据", tiktok_plan.data_source)
        self.assertIn("product_name", tiktok_plan.missing_fields)
        self.assertEqual({}, tiktok_plan.filenames)

    def test_product_override_does_not_open_input_workbook(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-output-name-product-override"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "TTCommentExporter-7371053582810008864-51-comments-replies.xlsx"
        input_path.write_bytes(b"not an xlsx archive")

        plan = build_naming_plan(
            [input_path],
            product_name="ScreenBar Series",
            data_source="TikTok评论数据",
            today=datetime(2026, 7, 10, 9, 30),
        )

        self.assertEqual("ScreenBar Series", plan.product_name)
        self.assertEqual([], plan.product_candidates)
        self.assertEqual([], plan.missing_fields)

    def test_product_discovery_closes_the_input_workbook(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-output-name-closes-workbook"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"

        source = Workbook()
        source.active.append(["产品名", "评论内容"])
        source.active.append(["ScreenBar", "评论内容足够完整"])
        source.save(input_path)
        source.close()

        opened = load_workbook(input_path, read_only=True, data_only=True)
        opened.close = mock.Mock(wraps=opened.close)
        with mock.patch(
            "tools.output_file_naming.load_workbook_for_processing",
            return_value=opened,
        ):
            candidates = product_candidates_from_workbook(input_path)

        self.assertEqual(["ScreenBar"], candidates)
        opened.close.assert_called_once_with()



if __name__ == "__main__":
    unittest.main()
