from __future__ import annotations

import unittest
from pathlib import Path

from openpyxl import Workbook, load_workbook

from tools.merge_excel_workbooks import HeaderMismatchError, OutputPathConflictError, merge_workbooks


def write_workbook(path: Path, sheets: dict[str, list[list[object]]]) -> None:
    workbook = Workbook()
    first = True
    for sheet_name, rows in sheets.items():
        sheet = workbook.active if first else workbook.create_sheet(sheet_name)
        sheet.title = sheet_name
        first = False
        for row in rows:
            sheet.append(row)
    workbook.save(path)


class MergeExcelWorkbooksTest(unittest.TestCase):
    def test_merges_all_workbooks_with_one_header(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-merge"
        tmp.mkdir(parents=True, exist_ok=True)
        first_path = tmp / "b_second.xlsx"
        second_path = tmp / "a_first.xlsx"
        output_path = tmp / "merged.xlsx"

        header = ["province", "user", "comment", "time"]
        write_workbook(
            first_path,
            {
                "main": [
                    header,
                    ["Shanghai", "u2", "comment 2", "2026-01-02"],
                ],
                "subcomments": [
                    header,
                    ["Shanghai", "u2-sub", "sub comment", "2026-01-03"],
                ],
            },
        )
        write_workbook(
            second_path,
            {
                "main": [
                    header,
                    ["Beijing", "u1", "comment 1", "2026-01-01"],
                    [None, None, None, None],
                ],
            },
        )

        result = merge_workbooks([first_path, second_path], output_path)

        merged = load_workbook(result.output_path, read_only=True, data_only=True)
        sheet = merged["总表"]
        rows = list(sheet.iter_rows(values_only=True))

        self.assertEqual(tuple(header), rows[0])
        self.assertEqual(
            [
                ("Shanghai", "u2", "comment 2", "2026-01-02"),
                ("Shanghai", "u2-sub", "sub comment", "2026-01-03"),
                ("Beijing", "u1", "comment 1", "2026-01-01"),
            ],
            rows[1:],
        )
        self.assertEqual(2, result.files_processed)
        self.assertEqual(3, result.data_rows_written)
        self.assertTrue(result.summary_path.exists())

    def test_rejects_header_mismatch_by_default(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-merge-header-mismatch"
        tmp.mkdir(parents=True, exist_ok=True)
        first_path = tmp / "first.xlsx"
        second_path = tmp / "second.xlsx"

        write_workbook(first_path, {"main": [["province", "user", "comment"], ["A", "u1", "c1"]]})
        write_workbook(second_path, {"main": [["province", "comment", "user"], ["B", "c2", "u2"]]})

        with self.assertRaises(HeaderMismatchError):
            merge_workbooks([first_path, second_path], tmp / "merged.xlsx")

    def test_rejects_output_path_that_would_overwrite_input_file(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-merge-output-conflict"
        tmp.mkdir(parents=True, exist_ok=True)
        first_path = tmp / "first.xlsx"
        second_path = tmp / "second.xlsx"

        write_workbook(first_path, {"main": [["province", "user", "comment"], ["A", "u1", "c1"]]})
        write_workbook(second_path, {"main": [["province", "user", "comment"], ["B", "u2", "c2"]]})
        before = first_path.read_bytes()

        with self.assertRaises(OutputPathConflictError):
            merge_workbooks([first_path, second_path], first_path)

        self.assertEqual(before, first_path.read_bytes())

    def test_merges_explicit_csv_inputs_through_compatibility_layer(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-merge-csv-inputs"
        tmp.mkdir(parents=True, exist_ok=True)
        first_path = tmp / "first.csv"
        second_path = tmp / "second.csv"
        output_path = tmp / "merged.xlsx"

        first_path.write_text("timestamp,content,like_count\n1678870952,评论 A,2\n", encoding="utf-8-sig")
        second_path.write_text("timestamp,content,like_count\n1678870960,评论 B,3\n", encoding="utf-8-sig")

        result = merge_workbooks([first_path, second_path], output_path)

        merged = load_workbook(result.output_path, read_only=True, data_only=True)
        rows = list(merged["总表"].iter_rows(values_only=True))

        self.assertEqual(("timestamp", "content", "like_count"), rows[0])
        self.assertEqual(
            [
                ("1678870952", "评论 A", "2"),
                ("1678870960", "评论 B", "3"),
            ],
            rows[1:],
        )
        self.assertEqual(2, result.files_processed)
        self.assertEqual(2, result.data_rows_written)

    def test_preserves_formula_cells_instead_of_turning_them_into_blanks(self) -> None:
        tmp = Path.cwd() / ".tmp-tests" / "case-merge-preserves-formulas"
        tmp.mkdir(parents=True, exist_ok=True)
        input_path = tmp / "source.xlsx"
        output_path = tmp / "merged.xlsx"

        write_workbook(
            input_path,
            {"main": [["评论日期", "评论内容", "点赞数"], ["2026-07-08", "评论内容足够完整", "=1+1"]]},
        )

        result = merge_workbooks([input_path], output_path)

        merged = load_workbook(result.output_path, read_only=True, data_only=False)
        self.assertEqual("=1+1", merged["总表"].cell(row=2, column=3).value)



if __name__ == "__main__":
    unittest.main()
